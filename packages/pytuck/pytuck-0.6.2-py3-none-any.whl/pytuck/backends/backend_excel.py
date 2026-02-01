"""
Pytuck Excel存储引擎

使用单个Excel工作簿（.xlsx），每个表一个工作表，可视化友好
"""

import json
import base64
from pathlib import Path
from typing import Any, Dict, List, Union, TYPE_CHECKING, Tuple, Optional
from datetime import datetime
from .base import StorageBackend
from ..common.exceptions import SerializationError
from .versions import get_format_version
from ..core.types import TypeRegistry

from ..common.options import ExcelBackendOptions

if TYPE_CHECKING:
    from ..core.storage import Table
    from openpyxl import Workbook


class ExcelBackend(StorageBackend):
    """Excel format storage engine (requires openpyxl)"""

    ENGINE_NAME = 'excel'
    REQUIRED_DEPENDENCIES = ['openpyxl']
    FORMAT_VERSION = get_format_version('excel')

    def __init__(self, file_path: Union[str, Path], options: ExcelBackendOptions):
        """
        初始化 Excel 后端

        Args:
            file_path: Excel 文件路径
            options: Excel 后端配置选项
        """
        assert isinstance(options, ExcelBackendOptions), "options must be an instance of ExcelBackendOptions"
        super().__init__(file_path, options)
        # 类型安全：将 options 转为具体的 ExcelBackendOptions 类型
        self.options: ExcelBackendOptions = options

    def save(self, tables: Dict[str, 'Table']) -> None:
        """保存所有表数据到Excel工作簿"""
        if self.options.read_only:
            raise SerializationError("Excel backend does not support read-only mode")
        try:
            from openpyxl import Workbook
        except ImportError:
            raise SerializationError("openpyxl is required for Excel backend. Install with: pip install pytuck[excel]")

        temp_path = self.file_path.parent / (self.file_path.name + '.tmp')
        try:
            wb = Workbook()
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 创建元数据工作表
            metadata_sheet = wb.create_sheet('_metadata', 0)
            metadata_sheet.append(['Key', 'Value'])
            metadata_sheet.append(['format_version', self.FORMAT_VERSION])
            metadata_sheet.append(['timestamp', datetime.now().isoformat()])
            metadata_sheet.append(['table_count', len(tables)])

            # 创建统一的表结构工作表 _pytuck_tables
            tables_sheet = wb.create_sheet('_pytuck_tables', 1)
            tables_sheet.append(['table_name', 'primary_key', 'next_id', 'comment', 'columns'])
            for table_name, table in tables.items():
                columns_json = json.dumps([
                    {
                        'name': col.name,
                        'type': col.col_type.__name__,
                        'nullable': col.nullable,
                        'primary_key': col.primary_key,
                        'index': col.index,
                        'comment': col.comment
                    }
                    for col in table.columns.values()
                ])
                tables_sheet.append([table_name, table.primary_key, table.next_id, table.comment or '', columns_json])

            # 根据配置隐藏元数据工作表
            if self.options.hide_metadata_sheets:
                metadata_sheet.sheet_state = 'hidden'
                tables_sheet.sheet_state = 'hidden'

            # 为每个表创建数据工作表
            for table_name, table in tables.items():
                self._save_table_to_workbook(wb, table_name, table)

            # 原子性保存
            wb.save(str(temp_path))

            if self.file_path.exists():
                self.file_path.unlink()
            temp_path.replace(self.file_path)

        except Exception as e:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except FileNotFoundError:
                    pass
            raise SerializationError(f"Failed to save Excel file: {e}")

    def load(self) -> Dict[str, 'Table']:
        """从Excel工作簿加载所有表数据"""
        if not self.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SerializationError("openpyxl is required for Excel backend. Install with: pip install pytuck[excel]")

        try:
            wb = load_workbook(
                filename=str(self.file_path), read_only=self.options.read_only, data_only=True, keep_links=False
            )

            # 从 _pytuck_tables 工作表读取所有表的 schema
            tables_schema: Dict[str, Dict[str, Any]] = {}
            if '_pytuck_tables' in wb.sheetnames:
                tables_sheet = wb['_pytuck_tables']
                rows = list(tables_sheet.iter_rows(min_row=2, values_only=True))
                for row in rows:
                    if row[0]:  # table_name 不为空
                        table_name = row[0]
                        tables_schema[table_name] = {
                            'primary_key': row[1],
                            'next_id': int(row[2]) if row[2] else 1,
                            'comment': row[3] if row[3] else None,
                            'columns': json.loads(row[4]) if row[4] else []
                        }

            # 获取所有数据表名（排除元数据表）
            table_names = [
                name for name in wb.sheetnames
                if not name.startswith('_')
            ]

            tables = {}
            for table_name in table_names:
                schema = tables_schema.get(table_name, {})
                table = self._load_table_from_workbook(wb, table_name, schema)
                tables[table_name] = table

            return tables

        except Exception as e:
            raise SerializationError(f"Failed to load Excel file: {e}")

    def exists(self) -> bool:
        """检查文件是否存在"""
        return self.file_path.exists()

    def delete(self) -> None:
        """删除文件"""
        if self.exists():
            self.file_path.unlink()

    def _save_table_to_workbook(self, wb: 'Workbook', table_name: str, table: 'Table') -> None:
        """保存单个表的数据到工作簿"""
        # 数据工作表
        data_sheet = wb.create_sheet(table_name)

        # 写入表头
        columns = list(table.columns.keys())
        # 确保有主键时主键列在列表中
        if table.primary_key and table.primary_key not in columns:
            columns.insert(0, table.primary_key)

        data_sheet.append(columns)

        # 写入数据行
        for record in table.data.values():
            row: List[Any] = []
            for col_name in columns:
                value = record.get(col_name)
                column = table.columns.get(col_name)

                if value is None:
                    row.append('')
                elif column and column.col_type == bool:
                    # Excel 特殊处理：bool 转字符串 'TRUE'/'FALSE'
                    row.append('TRUE' if value else 'FALSE')
                elif column:
                    # 使用 TypeRegistry 统一序列化
                    row.append(TypeRegistry.serialize_for_text(value, column.col_type))
                else:
                    row.append(value)
            data_sheet.append(row)

    def _load_table_from_workbook(
        self, wb: 'Workbook', table_name: str, schema: Dict[str, Any]
    ) -> 'Table':
        """从工作簿加载单个表"""
        from ..core.storage import Table
        from ..core.orm import Column
        from datetime import datetime, date, timedelta

        primary_key = schema.get('primary_key')  # 可能为 None（无主键表）
        next_id = schema.get('next_id', 1)
        table_comment = schema.get('comment')
        columns_data = schema.get('columns', [])

        # 重建列
        columns = []

        if columns_data:
            # 有 schema（Pytuck 创建的文件）
            for col_data in columns_data:
                col_type = TypeRegistry.get_type_by_name(col_data['type'])
                column = Column(
                    col_type,
                    name=col_data['name'],
                    nullable=col_data['nullable'],
                    primary_key=col_data['primary_key'],
                    index=col_data.get('index', False),
                    comment=col_data.get('comment')
                )
                columns.append(column)
        else:
            # 无 schema（外部 Excel），从 headers 构建列（无主键）
            data_sheet = wb[table_name]
            rows_preview = list(data_sheet.iter_rows(values_only=True, max_row=1))
            if rows_preview:
                headers = [h for h in rows_preview[0] if h]
                for name in headers:
                    columns.append(Column(str, name=name, nullable=True, primary_key=False))
            # 外部 Excel 不添加主键列，使用无主键模式

        # 创建表（primary_key 可能为 None）
        table = Table(table_name, columns, primary_key, comment=table_comment)
        table.next_id = next_id

        # 读取数据
        data_sheet = wb[table_name]
        rows = list(data_sheet.iter_rows(values_only=True))

        max_int_pk = 0  # 用于更新 next_id

        if len(rows) > 1:
            headers = rows[0]
            for row_data in rows[1:]:
                record: Dict[str, Any] = {}
                for col_name, value in zip(headers, row_data):
                    if col_name not in table.columns:
                        continue

                    column = table.columns[col_name]

                    # 处理空值
                    if value == '' or value is None:
                        value = None
                    elif column.col_type == bool:
                        # Excel 的 bool 特殊处理
                        if isinstance(value, bool):
                            pass  # 保持原样
                        elif isinstance(value, str):
                            value = (value.upper() == 'TRUE')
                        else:
                            value = bool(value)
                    elif column.col_type == bytes:
                        # bytes 需要特殊处理（base64 解码）
                        if value:
                            value = base64.b64decode(value)
                    elif column.col_type in (datetime, date, timedelta, list, dict, int, float):
                        # 使用 TypeRegistry 统一反序列化
                        value = TypeRegistry.deserialize_from_text(value, column.col_type)

                    record[col_name] = value

                # 确定主键/rowid
                pk: Any
                if primary_key:
                    # 有主键：从记录中获取
                    pk = record.get(primary_key)
                    if pk is None:
                        # 主键为空，使用自增
                        pk = table.next_id
                        table.next_id += 1
                        record[primary_key] = pk
                else:
                    # 无主键：使用内部自增 rowid
                    pk = table.next_id
                    table.next_id += 1

                table.data[pk] = record

                # 跟踪最大 int pk
                if isinstance(pk, int) and pk > max_int_pk:
                    max_int_pk = pk

        # 更新 next_id（如果主键是 int 类型）
        if max_int_pk >= table.next_id:
            table.next_id = max_int_pk + 1

        # 重建索引
        for col_name, column in table.columns.items():
            if column.index:
                if col_name in table.indexes:
                    del table.indexes[col_name]
                table.build_index(col_name)

        return table

    def get_metadata(self) -> Dict[str, Any]:
        """获取元数据"""
        if not self.exists():
            return {}

        try:
            file_stat = self.file_path.stat()
            file_size = file_stat.st_size
            modified_time = file_stat.st_mtime

            from openpyxl import load_workbook
            wb = load_workbook(str(self.file_path), read_only=True)

            metadata = {
                'engine': 'excel',
                'file_size': file_size,
                'modified': modified_time
            }

            # 尝试读取元数据工作表
            if '_metadata' in wb.sheetnames:
                sheet = wb['_metadata']
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        metadata[row[0]] = row[1]

            wb.close()
            return metadata

        except:
            return {}

    @classmethod
    def probe(cls, file_path: Union[str, Path]) -> Tuple[bool, Optional[Dict[str, Any]]]:
        """
        轻量探测文件是否为 Excel 引擎格式

        通过检查 Excel 文件（实际上是 ZIP）是否包含 _pytuck_tables 工作表来识别。
        使用 ZIP 方式而非 openpyxl 以避免依赖问题和提高性能。

        Returns:
            Tuple[bool, Optional[Dict]]: (是否匹配, 元数据信息或None)
        """
        try:
            file_path = Path(file_path).expanduser()
            if not file_path.exists():
                return False, {'error': 'file_not_found'}

            # 获取文件信息
            file_stat = file_path.stat()
            file_size = file_stat.st_size

            # 空文件不可能是有效的 Excel
            if file_size == 0:
                return False, {'error': 'empty_file'}

            # Excel 文件实际上是 ZIP 格式，先检查是否为 ZIP
            import zipfile
            if not zipfile.is_zipfile(file_path):
                return False, None

            try:
                with zipfile.ZipFile(str(file_path), 'r') as zf:
                    namelist = zf.namelist()

                    # 检查是否为 Excel 文件结构
                    if 'xl/workbook.xml' not in namelist:
                        return False, None

                    # 检查是否包含 _pytuck_tables 工作表的 XML 文件
                    # Excel 工作表在 ZIP 中存储为 xl/worksheets/sheetN.xml
                    # 工作表名称映射在 xl/workbook.xml 中
                    pytuck_tables_found = False

                    try:
                        # 读取 workbook.xml 来查找工作表名称
                        with zf.open('xl/workbook.xml') as f:
                            workbook_xml = f.read(8192).decode('utf-8')  # 只读前 8KB

                        # 简单的字符串检查，查找 _pytuck_tables 工作表
                        if '_pytuck_tables' in workbook_xml:
                            pytuck_tables_found = True

                    except (KeyError, UnicodeDecodeError):
                        pass

                    if not pytuck_tables_found:
                        return False, None

                    # 尝试获取更多元数据信息
                    format_version = None
                    table_count = None
                    timestamp = None

                    # 如果有 openpyxl 依赖，尝试获取更详细信息
                    if cls.is_available():
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook(str(file_path), read_only=True, data_only=True)

                            # 从 _metadata 工作表读取信息
                            if '_metadata' in wb.sheetnames:
                                metadata_sheet = wb['_metadata']
                                for row in metadata_sheet.iter_rows(min_row=2, values_only=True):
                                    if row[0] and row[1]:
                                        if row[0] == 'format_version':
                                            format_version = row[1]
                                        elif row[0] == 'timestamp':
                                            timestamp = row[1]
                                        elif row[0] == 'table_count':
                                            table_count = row[1]

                            wb.close()
                        except Exception:
                            # 如果 openpyxl 读取失败，仍然可以确认是 Pytuck Excel 格式
                            pass

                    # 成功识别为 Excel 格式
                    return True, {
                        'engine': 'excel',
                        'format_version': format_version,
                        'table_count': table_count,
                        'file_size': file_size,
                        'modified': file_stat.st_mtime,
                        'timestamp': timestamp,
                        'confidence': 'high' if cls.is_available() else 'medium'
                    }

            except zipfile.BadZipFile:
                return False, {'error': 'corrupted_excel_file'}

        except Exception as e:
            return False, {'error': f'probe_exception: {str(e)}'}

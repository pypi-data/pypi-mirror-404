"""
外部文件加载工具

将非 Pytuck 格式的 CSV/Excel 文件加载为模型对象列表
"""

from pathlib import Path
from typing import Any, Dict, List, Optional, Type, TypeVar

from ..core.orm import PureBaseModel, Column


T = TypeVar('T', bound=PureBaseModel)


def load_table(
    model: Type[T],
    file_path: str,
    *,
    sheet_name: Optional[str] = None,
    encoding: str = 'utf-8-sig',
    delimiter: str = ',',
) -> List[T]:
    """
    从外部 CSV/Excel 文件加载数据到模型对象列表

    Args:
        model: 模型类（必须继承自 PureBaseModel）
        file_path: 文件路径（.csv 或 .xlsx）
        sheet_name: Excel 工作表名（仅 Excel 有效）
        encoding: CSV 文件编码
        delimiter: CSV 分隔符

    Returns:
        List[T]: 模型对象列表

    Raises:
        FileNotFoundError: 文件不存在
        ValueError: 不支持的文件类型
        ValidationError: 类型转换失败

    Example:
        >>> from pytuck import Storage, declarative_base, Column
        >>> from pytuck.tools import load_table
        >>>
        >>> db = Storage(in_memory=True)
        >>> Base = declarative_base(db)
        >>>
        >>> class User(Base):
        ...     __tablename__ = 'users'
        ...     id = Column(int, primary_key=True)
        ...     name = Column(str)
        ...     age = Column(int)
        >>>
        >>> users = load_table(User, 'users.csv')
        >>> for user in users:
        ...     print(user.id, user.name, user.age)
    """
    path = Path(file_path).expanduser()

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    suffix = path.suffix.lower()

    # 根据文件类型加载原始数据
    if suffix == '.csv':
        rows = _load_csv(path, encoding, delimiter)
    elif suffix in ('.xls', '.xlsx'):
        rows = _load_excel(path, sheet_name)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    # 获取模型的列定义
    columns: Dict[str, Column] = model.__columns__

    # 转换每行数据为模型对象
    result: List[T] = []
    for row_idx, row in enumerate(rows):
        try:
            obj = _row_to_model(model, row, columns)
            result.append(obj)
        except Exception as e:
            raise ValueError(f"Error at row {row_idx + 2}: {e}") from e

    return result


def _load_csv(path: Path, encoding: str, delimiter: str) -> List[Dict[str, Any]]:
    """加载 CSV 文件，返回字典列表"""
    import csv

    rows: List[Dict[str, Any]] = []
    with open(path, 'r', encoding=encoding, newline='') as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            rows.append(dict(row))

    return rows


def _load_excel(path: Path, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    """加载 Excel 文件，返回字典列表"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel files. "
            "Install with: pip install pytuck[excel]"
        )

    wb = load_workbook(str(path), read_only=True, data_only=True)

    # 选择工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
        ws = wb[sheet_name]
    else:
        # 取第一个工作表
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # 读取 header
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return []

    headers = [str(h) if h is not None else '' for h in header_row]

    # 读取数据行
    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        row_dict: Dict[str, Any] = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = value
        rows.append(row_dict)

    wb.close()
    return rows


def _row_to_model(
    model: Type[T],
    row: Dict[str, Any],
    columns: Dict[str, Column]
) -> T:
    """将一行数据转换为模型对象"""
    kwargs: Dict[str, Any] = {}

    for attr_name, column in columns.items():
        # 优先使用 Column.name 查找（匹配 CSV/Excel 表头）
        # 如果没有指定 name，则使用属性名
        col_name = column.name if column.name else attr_name
        raw_value = row.get(col_name)

        # 如果 Column.name 找不到，尝试用属性名（兼容旧行为）
        if raw_value is None and column.name:
            raw_value = row.get(attr_name)

        # 空字符串视为 None
        if raw_value == '':
            raw_value = None

        # 使用 Column.validate 进行类型转换
        # 能转就转，不能转就抛出 ValidationError
        validated_value = column.validate(raw_value)
        kwargs[attr_name] = validated_value  # 用属性名作为 kwargs 键

    # 创建模型实例
    # 注意：PureBaseModel 的 __init__ 由 declarative_base 动态生成
    obj = model(**kwargs)

    return obj

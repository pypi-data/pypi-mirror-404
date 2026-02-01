"""
tactus.io.excel - Excel spreadsheet operations for Tactus.

Provides Excel read/write operations with sandboxing
to the procedure's base directory.

Requires: openpyxl

Usage in .tac files:
    local excel = require("tactus.io.excel")

    -- Read Excel file
    local data = excel.read("data.xlsx")

    -- Read specific sheet
    local data = excel.read("data.xlsx", {sheet = "Sheet2"})

    -- Write Excel file
    excel.write("output.xlsx", {
        {name = "Alice", score = 95},
        {name = "Bob", score = 87}
    })

    -- List sheet names
    local sheets = excel.sheets("data.xlsx")
"""

import os
import sys
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")

# Get context (injected by loader)
_ctx = getattr(sys.modules[__name__], "__tactus_context__", None)


def read(filepath: str, options: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """
    Read Excel file, returning list of dictionaries.

    Args:
        filepath: Path to Excel file (relative to working directory)
        options: Optional dict with 'sheet' key for sheet name

    Returns:
        List of dictionaries, one per row (excluding header row)

    Raises:
        FileNotFoundError: If file does not exist
        PermissionError: If path is outside working directory
    """
    if _ctx:
        filepath = _ctx.validate_path(filepath)

    options = options or {}
    sheet = options.get("sheet") if options else None

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # First row is headers
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def write(
    filepath: str, data: List[Dict[str, Any]], options: Optional[Dict[str, Any]] = None
) -> None:
    """
    Write list of dictionaries to Excel file.

    Args:
        filepath: Path to Excel file
        data: List of dictionaries to write
        options: Optional dict with 'sheet' key for sheet name (default "Sheet1")

    Raises:
        PermissionError: If path is outside working directory
        ValueError: If data is empty
    """
    if _ctx:
        filepath = _ctx.validate_path(filepath)

    if not data:
        raise ValueError("Cannot write empty data to Excel")

    options = options or {}
    sheet = options.get("sheet", "Sheet1")

    # Create parent directories
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    # Write headers and data
    headers = list(data[0].keys())
    ws.append(headers)
    for row in data:
        ws.append([row.get(h) for h in headers])

    wb.save(filepath)


def sheets(filepath: str) -> List[str]:
    """
    List sheet names in Excel file.

    Args:
        filepath: Path to Excel file

    Returns:
        List of sheet names

    Raises:
        FileNotFoundError: If file does not exist
        PermissionError: If path is outside working directory
    """
    if _ctx:
        filepath = _ctx.validate_path(filepath)

    wb = load_workbook(filepath, read_only=True)
    return wb.sheetnames


# Explicit exports
__tactus_exports__ = ["read", "write", "sheets"]

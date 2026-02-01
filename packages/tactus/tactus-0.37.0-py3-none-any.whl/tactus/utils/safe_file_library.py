"""Safe file I/O libraries for Lua sandbox.

Provides sandboxed file operations restricted to the working directory.
All paths are validated to prevent directory traversal attacks.

Supported formats:
- File: Raw text read/write
- Csv: CSV with automatic header handling
- Tsv: Tab-separated values
- Json: JSON read/write
- Parquet: Apache Parquet (via pyarrow)
- Hdf5: HDF5 datasets (via h5py)
- Excel: Excel spreadsheets (via openpyxl)
"""

import csv
import json
import logging
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class PathValidator:
    """Validates file paths are within allowed base directory."""

    def __init__(self, base_path: str):
        """
        Initialize path validator.

        Args:
            base_path: The base directory that all file operations are restricted to.
        """
        self.base_path = os.path.realpath(base_path)

    def validate(self, file_path: str) -> str:
        """
        Validate and resolve a file path.

        Args:
            file_path: Relative or absolute file path to validate.

        Returns:
            Resolved absolute path if valid.

        Raises:
            PermissionError: If path is outside the base directory.
        """
        # Join with base path and resolve to absolute
        resolved_path = os.path.realpath(os.path.join(self.base_path, file_path))

        # Check if resolved path is within base directory
        # Allow exact match (base_path itself) or paths that start with base_path + separator
        if resolved_path != self.base_path and not resolved_path.startswith(
            self.base_path + os.sep
        ):
            raise PermissionError(f"Access denied: path outside working directory: {file_path}")

        return resolved_path


def create_safe_file_library(base_path: str) -> Dict[str, Any]:
    """
    Create raw text file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of file operation functions.
    """
    validator = PathValidator(base_path)

    def read(file_path: str) -> str:
        """Read entire file as text."""
        validated_path = validator.validate(file_path)
        with open(validated_path, "r", encoding="utf-8") as file_handle:
            return file_handle.read()

    def write(file_path: str, content: str) -> None:
        """Write text to file."""
        validated_path = validator.validate(file_path)
        # Ensure parent directory exists
        parent_directory = os.path.dirname(validated_path)
        if parent_directory:
            os.makedirs(parent_directory, exist_ok=True)
        with open(validated_path, "w", encoding="utf-8") as file_handle:
            file_handle.write(content)

    def exists(file_path: str) -> bool:
        """Check if file exists."""
        try:
            validated_path = validator.validate(file_path)
            return os.path.exists(validated_path)
        except PermissionError:
            return False

    return {"read": read, "write": write, "exists": exists}


class LuaList:
    """
    Wrapper for Python lists that works better with Lua via lupa.

    Provides both 0-indexed access (Python style) and a len() method
    that can be called from Lua.
    """

    def __init__(self, data: List):
        self._data = data

    def __getitem__(self, key_or_index):
        # Lua method access comes through __getitem__ with string keys
        if isinstance(key_or_index, str):
            # Handle method access
            if key_or_index == "len":
                return self.len
            elif key_or_index == "get":
                return self.get
            else:
                raise KeyError(f"Unknown method: {key_or_index}")
        # Lua numbers are floats, convert to int for indexing
        if isinstance(key_or_index, float):
            key_or_index = int(key_or_index)
        return self._data[key_or_index]

    def __len__(self):
        return len(self._data)

    def __iter__(self):
        return iter(self._data)

    def len(self):
        """Return length - callable from Lua as data:len()."""
        return len(self._data)

    def get(self, requested_index):
        """Alternative access method - data:get(0) instead of data[0]."""
        if isinstance(requested_index, float):
            requested_index = int(requested_index)
        return self._data[requested_index]


def create_safe_csv_library(base_path: str) -> Dict[str, Any]:
    """
    Create CSV file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of CSV operation functions.
    """
    validator = PathValidator(base_path)

    def read(file_path: str) -> LuaList:
        """Read CSV file, returning list of dictionaries with headers as keys."""
        validated_path = validator.validate(file_path)
        with open(validated_path, "r", encoding="utf-8", newline="") as file_handle:
            reader = csv.DictReader(file_handle)
            return LuaList(list(reader))

    def write(file_path: str, data: List[Dict], options: Optional[Dict] = None) -> None:
        """Write list of dictionaries to CSV file."""
        validated_path = validator.validate(file_path)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        header_row = options.get("headers")

        # Convert headers from Lua table to Python list if needed
        if header_row and hasattr(header_row, "values"):
            header_row = list(header_row.values())

        # Convert Lua table to Python list if needed
        if hasattr(data, "values"):
            data = list(data.values())

        if not header_row and data:
            # Get headers from first row
            first_row = data[0]
            if hasattr(first_row, "keys"):
                header_row = list(first_row.keys())
            elif isinstance(first_row, dict):
                header_row = list(first_row.keys())
            else:
                raise ValueError("Cannot determine headers from data")

        # Ensure parent directory exists
        parent_directory = os.path.dirname(validated_path)
        if parent_directory:
            os.makedirs(parent_directory, exist_ok=True)

        with open(validated_path, "w", encoding="utf-8", newline="") as file_handle:
            writer = csv.DictWriter(file_handle, fieldnames=header_row)
            writer.writeheader()
            for row in data:
                # Convert Lua table to dict if needed
                if hasattr(row, "items"):
                    row = dict(row.items())
                writer.writerow(row)

    return {"read": read, "write": write}


def create_safe_tsv_library(base_path: str) -> Dict[str, Any]:
    """
    Create TSV (tab-separated values) file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of TSV operation functions.
    """
    validator = PathValidator(base_path)

    def read(file_path: str) -> LuaList:
        """Read TSV file, returning list of dictionaries with headers as keys."""
        validated_path = validator.validate(file_path)
        with open(validated_path, "r", encoding="utf-8", newline="") as file_handle:
            reader = csv.DictReader(file_handle, delimiter="\t")
            return LuaList(list(reader))

    def write(file_path: str, data: List[Dict], options: Optional[Dict] = None) -> None:
        """Write list of dictionaries to TSV file."""
        validated_path = validator.validate(file_path)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        header_row = options.get("headers")

        # Convert headers from Lua table to Python list if needed
        if header_row and hasattr(header_row, "values"):
            header_row = list(header_row.values())

        # Convert Lua table to Python list if needed
        if hasattr(data, "values"):
            data = list(data.values())

        if not header_row and data:
            first_row = data[0]
            if hasattr(first_row, "keys"):
                header_row = list(first_row.keys())
            elif isinstance(first_row, dict):
                header_row = list(first_row.keys())
            else:
                raise ValueError("Cannot determine headers from data")

        # Ensure parent directory exists
        parent_directory = os.path.dirname(validated_path)
        if parent_directory:
            os.makedirs(parent_directory, exist_ok=True)

        with open(validated_path, "w", encoding="utf-8", newline="") as file_handle:
            writer = csv.DictWriter(
                file_handle,
                fieldnames=header_row,
                delimiter="\t",
            )
            writer.writeheader()
            for row in data:
                if hasattr(row, "items"):
                    row = dict(row.items())
                writer.writerow(row)

    return {"read": read, "write": write}


def create_safe_json_library(base_path: str) -> Dict[str, Any]:
    """
    Create JSON file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of JSON operation functions.
    """
    validator = PathValidator(base_path)

    def read(file_path: str) -> Any:
        """Read JSON file and return parsed data."""
        validated_path = validator.validate(file_path)
        with open(validated_path, "r", encoding="utf-8") as file_handle:
            return json.load(file_handle)

    def write(file_path: str, data: Any, options: Optional[Dict] = None) -> None:
        """Write data to JSON file."""
        validated_path = validator.validate(file_path)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        indent = options.get("indent", 2)

        # Convert Lua tables to Python dicts/lists
        data = _lua_to_python(data)

        # Ensure parent directory exists
        parent_directory = os.path.dirname(validated_path)
        if parent_directory:
            os.makedirs(parent_directory, exist_ok=True)

        with open(validated_path, "w", encoding="utf-8") as file_handle:
            json.dump(data, file_handle, indent=indent, default=str)

    return {"read": read, "write": write}


def create_safe_parquet_library(base_path: str) -> Dict[str, Any]:
    """
    Create Parquet file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of Parquet operation functions.
    """
    import pyarrow as pa
    import pyarrow.parquet as pq

    validator = PathValidator(base_path)

    def read(file_path: str) -> LuaList:
        """Read Parquet file, returning list of dictionaries."""
        validated_path = validator.validate(file_path)
        table = pq.read_table(validated_path)
        return LuaList(table.to_pylist())

    def write(file_path: str, data: List[Dict]) -> None:
        """Write list of dictionaries to Parquet file."""
        validated_path = validator.validate(file_path)

        # Convert Lua tables to Python
        data = _lua_to_python(data)

        # Ensure parent directory exists
        parent_directory = os.path.dirname(validated_path)
        if parent_directory:
            os.makedirs(parent_directory, exist_ok=True)

        table = pa.Table.from_pylist(data)
        pq.write_table(table, validated_path)

    return {"read": read, "write": write}


def create_safe_hdf5_library(base_path: str) -> Dict[str, Any]:
    """
    Create HDF5 file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of HDF5 operation functions.
    """
    import h5py
    import numpy as np

    validator = PathValidator(base_path)

    def read(file_path: str, dataset: str) -> List[Any]:
        """Read dataset from HDF5 file."""
        validated_path = validator.validate(file_path)
        with h5py.File(validated_path, "r") as hdf5_file:
            return hdf5_file[dataset][:].tolist()

    def write(file_path: str, dataset: str, data: List) -> None:
        """Write data to HDF5 dataset."""
        validated_path = validator.validate(file_path)

        # Convert Lua tables to Python
        data = _lua_to_python(data)

        # Ensure parent directory exists
        parent_directory = os.path.dirname(validated_path)
        if parent_directory:
            os.makedirs(parent_directory, exist_ok=True)

        with h5py.File(validated_path, "a") as hdf5_file:
            if dataset in hdf5_file:
                del hdf5_file[dataset]
            hdf5_file.create_dataset(dataset, data=np.array(data))

    def list_datasets(file_path: str) -> List[str]:
        """List all datasets in HDF5 file."""
        validated_path = validator.validate(file_path)
        datasets = []
        with h5py.File(validated_path, "r") as hdf5_file:

            def visitor(name: str, hdf5_object: Any) -> None:
                if isinstance(hdf5_object, h5py.Dataset):
                    datasets.append(name)

            hdf5_file.visititems(visitor)
        return datasets

    return {"read": read, "write": write, "list": list_datasets}


def create_safe_excel_library(base_path: str) -> Dict[str, Any]:
    """
    Create Excel file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of Excel operation functions.
    """
    from openpyxl import Workbook, load_workbook

    validator = PathValidator(base_path)

    def read(file_path: str, options: Optional[Dict] = None) -> LuaList:
        """Read Excel file, returning list of dictionaries."""
        validated_path = validator.validate(file_path)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        sheet_name = options.get("sheet")

        workbook = load_workbook(validated_path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return LuaList([])

        # First row is headers
        header_row = [
            str(header_value) if header_value is not None else f"col_{index}"
            for index, header_value in enumerate(rows[0])
        ]
        data_rows = rows[1:]
        return LuaList([dict(zip(header_row, row_values)) for row_values in data_rows])

    def write(file_path: str, data: List[Dict], options: Optional[Dict] = None) -> None:
        """Write list of dictionaries to Excel file."""
        validated_path = validator.validate(file_path)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        sheet_name = options.get("sheet", "Sheet1")

        # Convert Lua tables to Python
        data = _lua_to_python(data)

        # Ensure parent directory exists
        parent_directory = os.path.dirname(validated_path)
        if parent_directory:
            os.makedirs(parent_directory, exist_ok=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        if data:
            header_row = list(data[0].keys())
            worksheet.append(header_row)
            for row in data:
                worksheet.append([row.get(header) for header in header_row])

        workbook.save(validated_path)

    def sheets(file_path: str) -> List[str]:
        """List sheet names in Excel file."""
        validated_path = validator.validate(file_path)
        workbook = load_workbook(validated_path, read_only=True)
        return workbook.sheetnames

    return {"read": read, "write": write, "sheets": sheets}


def _lua_to_python(value: Any) -> Any:
    """
    Recursively convert Lua table-like objects to Python dicts/lists.

    Args:
        value: Object to convert (may be Lua table or Python object).

    Returns:
        Python dict, list, or original value.
    """
    # Check if it's a Lua table (has values() or items() method)
    if hasattr(value, "items"):
        # Could be a dict-like Lua table
        try:
            lua_items = list(value.items())
            # Check if it's array-like (all integer keys starting from 1)
            if lua_items and all(isinstance(key, (int, float)) for key, _ in lua_items):
                lua_keys = [int(key) for key, _ in lua_items]
                if lua_keys == list(range(1, len(lua_keys) + 1)):
                    # It's an array-like table, convert to list
                    return [_lua_to_python(value[key]) for key in range(1, len(lua_keys) + 1)]
            # It's a dict-like table
            return {key: _lua_to_python(item) for key, item in lua_items}
        except (TypeError, AttributeError):
            pass

    if hasattr(value, "values") and not isinstance(value, (dict, str)):
        try:
            return [_lua_to_python(item) for item in value.values()]
        except (TypeError, AttributeError):
            pass

    if isinstance(value, dict):
        return {key: _lua_to_python(item) for key, item in value.items()}

    if isinstance(value, list):
        return [_lua_to_python(item) for item in value]

    return value

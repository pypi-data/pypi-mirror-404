import csv
import sys
import importlib
importlib.reload(sys)
from pdfminer.pdfparser import PDFParser, PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LTTextBoxHorizontal, LAParams
from pdfminer.pdfinterp import PDFTextExtractionNotAllowed
import win32com
import win32com.client
from openpyxl.reader.excel import load_workbook
from pyexcel_xls import get_data
from collections import OrderedDict  # 导入有序字典
from pyexcel_xls import save_data
import pickle

class RwFile(object):
    #读csv
    def readCsv(self, path):
        infoList = []
        with open(path, 'r') as f:
            allFileInfo = csv.reader(f)
            for row in allFileInfo:
                infoList.append(row)
        return infoList


    #写csv,data的格式为嵌套列表[['',''],['','']]
    def writeCsv(self, path, data):
        with open(path, 'a') as f: #a追加
            writer = csv.writer(f)
            for rowData in data:
                writer.writerow(rowData)


    #读pdf，内容写成一个文件
    def readPDF(self, path, callback=None, toPath=''):
        # 以二进制形式打开pdf文件
        with open(path, 'rb') as f:
            # 创建一个pdf文档分析器
            parser = PDFParser(f)
            # 创建一个pdf文档
            pdfFile = PDFDocument()
            # 将新建pdf与分析器进行链接
            parser.set_document(pdfFile)
            pdfFile.set_parser(parser)
            # 提供初始化密码
            pdfFile.initialize()
            # 检测文档是否提供Text转换
            if not pdfFile.is_extractable:  # 不能提供转换功能返回False，not之后为True，不在看此pdf
                raise PDFTextExtractionNotAllowed  # 结束
            else:
                # 解析数据
                # 创建数据管理器
                manager = PDFResourceManager()
                # 创建PDF设备对象
                laparams = LAParams()
                device = PDFPageAggregator(manager, laparams=laparams)
                # 创建解释器对象
                interpreter = PDFPageInterpreter(manager, device)
                # 开始循环处理，每次处理一页
                for page in pdfFile.get_pages():  # 获取pdf的页数
                    # 解释器处理一页内容
                    interpreter.process_page(page)
                    # 创建图层对象
                    layout = device.get_result()
                    for x in layout:
                        # isinstance判断x是否为LTTextBoxHorizontal类型
                        if (isinstance(x, LTTextBoxHorizontal)):
                            if toPath == '':
                            #处理数据,因为每次处理数据的方式不一样，又不想修改类中的代码，因此传入一个callback函数，在main中自己写，再传入类中
                                str_ = x.get_text()
                                if callback != None:
                                    callback(str_)
                                else:
                                    print(str_)

                            else:
                                print('将输入存入文件')
                                with open(toPath, 'a', encoding='utf-8') as f:
                                    # 读pdf的内容
                                    str_ = x.get_text()
                                    f.write(str_ + '\n')



    #读word
    def readWord(slef, path):
        word = []
        # 实例化系统中word的功能，即打开word软件，可以处理doc和docx两种文件
        mw = win32com.client.Dispatch('Word.Application')
        # 调用实例对象的Open方法打开指定word文件
        doc = mw.Documents.Open(path)
        for paragraph in doc.Paragraphs:
            line = paragraph.Range.Text
            word.append(line)
        # 关闭文件
        doc.Close()
        # 退出word
        mw.Quit()
        return word

    #读word另存为word
    def readWordToOther(self, path, toPath):
        mw = win32com.client.Dispatch('Word.Application')
        doc = mw.Documents.Open(path)
        # 将读取到的数据保存到(另一个文件中, 2=txt文件)
        doc.SaveAs(toPath, 2)
        doc.Close()
        mw.Quit()


    #创建word
    def makeWorkFile(self, path):
        word = win32com.client.Dispatch('Word.Application')
        word.Visible = True
        doc = word.Documents.Add()
        r = doc.Range(0, 0)
        r.InsertAfter('亲爱的%s\n')
        r.InsertAfter('    我想你。。。\n')
        doc.SaveAs(path)
        doc.Close()
        word.Quit()


    #读xlsx并返回字典
    def readXlsxFile(self, path):
        dict_ = {}
        file = load_workbook(filename=path)
        sheets = file.get_sheet_names()
        for sheetName in sheets:
            sheet = file.get_sheet_by_name(sheetName)
            sheetInfo = []
            for lineNum in range(1, sheet.max_row + 1):
                lineList = []
                for columnNum in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=lineNum, column=columnNum).value
                    lineList.append(value)
                sheetInfo.append(lineList)
            dict_[sheetName] = sheetInfo
        return dict_


    #读xls和xlsx文件，返回字典
    def readXlsAndXlsxFile(slef, path):
        dict_ = OrderedDict()
        xdata = get_data(path) #抓取所有数据
        for sheet in xdata:
            dict_[sheet] = xdata[sheet]
        return dict_


    #创建xls文件
    def makeXlsFile(slef, path, data):
        dic = OrderedDict()
        for sheetName, sheetValue in data.items():
            d = {}
            d[sheetName] = sheetValue
            dic.update(d)  # .update() 似 .append()
        save_data(path, dic)


    #写ppt
    def makePPT(self, path):
        ppt = win32com.client.Dispatch('PowerPoint.Application')
        ppt.Visible = True
        # 增加一个文件
        pptFile = ppt.Presentations.Add()
        # 新建幻灯片
        pag1 = pptFile.Slides.Add(1, 1)
        # 幻灯片的第一个文本输入框的输入文字的位置
        t1 = pag1.Shapes[0].TextFrame.TextRange
        # 写文字
        t1.Text = '陈艺龙'
        t2 = pag1.Shapes[1].TextFrame.TextRange
        # 写文字
        t2.Text = '彭淑贤'
        # 保存
        pptFile.SaveAs(path)
        # 关闭幻灯片
        pptFile.Close()
        # 关闭软件
        ppt.Quit()

    #列表形式的数据以循环方式一行一行写入txt
    def listToTxt(self, path, dataList):
        for data in dataList:
            with open(path, 'a') as f:
                f.write(data + '\n')


    #保存python数据类型
    def dumpPickle(self, path, data):
        with open(path, 'a') as f:
            pickle.dump(data, f)

    #读取python数据类型
    def loadPickle(self, path):
        with open(path, 'rb') as f:
            data = pickle.load(f)
        return data
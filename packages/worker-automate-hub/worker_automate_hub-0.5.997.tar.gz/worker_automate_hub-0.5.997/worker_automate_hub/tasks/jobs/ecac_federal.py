import asyncio
import io
import mimetypes
import os
import re
import shutil
import smtplib
import time
import uuid
import zipfile
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
import pandas as pd
import pyautogui
import timedelta
from bs4 import BeautifulSoup
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from playwright.async_api import async_playwright
from rich.console import Console
from xhtml2pdf import pisa

from worker_automate_hub.api.client import (
    get_config_by_name,
    sync_get_config_by_name,
)
from worker_automate_hub.models.dto.rpa_historico_request_dto import (
    RpaHistoricoStatusEnum,
    RpaRetornoProcessoDTO,
    RpaTagDTO,
    RpaTagEnum,
)
from worker_automate_hub.models.dto.rpa_processo_entrada_dto import (
    RpaProcessoEntradaDTO,
)
from worker_automate_hub.utils.get_creds_gworkspace import GetCredsGworkspace
from worker_automate_hub.utils.logger import logger
from worker_automate_hub.utils.util import (
    find_target_position,
    take_screenshot,
)

pyautogui.PAUSE = 0.5
pyautogui.FAILSAFE = False
console = Console()


def find_and_click(
    image_path: str,
    description: str,
    confidence: float = 0.8,
    write_text: str | None = None,
    wait_time: float = 2,
) -> None:
    """
    Encontra um elemento na tela com base em uma imagem e clica nele.

    Args:
        image_path (str): Caminho para a imagem do elemento.
        description (str): Descrição do elemento.
        confidence (float, optional): Nível de confiança para encontrar o elemento. Defaults to 0.8.
        write_text (str | None, optional): Texto a ser inserido no campo. Defaults to None.
        wait_time (float, optional): Tempo de espera após clicar no elemento. Defaults to 2.

    Raises:
        ValueError: Se o caminho da imagem for nulo ou vazio.
        FileNotFoundError: Se o arquivo da imagem não existir.
        Exception: Se o elemento não for encontrado.
    """
    if not image_path or not isinstance(image_path, str):
        raise ValueError("O caminho da imagem não pode ser nulo ou vazio")

    if not Path(image_path).exists():
        raise FileNotFoundError(f"O arquivo {image_path} não existe")

    element = pyautogui.locateCenterOnScreen(image_path, confidence=confidence)
    if element:
        pyautogui.click(element)
        console.print(f"[green]'{description}' encontrado e clicado.[/green]")
        if write_text:
            pyautogui.write(write_text, interval=0.05)
            console.print(
                f"[green]Texto '{write_text}' inserido em '{description}'.[/green]"
            )
        time.sleep(wait_time)
    else:
        raise Exception(f"[red]Elemento '{description}' não encontrado![/red]")


def extract_table(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    table = soup.find("table", {"id": "gridMensagens"})
    if not table:
        print("Tabela não encontrada")
        return []

    table_data = []
    rows = table.find_all("tr")

    header_row = rows[0]
    headers = [cell.get_text(strip=True) for cell in header_row.find_all(["td", "th"])]

    try:
        start_index = headers.index("Remetente")
    except ValueError:
        print("Coluna 'Remetente' não encontrada")
        return []

    for row in rows:
        row_data = []
        cells = row.find_all(["td", "th"])
        for i in range(start_index, len(cells)):
            cell = cells[i]
            link = cell.find("a")
            if link and "href" in link.attrs:
                link_text = link.get_text(strip=True)
                link_url = link["href"]
                row_data.append(link_text)
                row_data.append(link_url)
            else:
                row_data.append(cell.get_text(strip=True))
                row_data.append("")
        table_data.append(row_data)

    return table_data


def report_to_user(filename, report, path_directory):
    fullpath = os.path.join(path_directory, filename)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        "Status",
        "Destinatario",
        "Remetente",
        "Assunto",
        "Enviada em",
        "Aberta em",
        "AVISO",
    ]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num).value = header

    row_num = 2
    for page_data in report:
        for data in page_data:
            col_num = 1
            for key in headers:
                ws.cell(row=row_num, column=col_num).value = data.get(key, "")
                col_num += 1
            row_num += 1

    wb.save(fullpath)


def compactar_diretorio(
    diretorio: str,
    nome_arquivo_zip: str,
) -> None:
    """
    Compacta todos os arquivos em um diretório em um arquivo ZIP.

    Args:
        diretorio (str): O caminho do diretório a ser compactado.
        nome_arquivo_zip (str): O nome do arquivo ZIP a ser criado.

    Returns:
        None
    """
    if not diretorio or not isinstance(diretorio, str):
        raise ValueError("O caminho do diretório não pode ser nulo ou vazio")

    if not nome_arquivo_zip or not isinstance(nome_arquivo_zip, str):
        raise ValueError("O nome do arquivo ZIP não pode ser nulo ou vazio")

    arquivos = [
        os.path.join(diretorio, arquivo)
        for arquivo in os.listdir(diretorio)
        if os.path.isfile(os.path.join(diretorio, arquivo))
    ]

    if not arquivos:
        raise FileNotFoundError(
            f"O diretório {diretorio} não contém arquivos a serem compactados"
        )

    try:
        with zipfile.ZipFile(nome_arquivo_zip, "w") as zipf:
            for arquivo in arquivos:
                zipf.write(arquivo, os.path.basename(arquivo))
    except Exception as e:
        raise Exception(f"Erro ao compactar o diretório {diretorio}: {e}")


def send_email(
    smtp_server: str,
    smtp_port: int,
    smtp_user: str,
    smtp_password: str,
    message_text: str,
    subject: str,
    to: str,
    diretorio: str,
) -> None:
    """
    Envia um e-mail com os dados recebidos.

    Args:
        smtp_server (str): Servidor SMTP
        smtp_port (int): Porta do servidor SMTP
        smtp_user (str): Usuário do servidor SMTP
        smtp_password (str): Senha do usuário do servidor SMTP
        message_text (str): Corpo do e-mail
        subject (str): Assunto do e-mail
        to (str): Destinatário do e-mail
        diretorio (str): Caminho do diretório a ser compactado

    Returns:
        None
    """
    if not smtp_server or not smtp_port or not smtp_user or not smtp_password:
        raise ValueError("Parâmetros SMTP não podem ser nulos")

    if not message_text or not subject or not to or not diretorio:
        raise ValueError("Parâmetros de e-mail não podem ser nulos")

    nome_arquivo_zip = "Consulta ECAC.zip"
    full_path = os.path.join(diretorio, nome_arquivo_zip)
    compactar_diretorio(diretorio, full_path)

    message = create_message_with_attachment(
        smtp_user, to, subject, message_text, full_path
    )
    try:
        send_message(
            smtp_server, smtp_port, smtp_user, smtp_password, smtp_user, to, message
        )
    except Exception as e:
        err_msg = f"Erro ao enviar e-mail: {e}"
        console.print(err_msg, style="bold red")
        logger.error(err_msg)


def create_message_with_attachment(
    sender: str, to: str, subject: str, message_text: str, file: str | None = None
) -> str:
    """
    Cria uma mensagem com um anexo.

    Args:
        sender (str): Remetente do email
        to (str): Destinatário do email
        subject (str): Assunto do email
        message_text (str): Corpo do email
        file (str | None): Caminho do arquivo a ser anexado (opcional)

    Returns:
        str: A mensagem com o anexo no formato de string
    """
    if not sender:
        raise ValueError("Remetente do email não pode ser nulo")
    if not to:
        raise ValueError("Destinatário do email não pode ser nulo")
    if not subject:
        raise ValueError("Assunto do email não pode ser nulo")
    if not message_text:
        raise ValueError("Corpo do email não pode ser nulo")

    message = MIMEMultipart()
    message["to"] = to
    message["from"] = sender
    message["subject"] = subject

    msg = MIMEText(message_text)
    message.attach(msg)

    if file:
        if not os.path.exists(file):
            raise FileNotFoundError(f"Arquivo {file} não encontrado")
        if not os.path.isfile(file):
            raise ValueError(f"{file} não é um arquivo")

        content_type, encoding = mimetypes.guess_type(file)

        if content_type is None or encoding is not None:
            content_type = "application/octet-stream"

        main_type, sub_type = content_type.split("/", 1)
        with open(file, "rb") as f:
            attachment = MIMEBase(main_type, sub_type)
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header(
                "Content-Disposition", "attachment", filename=os.path.basename(file)
            )
            attachment.add_header("Content-Transfer-Encoding", "base64")
            message.attach(attachment)

    return message.as_string()


def verificar_vencimento(cert_str: str) -> bool:
    """
    Verifica se o certificado (representado como uma string) está vencido.

    Args:
        cert_str (str): String que representa o certificado. Deve conter
            a data de vencimento no formato "Val. DD MM AAAA.pfx".

    Returns:
        bool: True se o certificado est  vencido, False caso contr rio.
    """
    try:
        partes = cert_str.split("Val. ")[-1]
        print(f"Partes: {partes}")
        dia, mes, ano = partes.replace(".pfx", "").split(" ")
        data_vencimento = datetime(int(f"20{ano}"), int(mes), int(dia))
        data_atual = datetime.now()
        print(f"Data venc: {data_vencimento}  Data atual: {data_atual}")

        if data_vencimento < data_atual:
            return True
        else:
            return False
    except Exception as e:
        console.print(
            f"Erro ao buscar data de vencimento no certificado: {compactar_diretorio}. Erro: {e}",
            style="bold red",
        )


def send_message(
    smtp_server: str,
    smtp_port: int,
    smtp_user: str,
    smtp_password: str,
    sender: str,
    recipient: str,
    message: str,
) -> None:
    """
    Envia uma mensagem via SMTP.

    Args:
        smtp_server (str): Servidor SMTP
        smtp_port (int): Porta do servidor SMTP
        smtp_user (str): Usuário do servidor SMTP
        smtp_password (str): Senha do usuário do servidor SMTP
        sender (str): Remetente do e-mail
        recipient (str): Destinatário do e-mail
        message (str): Corpo do e-mail

    Returns:
        None
    """
    if not smtp_server or not smtp_port or not smtp_user or not smtp_password:
        raise ValueError("Parâmetros SMTP não podem ser nulos")

    if not sender or not recipient or not message:
        raise ValueError("Parâmetros de e-mail não podem ser nulos")

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(sender, recipient, message)
            logger.info("Mensagem enviada com sucesso")
    except smtplib.SMTPAuthenticationError:
        console.print(f"Erro de autenticação. Verifique seu nome de usuário e senha.\n")
        logger.error("Erro de autenticação. Verifique seu nome de usuário e senha.")
    except smtplib.SMTPConnectError:
        console.print(
            f"Não foi possível conectar ao servidor SMTP. Verifique o endereço e a porta.\n"
        )
        logger.error(
            "Não foi possível conectar ao servidor SMTP. Verifique o endereço e a porta."
        )
    except smtplib.SMTPRecipientsRefused:
        console.print(
            f"O destinatário foi rejeitado. Verifique o endereço do destinatário.\n"
        )
        logger.error(
            "O destinatário foi rejeitado. Verifique o endereço do destinatário."
        )
    except smtplib.SMTPSenderRefused:
        console.print(
            f"O remetente foi rejeitado. Verifique o endereço do remetente..\n"
        )
        logger.error("O remetente foi rejeitado. Verifique o endereço do remetente.")
    except smtplib.SMTPDataError as e:
        console.print(f"Erro ao enviar dados: {e}\n")
        logger.error(f"Erro ao enviar dados: {e}")
    except Exception as error:
        console.print(f"Ocorreu um erro ao enviar a mensagem: {error} \n")
        logger.error("Ocorreu um erro ao enviar a mensagem: %s" % error)


async def ecac_federal(task: RpaProcessoEntradaDTO) -> RpaRetornoProcessoDTO:
    """
    Processo que realiza a consulta de caixa postal do ECAC Federal.

    """
    try:
        # Obtém a resolução da tela
        screen_width, screen_height = pyautogui.size()
        console.print(
            f"Resolução da tela: Width: {screen_width} - Height{screen_height}...\n"
        )
        console.print(f"Task:{task}")

        console.print("Realizando as validações inicias para execução do processo\n")
        console.print("Criando o diretório temporario ...\n")
        cd = Path.cwd()
        temp_dir = f"{cd}\\temp_certificates\\{str(uuid.uuid4())}"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        console.print("Obtendo configuração para execução do processo ...\n")
        try:
            config = await get_config_by_name("Ecac-Federal")
            emails_to = config.conConfiguracao.get("emails")
            certificado_path = config.conConfiguracao.get("CertificadoPath")
            planilha_id = config.conConfiguracao.get("SpreadSheets")

            console.print(
                "Obtendo configuração de email para execução do processo ...\n"
            )
            smtp_config = await get_config_by_name("SMTP")

            smtp_server = smtp_config.conConfiguracao.get("server")
            smtp_user = smtp_config.conConfiguracao.get("user")
            smtp_pass = smtp_config.conConfiguracao.get("password")
            smtp_port = smtp_config.conConfiguracao.get("port")

            get_gcp_token = sync_get_config_by_name("GCP_SERVICE_ACCOUNT")
            get_gcp_credentials = sync_get_config_by_name("GCP_CREDENTIALS")
        except Exception as e:
            console.print(
                f"Erro ao obter as configurações para execução do processo, erro: {e}...\n"
            )
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Erro ao obter as configurações para execução do processo, erro: {e}",
                status=RpaHistoricoStatusEnum.Falha, tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)]
            )

        try:
            console.print("Realizando a leitura da planilha com os certificados...\n")
            gcp_credencial = GetCredsGworkspace(
                token_dict=get_gcp_token.conConfiguracao,
                credentials_dict=get_gcp_credentials.conConfiguracao,
            )
            creds = gcp_credencial.get_creds_gworkspace()

            if not creds:
                console.print(f"Erro ao obter autenticação para o GCP...\n")
                return RpaRetornoProcessoDTO(
                    sucesso=False,
                    retorno=f"Erro ao obter autenticação para o GCP",
                    status=RpaHistoricoStatusEnum.Falha,
                    tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
                )

            spreadsheet_id = planilha_id
            sheet_name = "ECACFederal"

            service = build("sheets", "v4", credentials=creds, static_discovery=False)
            sheet = service.spreadsheets()
            result = (
                sheet.values()
                .get(spreadsheetId=spreadsheet_id, range=f"{sheet_name}")
                .execute()
            )
            values = result.get("values", [])
        except Exception as e:
            console.print(
                f"Não foi possivel obter a planilha no Google Drive para execução do processo, erro: {e}...\n"
            )
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Não foi possivel obter a planilha no Google Drive para execução do processo, erro: {e}",
                status=RpaHistoricoStatusEnum.Falha, tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)]
            )

        console.print("Verificando a existência do diretório com locators...\n")
        if os.path.exists("assets/ecac_federal"):
            console.print("Diretório existe..\n")
        else:
            console.print("Diretório não existe..\n")
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Não foi possivel encontrar o diretório com os locators para continuidade do processo, diretório: 'assets/ecac_federal'",
                status=RpaHistoricoStatusEnum.Falha, tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)]
            )

        console.print(
            "Verificando a existência dos arquivos de certificados no Google Drive...\n"
        )
        drive_service = build("drive", "v3", credentials=creds)
        query = f"'{certificado_path}' in parents"
        results = (
            drive_service.files()
            .list(
                q=query,
                pageSize=1000,
                supportsAllDrives=True,  # Habilita suporte para Shared Drives
                includeItemsFromAllDrives=True,  # Inclui itens de todos os drives
                fields="files(id, name)",
            )
            .execute()
        )

        items = results.get("files", [])

        if not items:
            console.print(f"Nenhum certificado encontrado...\n")
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Nenhum certificado encontrado no diretório do Google Drive",
                status=RpaHistoricoStatusEnum.Falha, tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)]
            )

        certificate_files = pd.DataFrame(values[1:], columns=values[0])
        report = []
        feedback_cnpj = []
        count_error = 0

        for index, row in certificate_files.iterrows():
            try:
                empresa = row["Empresa"]
                certificado = row["Certificado"]
                # validade = row['Validade']
                senha = row["Senha"]

                console.print(
                    f"Procurando certificado: {certificado} para a empresa {empresa}...\n"
                )

                drive_service = build("drive", "v3", credentials=creds)
                query = f"'{certificado_path}' in parents and name='{certificado}'"
                results = (
                    drive_service.files()
                    .list(
                        q=query,
                        pageSize=1000,
                        supportsAllDrives=True,  # Habilita suporte para Shared Drives
                        includeItemsFromAllDrives=True,  # Inclui itens de todos os drives
                        fields="files(id, name)",
                    )
                    .execute()
                )

                items = results.get("files", [])

                if items:
                    console.print(
                        f"Certificado: {certificado} para a empresa {empresa} encontrado...\n"
                    )
                    file_id = items[0]["id"]
                    console.print(
                        f"Certificado {certificado} encontrado. Iniciando download...\n"
                    )

                    file_path = f"{temp_dir}\\{certificado}"

                    request = drive_service.files().get_media(fileId=file_id)

                    fh = io.FileIO(file_path, "wb")
                    downloader = MediaIoBaseDownload(fh, request)

                    done = False
                    while done is False:
                        status, done = downloader.next_chunk()
                        console.print(f"Download {int(status.progress() * 100)}%.")

                    fh.close()

                    console.print(f"Certificado {certificado} baixado com sucesso.\n")

                    certificado_vencido = verificar_vencimento(certificado)

                    if certificado_vencido:
                        console.print(
                            f"Não foi possivel realizar a importação do certificado, pois está vencido: {certificado}",
                            style="bold yellow",
                        )
                        result = {
                            "error": True,
                            "message": "Não foi possivel realizar a importação do certificado, pois está vencido:",
                            "certificado": certificado,
                        }
                        feedback_cnpj.append(
                            f"Não foi possivel realizar a importação do certificado, pois está vencido: {certificado}"
                        )
                        continue

                    async with async_playwright() as p:
                        browser = await p.firefox.launch(headless=False, timeout=15000)
                        console.print("Iniciando o browser")
                        context = await browser.new_context(
                            accept_downloads=True, ignore_https_errors=True
                        )
                        page = await context.new_page()

                        await page.goto(
                            "about:preferences#privacy",
                            timeout=15000,
                            wait_until="load",
                        )
                        locator_path = f"{cd}\\assets\\ecac_federal"

                        element_find_settings_image = (
                            f"{locator_path}/FindInSettings.PNG"
                        )
                        element_view_certificates = (
                            f"{locator_path}/ViewCertificates.PNG"
                        )
                        element_your_certificates = (
                            f"{locator_path}/YourCertificates.PNG"
                        )
                        element_import_certificates = (
                            f"{locator_path}/ImportCertificate.PNG"
                        )
                        element_insert_path_certificates = (
                            f"{locator_path}/InsertPathCertificate.PNG"
                        )
                        element_open_certificates = (
                            f"{locator_path}/OpenCertificate.PNG"
                        )
                        element_pwd_certificates = (
                            f"{locator_path}/EnterPasswordCertificate.PNG"
                        )
                        element_sign_in_certificates = (
                            f"{locator_path}/SignCertificate.PNG"
                        )
                        element_confirm_certificates = f"{locator_path}/Confirm.PNG"
                        element_sign_in_certificates = (
                            f"{locator_path}/SignCertificate.PNG"
                        )
                        element_popup_certificate = (
                            f"{locator_path}/ImportCertificate_popup.PNG"
                        )
                        element_certificate_expired = (
                            f"{locator_path}/CertificateNotFoundOrExpired.PNG"
                        )
                        element_forbidden = f"{locator_path}/Forbidden.PNG"
                        element_seu_certificado_digital = (
                            f"{locator_path}/SeuCertificadoDigital.png"
                        )

                        console.print(
                            "Importando o certificado no browser (playwright) "
                        )
                        # Configurações
                        find_and_click(
                            element_find_settings_image, "Configurações", wait_time=10
                        )
                        pyautogui.write("Cert", interval=0.05)
                        await asyncio.sleep(1)

                        # Certificados
                        find_and_click(element_view_certificates, "Certificados")
                        await asyncio.sleep(1)

                        # Meus Certificados
                        find_and_click(element_your_certificates, "Meus Certificados")
                        await asyncio.sleep(1)

                        # Importar Certificado
                        find_and_click(
                            element_import_certificates, "Importar Certificado"
                        )
                        await asyncio.sleep(2)

                        # Inserir Caminho do Certificado e escrever o caminho do arquivo
                        find_and_click(
                            element_insert_path_certificates,
                            "Inserir Caminho do Certificado",
                            write_text=file_path,
                        )
                        await asyncio.sleep(2)

                        # Abrir Certificado
                        find_and_click(element_open_certificates, "Abrir Certificado")
                        await asyncio.sleep(2)

                        # Inserir Senha do Certificado
                        find_and_click(
                            element_pwd_certificates,
                            "Inserir Senha do Certificado",
                            write_text=senha,
                        )
                        await asyncio.sleep(2)

                        # Assinar Certificado
                        find_and_click(
                            element_sign_in_certificates, "Assinar Certificado"
                        )
                        await asyncio.sleep(2)

                        # Confirmar Importação
                        find_and_click(
                            element_confirm_certificates, "Confirmar Importação"
                        )
                        await asyncio.sleep(2)

                        try:
                            element_certificate_expired_pos = (
                                pyautogui.locateCenterOnScreen(
                                    element_certificate_expired, confidence=0.8
                                )
                            )
                        except:
                            element_certificate_expired_pos = None

                        if element_certificate_expired_pos != None:
                            console.print(
                                f"Não foi possivel realizar a importação do certificado, pois está vencido: {certificado}",
                                style="bold yellow",
                            )
                            result = {
                                "error": True,
                                "message": "Não foi possivel realizar a importação do certificado, pois está vencido:",
                                "certificado": certificado,
                            }
                            feedback_cnpj.append(
                                f"Não foi possivel realizar a importação do certificado, pois está vencido: {certificado}"
                            )
                            continue

                        try:
                            element_forbidden_pos = pyautogui.locateCenterOnScreen(
                                element_forbidden, confidence=0.8
                            )
                        except:
                            element_forbidden_pos = None

                        if element_forbidden_pos != None:
                            console.print(
                                f"[403 - Forbidden] Não foi possivel realizar o login com o certificado: {certificado}",
                                style="bold yellow",
                            )
                            result = {
                                "error": True,
                                "message": "[403 - Forbidden] Não foi possivel realizar o login com o certificado:",
                                "certificado": certificado,
                            }
                            feedback_cnpj.append(
                                f"[403 - Forbidden] Não foi possivel realizar o login com o certificado: {certificado}"
                            )
                            continue

                        try:
                            element_popup_certificate_pos = (
                                pyautogui.locateCenterOnScreen(
                                    element_popup_certificate, confidence=0.8
                                )
                            )
                        except:
                            element_popup_certificate_pos = None

                        if element_popup_certificate_pos == None:
                            await asyncio.sleep(3)
                            try:
                                await page.goto(
                                    "https://certificado.sso.acesso.gov.br",
                                    timeout=10000,
                                    wait_until="load",
                                )
                                await page.keyboard.press("Enter")
                            except:
                                await page.keyboard.press("Enter")
                                await page.goto(
                                    "https://certificado.sso.acesso.gov.br",
                                    timeout=30000,
                                    wait_until="load",
                                )
                                await page.keyboard.press("Enter")

                                await asyncio.sleep(5)
                                XPATH_ECAC_ENTRAR = (
                                    '//*[@id="login-dados-certificado"]/p[2]/input'
                                )
                                XPATH_SSO_GOV = '//*[@id="login-certificate"]'
                                XPATH_LOGIN_EFETUADO = '//*[@id="sairSeguranca"]'

                                await page.keyboard.press("Enter")
                                await asyncio.sleep(3)
                                ecac_login_url = "https://cav.receita.fazenda.gov.br/autenticacao/login"
                                await page.goto(
                                    ecac_login_url, timeout=80000, wait_until="load"
                                )
                                try:
                                    await page.wait_for_load_state("networkidle")
                                except:
                                    await asyncio.sleep(3)
                                await asyncio.sleep(6)

                                await page.locator(XPATH_ECAC_ENTRAR).click(
                                    timeout=80000
                                )
                                try:
                                    await page.wait_for_load_state("networkidle")
                                except:
                                    await asyncio.sleep(3)

                                await asyncio.sleep(9)
                                await page.locator(XPATH_SSO_GOV).dblclick(
                                    timeout=90000
                                )
                                await asyncio.sleep(9)

                                try:
                                    await page.locator(
                                        '//*[@id="identidade-govbr"]'
                                    ).click()
                                except:
                                    pass

                                await asyncio.sleep(3)

                                try:
                                    element = pyautogui.locateCenterOnScreen(
                                        element_seu_certificado_digital, confidence=0.8
                                    )
                                    if element:
                                        pyautogui.click(element)
                                        console.print(
                                            f"[green] Seu certificado digital encontrado e clicado.[/green]"
                                        )
                                        await asyncio.sleep(3)
                                except:
                                    await page.locator(
                                        '//*[@id="login-certificate"]'
                                    ).click()
                                    await asyncio.sleep(2)

                                await page.keyboard.press("Enter")

                                await asyncio.sleep(5)
                                login_efetuado = False
                                try:
                                    await page.wait_for_selector(
                                        XPATH_LOGIN_EFETUADO, timeout=80000
                                    )
                                    login_efetuado = True
                                except Exception:
                                    login_efetuado = False

                                if login_efetuado:
                                    console.print(
                                        "Navegando para aba de leitura das mensagens"
                                    )

                                    XPATH_ECAC_OPEN_MESSAGES = (
                                        '//*[@id="btnCaixaPostal"]'
                                    )
                                    XPATH_TABLE_MESSAGES = '//*[@id="gridMensagens"]'
                                    XPATH_GO_TO_LOGOFF_EFETUADO = (
                                        '//*[@id="sairSeguranca"]'
                                    )
                                    XPATH_MESSAGE = '//*[@id="lbDtEnvio"]'
                                    XPATH_DESTINATION_MESSAGE = (
                                        '//*[@id="lbValorCNPJReferencia"]'
                                    )
                                    XPATH_INFO_TITULAR_ACESSO = (
                                        '//*[@id="informacao-perfil"]'
                                    )
                                    XPATH_PRIMEIRA_LEITURA_MESSAGE = (
                                        '//*[@id="lbValorPrimeiraLeitura"]'
                                    )
                                    XPATH_INFO_CERTIFICADO = (
                                        '//*[@id="avisoCertificado"]'
                                    )

                                    console.print(f"Capturando o responsavel legal")
                                    try:
                                        await page.wait_for_selector(
                                            XPATH_INFO_TITULAR_ACESSO, timeout=30000
                                        )
                                        titular_accesso = await page.locator(
                                            XPATH_INFO_TITULAR_ACESSO
                                        ).text_content(timeout=30000)
                                        regex = r"Titular\s+\(.*?\):\s+[\d./-]+\s*-\s*(.+?)(?=\s+Responsável Legal:)"
                                        match = re.search(
                                            regex, titular_accesso, re.DOTALL
                                        )
                                        titular_accesso = match.group(1).strip()
                                        console.print(
                                            f"Responsavel legal: {titular_accesso}"
                                        )
                                    except:
                                        titular_accesso = ""
                                        console.print(
                                            f"Não foi possivel capturar o responsavel legal: {titular_accesso}"
                                        )

                                    try:
                                        certificate_message = await page.text_content(
                                            XPATH_INFO_CERTIFICADO, timeout=30000
                                        )
                                    except:
                                        certificate_message = ""

                                    await page.goto(
                                        "https://cav.receita.fazenda.gov.br/Servicos/ATSDR/CaixaPostal.app/Action/ListarMensagensAction_V20230215.aspx",
                                        timeout=30000,
                                        wait_until="load",
                                    )

                                    await page.locator(XPATH_TABLE_MESSAGES)
                                    await asyncio.sleep(3)

                                    # Obter o HTML da tabela
                                    html_content = await page.content()

                                    with open(
                                        "filename.html", "w", encoding="utf-8"
                                    ) as file:
                                        file.write(html_content)

                                    # Extrair os dados da tabela do HTML
                                    table_data = extract_table(html_content)

                                    if table_data:
                                        console.print(
                                            f"Dados da tabela extraídos com sucesso: {certificado}"
                                        )
                                        for row in table_data:
                                            if len(row[1]) > 10:
                                                today_date = datetime.now()
                                                formatted_date = today_date.strftime(
                                                    "%d/%m/%Y"
                                                )
                                                formatted_date = datetime.strptime(
                                                    formatted_date, "%d/%m/%Y"
                                                )
                                                today_date = formatted_date

                                                sent_message_date = datetime.strptime(
                                                    row[4], "%d/%m/%Y"
                                                )
                                                seven_days_ago = today_date - timedelta(
                                                    days=5
                                                )

                                                if (
                                                    seven_days_ago
                                                    <= sent_message_date
                                                    <= today_date
                                                ):
                                                    await page.goto(
                                                        f"https://cav.receita.fazenda.gov.br/Servicos/ATSDR/CaixaPostal.app/Action/{row[1]}",
                                                        wait_until="load",
                                                    )
                                                    await page.wait_for_selector(
                                                        XPATH_MESSAGE, timeout=10000
                                                    )

                                                    if await page.query_selector(
                                                        XPATH_MESSAGE
                                                    ):
                                                        destination_message = await page.text_content(
                                                            XPATH_DESTINATION_MESSAGE,
                                                            timeout=30000,
                                                        )
                                                        first_read_date_message = await page.text_content(
                                                            XPATH_PRIMEIRA_LEITURA_MESSAGE,
                                                            timeout=30000,
                                                        )

                                                        message_html_content = (
                                                            await page.content()
                                                        )
                                                        temp_htmlfile = "temp_page.html"

                                                        cnpj = re.sub(r"\W+", "", cnpj)
                                                        today_date = datetime.now()
                                                        date = today_date.strftime(
                                                            "%d-%m-%Y"
                                                        )
                                                        date = re.sub(r"\W+", "", date)
                                                        unique_id = uuid.uuid4()
                                                        first_4_digits = str(
                                                            unique_id
                                                        ).replace("-", "")[:4]
                                                        os.makedirs(
                                                            os.path.join(
                                                                temp_dir, empresa
                                                            ),
                                                            exist_ok=True,
                                                        )
                                                        file_fullname = os.path.join(
                                                            temp_dir,
                                                            empresa,
                                                            f"{date}_{cnpj}_{first_4_digits}.pdf",
                                                        )

                                                        with open(
                                                            temp_htmlfile,
                                                            "w",
                                                            encoding="utf-8",
                                                        ) as f:
                                                            f.write(
                                                                message_html_content
                                                            )

                                                        with open(
                                                            temp_htmlfile,
                                                            "r",
                                                            encoding="utf-8",
                                                        ) as f:
                                                            html_content = f.read()

                                                        with open(
                                                            file_fullname, "wb"
                                                        ) as pdf:
                                                            pisa_status = (
                                                                pisa.CreatePDF(
                                                                    html_content,
                                                                    dest=pdf,
                                                                )
                                                            )

                                                        dados = {
                                                            "Status": "Possui mensagem a ser lida",
                                                            "Destinatario": destination_message,
                                                            "Remetente": row[0],
                                                            "Assunto": row[2],
                                                            "Enviada em": row[4],
                                                            "Aberta em": first_read_date_message,
                                                            "AVISO": certificate_message,
                                                            "Titular": titular_accesso,
                                                        }

                                                        report.append(dados)

                                    else:
                                        dados = {
                                            "Status": "Nenhum  mensagem a ser lida",
                                            "Destinatario": destination_message,
                                            "Remetente": "",
                                            "Assunto": "",
                                            "Enviada em": "",
                                            "Aberta em": "",
                                            "AVISO": certificate_message,
                                            "Titular": titular_accesso,
                                        }
                                        report.append(dados)

                                        result = {
                                            "error": False,
                                            "message": f"Proceso completado com sucesso para o certificado {certificado}",
                                        }
                                        feedback_cnpj.append(
                                            f"Processo finalizado com sucesso para o certificado: {certificado}"
                                        )

                                    await asyncio.sleep(3)
                                    await page.goto(
                                        "https://cav.receita.fazenda.gov.br/ecac/",
                                        wait_until="load",
                                    )
                                    await asyncio.sleep(1)
                                    await page.locator(
                                        XPATH_GO_TO_LOGOFF_EFETUADO
                                    ).click()
                                    await browser.close()

                                else:
                                    count_error += 1
                                    console.print(
                                        f"Não foi possivel realizar o login no site do ECAC, devido a bloqueio de captcha: {certificado}"
                                    )
                                    result = {
                                        "error": True,
                                        "message": "Não foi possivel realizar o login no site do ECAC, devido a bloqueio de captcha",
                                        "certificado": certificado,
                                    }
                                    feedback_cnpj.append(
                                        f"Não foi possivel realizar o login no site do ECAC, devido a bloqueio de captcha: {certificado}"
                                    )
                                    await browser.close()

                        else:
                            count_error += 1
                            console.print(
                                f"Não foi possivel realizar a importação do certificado: {certificado}"
                            )
                            result = {
                                "error": True,
                                "message": "Não foi possivel realizar a importação do certificado",
                                "certificado": certificado,
                            }
                            feedback_cnpj.append(
                                f"Não foi possivel realizar a importação do certificado: {certificado}"
                            )

                else:
                    count_error += 1
                    console.print(
                        f"Erro para o {certificado} - Certificado não encontrado no Google Drive"
                    )
                    result = {
                        "error": False,
                        "message": "Certificado não encontrado no Google Drive",
                        "certificado": certificado,
                    }
                    feedback_cnpj.append(
                        f"Certificado não encontrado no Google Drive: {certificado}"
                    )

            except Exception as e:
                count_error += 1
                console.print(f"Erro para {certificado}: {str(e)}, Mensagem: {e}")
                result = {
                    "error": True,
                    "error_message": str(e),
                    "certificate": certificado,
                }
                feedback_cnpj.append(
                    f" Não foi possivel concluir o processo para o centificado: {certificado}"
                )

        if count_error == len(certificate_files):
            log_msg = f"Erro Processo do Ecac Federal, não obtivemos exito em nenhum certificado executado, erro: {feedback_cnpj}"
            logger.error(log_msg)
            console.print(log_msg, style="bold red")
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=log_msg,
                status=RpaHistoricoStatusEnum.Falha, tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)]
            )

        else:
            messages_found = False
            for sublista in report:
                if sublista:
                    messages_found = True

            report_to_user("consultaFederal_ecac.xlsx", report, os.path.join(temp_dir))
            feedback_cnpj_str = "\n".join(feedback_cnpj)
            corpo_email = f""" Caros(as), \nEspero que este e-mail o encontre bem ! \n\n Abaixo, segue um resumo de facil acesso as consultas das caixas federeias.
            \n Controladoria Federal:
            1 - Para consulta temos acesso a {len(certificate_files)} certificados e o RPA tem a responsabilidade de acessar diariamente o sistema e-cac para consulta de caixa postais.\n
            CNPJs buscados e seus respectivos status: \n
            {feedback_cnpj_str}

            Feedback de execução:
            """

            if messages_found:

                feedback_agrupado = {}

                for sublist in report:
                    for message in sublist:
                        destinatario = message["Destinatario"]
                        if destinatario not in feedback_agrupado:
                            feedback_agrupado[destinatario] = []
                        feedback_agrupado[destinatario].append(message)

                # Adicionar feedback ao corpo do e-mail
                for destinatario, messages in feedback_agrupado.items():
                    corpo_email += f"\nPara o destinatário: {destinatario}, foram encontradas as seguintes mensagens:\n"
                    for message in messages:
                        acesso = message["Titular"]
                        remetente = message["Remetente"]
                        assunto = message["Assunto"]
                        enviada_em = message["Aberta em"]
                        certificado_expira_em = message["AVISO"]
                        corpo_email += f"\n- Acesso: {acesso}\n Remetente: {remetente}\n  Assunto: {assunto}\n  Enviada em: {enviada_em}\n  Obs: {certificado_expira_em}\n"

                corpo_email += (
                    f"\nEstamos disponivel para quaisquer duvidas ou sugestões.\n"
                )
            else:
                corpo_email += "Nenhuma mensagem para fornecer resumo da mensagem. \nEstamos disponivel para quaisquer duvidas ou sugestões.\n"

            try:
                send_email(
                    smtp_server,
                    smtp_port,
                    smtp_user,
                    smtp_pass,
                    corpo_email,
                    "RPA - ECAC Federal",
                    emails_to,
                    temp_dir,
                )
                log_msg = f"Processo concluido e e-mail disparado para area de negócio"
                console.print(log_msg)
                return RpaRetornoProcessoDTO(
                    sucesso=True,
                    retorno=log_msg,
                    status=RpaHistoricoStatusEnum.Sucesso,
                )
            except Exception as e:
                log_msg = f"Processo concluido porém não foi possivel enviar o e-mail para area de negóciol: {e}"
                logger.error(log_msg)
                console.print(log_msg, style="bold red")
                return RpaRetornoProcessoDTO(
                    sucesso=False,
                    retorno=log_msg,
                    status=RpaHistoricoStatusEnum.Falha, tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)]
                ) 

    except Exception as e:
        log_msg = f"Erro Processo do Ecac Federal: {e}"
        logger.error(log_msg)
        console.print(log_msg, style="bold red")
        return RpaRetornoProcessoDTO(
            sucesso=False,
            retorno=log_msg,
            status=RpaHistoricoStatusEnum.Falha, tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)]
        )

    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

import asyncio
import io
import mimetypes
import os
import shutil
import smtplib
import subprocess
import time
import uuid
import zipfile
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
import pandas as pd
import psutil
import pyautogui
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from playwright.async_api import async_playwright
from rich.console import Console
from worker_automate_hub.api.client import get_config_by_name, sync_get_config_by_name
from worker_automate_hub.models.dto.rpa_historico_request_dto import (
    RpaHistoricoStatusEnum,
    RpaRetornoProcessoDTO,
    RpaTagDTO,
    RpaTagEnum,
)
from worker_automate_hub.models.dto.rpa_processo_entrada_dto import (
    RpaProcessoEntradaDTO,
)
from worker_automate_hub.utils.get_creds_gworkspace import GetCredsGworkspace
from worker_automate_hub.utils.logger import logger

console = Console()
pyautogui.FAILSAFE = False


def fechar_aplicativo(nome_processo: str) -> None:
    """
    Termina o processo com o nome especificado.

    Args:
        nome_processo (str): Nome do processo a ser terminado.

    Returns:
        None
    """
    encontrado = False
    for proc in psutil.process_iter(["name"]):
        try:
            if proc.info["name"].lower() == nome_processo.lower():
                console.print(f"Fechando o processo {nome_processo}...", style="bold red")
                proc.terminate()
                proc.wait()
                encontrado = True
                break
            return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
            console.print(f"Erro ao tentar finalizar o processo: {str(e)}", style="bold yellow")
            return False

    if not encontrado:
        console.print(f"O processo {nome_processo} não está em execução.", style="bold yellow")
        return True

def find_and_click(
    image_path, description, confidence=0.8, write_text=None, wait_time=2
):
    element = pyautogui.locateCenterOnScreen(image_path, confidence=confidence)
    if element:
        pyautogui.click(element)
        console.print(f"[green]'{description}' encontrado e clicado.[/green]")
        if write_text:
            pyautogui.write(write_text, interval=0.05)
            console.print(
                f"[green]Texto '{write_text}' inserido em '{description}'.[/green]"
            )
        time.sleep(wait_time)
    else:
        raise Exception(f"[red]Elemento '{description}' não encontrado![/red]")


def report_to_user(filename, report, path_directory):
    fullpath = os.path.join(path_directory, filename)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        "Número do processo",
        "Assunto",
        "Tipo de Comunicação",
        "Data Final Ciência",
        "Forma Ciência",
        "Detalhamento",
        "Inteiro Teor",
    ]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num).value = header

    row_num = 2
    for page_data in report:
        for data in page_data:
            col_num = 1
            for key in headers:
                ws.cell(row=row_num, column=col_num).value = data.get(key, "")
                col_num += 1
            row_num += 1

    wb.save(fullpath)


def compactar_diretorio(diretorio, nome_arquivo_zip):
    arquivos = [
        os.path.join(diretorio, arquivo)
        for arquivo in os.listdir(diretorio)
        if os.path.isfile(os.path.join(diretorio, arquivo))
    ]

    with zipfile.ZipFile(nome_arquivo_zip, "w") as zipf:
        for arquivo in arquivos:
            zipf.write(arquivo, os.path.basename(arquivo))


def send_email(
    smtp_server,
    smtp_port,
    smtp_user,
    smtp_password,
    message_text,
    subject,
    to,
    diretorio,
):

    nome_arquivo_zip = "Consulta DJE.zip"
    full_path = os.path.join(diretorio, nome_arquivo_zip)
    compactar_diretorio(diretorio, full_path)

    message = create_message_with_attachment(
        smtp_user, to, subject, message_text, full_path
    )
    send_message(
        smtp_server, smtp_port, smtp_user, smtp_password, smtp_user, to, message
    )


def create_message_with_attachment(sender, to, subject, message_text, file):
    message = MIMEMultipart()
    message["to"] = to
    message["from"] = sender
    message["subject"] = subject

    msg = MIMEText(message_text)
    message.attach(msg)

    if file:
        content_type, encoding = mimetypes.guess_type(file)

        if content_type is None or encoding is not None:
            content_type = "application/octet-stream"

        main_type, sub_type = content_type.split("/", 1)
        with open(file, "rb") as f:
            attachment = MIMEBase(main_type, sub_type)
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header(
                "Content-Disposition", "attachment", filename=os.path.basename(file)
            )
            attachment.add_header("Content-Transfer-Encoding", "base64")
            message.attach(attachment)

    # raw_message = base64.urlsafe_b64encode(message.as_bytes())
    # raw_message = raw_message.decode()
    # return {'raw': raw_message}
    return message.as_string()


def send_message(
    smtp_server, smtp_port, smtp_user, smtp_password, sender, recipient, message
):
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(sender, recipient, message)
            logger.info("Mensagem enviada com sucesso")
    except smtplib.SMTPAuthenticationError:
        console.print(f"Erro de autenticação. Verifique seu nome de usuário e senha.\n")
        logger.error("Erro de autenticação. Verifique seu nome de usuário e senha.")
    except smtplib.SMTPConnectError:
        console.print(
            f"Não foi possível conectar ao servidor SMTP. Verifique o endereço e a porta.\n"
        )
        logger.error(
            "Não foi possível conectar ao servidor SMTP. Verifique o endereço e a porta."
        )
    except smtplib.SMTPRecipientsRefused:
        console.print(
            f"O destinatário foi rejeitado. Verifique o endereço do destinatário.\n"
        )
        logger.error(
            "O destinatário foi rejeitado. Verifique o endereço do destinatário."
        )
    except smtplib.SMTPSenderRefused:
        console.print(
            f"O remetente foi rejeitado. Verifique o endereço do remetente..\n"
        )
        logger.error("O remetente foi rejeitado. Verifique o endereço do remetente.")
    except smtplib.SMTPDataError as e:
        console.print(f"Erro ao enviar dados: {e}\n")
        logger.error(f"Erro ao enviar dados: {e}")
    except Exception as error:
        console.print(f"Ocorreu um erro ao enviar a mensagem: {error} \n")
        logger.error("Ocorreu um erro ao enviar a mensagem: %s" % error)


async def coleta_dje_start_update(task: RpaProcessoEntradaDTO) -> RpaRetornoProcessoDTO:
    from pywinauto.application import Application

    """
    Processo que realiza a consulta de processo no DJE.

    """
    try:
        # Obtém a resolução da tela
        screen_width, screen_height = pyautogui.size()
        console.print(
            f"Resolução da tela: Width: {screen_width} - Height{screen_height}...\n"
        )
        console.print(f"Task:{task}")

        console.print("Realizando as validações inicias para execução do processo\n")

        console.print("Criando o diretório temporario ...\n")
        cd = Path.cwd()
        temp_dir = f"{cd}\\temp_certificates\\{str(uuid.uuid4())}"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        console.print("Obtendo configuração para execução do processo ...\n")
        try:
            config = await get_config_by_name("DJE")
            emails_to = config.conConfiguracao.get("emails")
            certificado_path = config.conConfiguracao.get("CertificadoPath")
            planilha_id = config.conConfiguracao.get("SpreadSheets")

            console.print(
                "Obtendo configuração de email para execução do processo ...\n"
            )
            smtp_config = await get_config_by_name("SMTP")

            smtp_server = smtp_config.conConfiguracao.get("server")
            smtp_user = smtp_config.conConfiguracao.get("user")
            smtp_pass = smtp_config.conConfiguracao.get("password")
            smtp_port = smtp_config.conConfiguracao.get("port")

            get_gcp_token = sync_get_config_by_name("GCP_SERVICE_ACCOUNT")
            get_gcp_credentials = sync_get_config_by_name("GCP_CREDENTIALS")
        except Exception as e:
            console.print(
                f"Erro ao obter as configurações para execução do processo, erro: {e}...\n"
            )
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Erro ao obter as configurações para execução do processo, erro: {e}",
                status=RpaHistoricoStatusEnum.Falha,
                tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
            )

        try:
            console.print("Realizando a leitura da planilha com os certificados...\n")
            gcp_credencial = GetCredsGworkspace(
                token_dict=get_gcp_token.conConfiguracao,
                credentials_dict=get_gcp_credentials.conConfiguracao,
            )
            creds = gcp_credencial.get_creds_gworkspace()

            if not creds:
                console.print(f"Erro ao obter autenticação para o GCP...\n")
                return RpaRetornoProcessoDTO(
                    sucesso=False,
                    retorno=f"Erro ao obter autenticação para o GCP",
                    status=RpaHistoricoStatusEnum.Falha,
                    tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
                )

            spreadsheet_id = planilha_id
            sheet_name = "ECACFederal"

            service = build("sheets", "v4", credentials=creds, static_discovery=False)
            sheet = service.spreadsheets()
            result = (
                sheet.values()
                .get(spreadsheetId=spreadsheet_id, range=f"{sheet_name}")
                .execute()
            )
            values = result.get("values", [])
        except Exception as e:
            console.print(
                f"Não foi possivel obter a planilha no Google Drive para execução do processo, erro: {e}...\n"
            )
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Não foi possivel obter a planilha no Google Drive para execução do processo, erro: {e}",
                status=RpaHistoricoStatusEnum.Falha,
                tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
            )

        console.print("Verificando a existência do diretório com locators...\n")
        if os.path.exists("assets/dje"):
            console.print("Diretório existe..\n")
        else:
            console.print("Diretório não existe..\n")
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Não foi possivel encontrar o diretório com os locators para continuidade do processo, diretório: 'assets/dje'",
                status=RpaHistoricoStatusEnum.Falha,
                tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
            )

        console.print(
            "Verificando a existência dos arquivos de certificados no Google Drive...\n"
        )
        drive_service = build("drive", "v3", credentials=creds)
        query = f"'{certificado_path}' in parents"
        results = (
            drive_service.files()
            .list(
                q=query,
                pageSize=1000,
                supportsAllDrives=True,  # Habilita suporte para Shared Drives
                includeItemsFromAllDrives=True,  # Inclui itens de todos os drives
                fields="files(id, name)",
            )
            .execute()
        )

        items = results.get("files", [])

        if not items:
            console.print(f"Nenhum certificado encontrado...\n")
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=f"Nenhum certificado encontrado no diretório do Google Drive",
                status=RpaHistoricoStatusEnum.Falha,
                tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
            )

        certificate_files = pd.DataFrame(values[1:], columns=values[0])
        report = []
        feedback_cnpj = []
        count_error = 0

        for index, row in certificate_files.iterrows():
            try:
                empresa = row["Empresa"]
                certificado = row["Certificado"]
                senha = row["Senha"]
                console.print(
                    f"Certificado a ser executado: {certificado} para a empresa {empresa}...\n"
                )

                nome_aplicativo = "javaw.exe"
                console.print("\nPJeOffice PRO encerrando...", style="bold green")

                fechar_dje =  fechar_aplicativo(nome_aplicativo)
                if fechar_dje:
                    console.print("\nPJeOffice PRO encerrado com sucesso...", style="bold green")
                else:
                    log_msg = f"Erro Processo do DJE, não foi possivel encerrar a aplicação: {e}"
                    return RpaRetornoProcessoDTO(
                        sucesso=False,
                        retorno=log_msg,
                        status=RpaHistoricoStatusEnum.Falha,
                        tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
                    )

                # Abre o PJeOffice
                console.print("\nPJeOffice PRO iniciando...", style="bold green")
                try:
                    subprocess.Popen(
                        ["C:\\Program Files\\PJeOffice Pro\\pjeoffice-pro.exe"],
                        shell=True,
                    )
                except Exception as e:
                    log_msg = f"Erro Processo do DJE, não foi possivel iniciar a aplicação: {e}"
                    console.print(log_msg, style="bold red")
                    return RpaRetornoProcessoDTO(
                        sucesso=False,
                        retorno=log_msg,
                        status=RpaHistoricoStatusEnum.Falha,
                        tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
                    )

                console.print(
                    f"Procurando certificado: {certificado} para a empresa {empresa}...\n"
                )
                drive_service = build("drive", "v3", credentials=creds)
                query = f"'{certificado_path}' in parents and name='{certificado}'"
                results = (
                    drive_service.files()
                    .list(
                        q=query,
                        pageSize=1000,
                        supportsAllDrives=True,  # Habilita suporte para Shared Drives
                        includeItemsFromAllDrives=True,  # Inclui itens de todos os drives
                        fields="files(id, name)",
                    )
                    .execute()
                )

                items = results.get("files", [])

                if items:
                    console.print(
                        f"Certificado: {certificado} para a empresa {empresa} encontrado...\n"
                    )
                    file_id = items[0]["id"]
                    console.print(
                        f"Certificado {certificado} encontrado. Iniciando download...\n"
                    )

                    file_path = os.path.join(temp_dir, certificado)

                    request = drive_service.files().get_media(fileId=file_id)

                    fh = io.FileIO(file_path, "wb")
                    downloader = MediaIoBaseDownload(fh, request)

                    done = False
                    while done is False:
                        status, done = downloader.next_chunk()
                        console.print(f"Download {int(status.progress() * 100)}%.")

                    fh.close()

                    console.print(f"Certificado {certificado} baixado com sucesso.\n")

                    locator_path = "assets/dje"
                    autorizacao_site = f"{locator_path}/AutorizacaoSite_Pje.png"
                    sim_destavez = f"{locator_path}/Sim_destavez.png"
                    informar_senha_existe = f"{locator_path}/InformeSenha.png"
                    cancelar = f"{locator_path}/Cancelar.png"
                    certificados_disponiveis = (
                        f"{locator_path}/CertificadosDisponiveis.png"
                    )
                    engrenagem_certificado = (
                        f"{locator_path}/Engrenagem_Configuracao.png"
                    )
                    tipo_certificado = f"{locator_path}/TipoCertificado.png"
                    remover_certificado = f"{locator_path}/RemoverCertificado.png"
                    localizar = f"{locator_path}/SelecionarCertificado.png"
                    cert_fullpath = f"{locator_path}/DiretorioCertificado.png"
                    confirmar_certificado = f"{locator_path}/Abrir.png"
                    ok_confirmar_certificado = f"{locator_path}/OK_Importar.png"
                    senha_certificado = f"{locator_path}/DigitarSenha.png"

                    async with async_playwright() as p:
                        browser = await p.firefox.launch(headless=False)
                        context = await browser.new_context()
                        page = await context.new_page()
                        await page.goto(
                            "https://sso.cloud.pje.jus.br/auth/realms/pje/protocol/openid-connect/auth?client_id=domicilio-eletronico-frontend&redirect_uri=https%3A%2F%2Fdomicilio-eletronico.pdpj.jus.br%2Fcomunicacoes%3Btipo%3DComunicacao&state=35cb2d31-0ded-44f2-88e0-d4a02c3b445b&response_mode=fragment&response_type=code&scope=openid&nonce=b38f5525-8e57-4aea-8fbd-e7f3cfebc9ef"
                        )
                        await page.get_by_text("Seu certificado digital").click()
                        await asyncio.sleep(3)

                        validador_site = pyautogui.locateOnScreen(
                            autorizacao_site, confidence=0.7
                        )

                        if validador_site:
                            find_and_click(sim_destavez, "Sim desta vez")

                        try:
                            while True:
                                solicitacao_senha = pyautogui.locateOnScreen(
                                    informar_senha_existe, confidence=0.7
                                )
                                if solicitacao_senha:
                                    pyautogui.click(
                                        pyautogui.locateCenterOnScreen(
                                            cancelar, confidence=0.7
                                        )
                                    )
                                    await asyncio.sleep(1)
                                else:
                                    break
                        except Exception as e:
                            pass

                        time.sleep(2)

                        certificados_pje = pyautogui.locateOnScreen(
                            certificados_disponiveis, confidence=0.7
                        )
                        if certificados_pje:
                            find_and_click(
                                engrenagem_certificado,
                                "Engrenagem, menu de configuração",
                            )
                            await asyncio.sleep(3)

                        find_and_click(tipo_certificado, "Tipo certificado...")
                        await asyncio.sleep(2)

                        try:
                            while True:
                                remover_certificado_exist = pyautogui.locateOnScreen(
                                    remover_certificado, confidence=0.7
                                )
                                if remover_certificado_exist:
                                    find_and_click(
                                        remover_certificado, "Remover Certificados"
                                    )
                                    console.print(
                                        "Certificado ja importado, removendo...\n"
                                    )
                                    await asyncio.sleep(2)
                                else:
                                    console.print(
                                        "Certificado ja importado, removidos...\n"
                                    )
                                    break
                        except Exception as e:
                            pass

                        find_and_click(localizar, "Importar certificado..")
                        find_and_click(
                            cert_fullpath,
                            "Inserir Caminho do Certificado",
                            write_text=file_path,
                        )
                        find_and_click(confirmar_certificado, "Confirmar..")

                        remover_certificado_exist = pyautogui.locateOnScreen(
                            remover_certificado, confidence=0.7
                        )
                        if remover_certificado_exist:
                            find_and_click(ok_confirmar_certificado, "OK Confirmar..")

                        await asyncio.sleep(3)

                        if informar_senha_existe:
                            find_and_click(
                                senha_certificado,
                                "Inserir Senha do certificado",
                                write_text=senha,
                            )
                            await asyncio.sleep(3)
                            find_and_click(
                                ok_confirmar_certificado, "Confirmar seleção"
                            )

                        await asyncio.sleep(5)
                        await page.wait_for_load_state("networkidle")
                        await page.wait_for_selector(
                            "text=Comunicação Processual Acompanhe as citações, intimações e ofícios direcionados",
                            timeout=15000,
                        )
                        await page.get_by_text(
                            "Comunicação Processual Acompanhe as citações, intimações e ofícios direcionados"
                        ).click()

                        await asyncio.sleep(5)
                        await page.wait_for_load_state("networkidle")
                        await page.wait_for_selector("button.buscar", timeout=15000)
                        await asyncio.sleep(3)
                        page.locator("button.buscar").click()

                        await asyncio.sleep(20)
                        await page.wait_for_load_state("networkidle")

                        # Captura todas as linhas da tabela
                        linhas = await page.query_selector_all("table tbody tr")

                        # Extrai informações de cada linha
                        for linha in linhas:
                            colunas = linha.query_selector_all("td")
                            dados_linha = [
                                coluna.inner_text().strip() for coluna in colunas
                            ]

                            if dados_linha and len(dados_linha) > 1:
                                nova_linha = [certificado, empresa] + dados_linha[1:]
                                report.append(nova_linha)

                        await page.get_by_label("Usuário Autenticado").click()
                        await page.locator("a").filter(has_text="Sair").click()

                        await asyncio.sleep(2)

                        await context.close()
                        await browser.close()

                        result = {
                            "error": False,
                            "message": f"Proceso completado com sucesso para o certificado {certificado}",
                        }
                        feedback_cnpj.append(
                            f"Processo finalizado com sucesso para o certificado: {certificado}"
                        )

                else:
                    count_error += 1
                    console.print(
                        f"Erro para o {certificado} - Certificado não encontrado no Google Drive"
                    )
                    result = {
                        "error": True,
                        "message": "Certificado não encontrado no Google Drive",
                        "certificado": certificado,
                    }
                    feedback_cnpj.append(
                        f"Certificado não encontrado no Google Drive: {certificado}"
                    )

            except Exception as e:
                count_error += 1
                console.print(f"Erro para {certificado}: {str(e)}, Mensagem: {e}")
                result = {
                    "error": True,
                    "error_message": str(e),
                    "certificate": certificado,
                }
                feedback_cnpj.append(
                    f" Não foi possivel concluir o processo para o centificado: {certificado}"
                )

        if count_error == len(certificate_files):
            log_msg = f"Erro Processo do DJE, não obtivemos exito em nenhum certificado executado, erro: {result}"
            logger.error(log_msg)
            console.print(log_msg, style="bold red")
            return RpaRetornoProcessoDTO(
                sucesso=False,
                retorno=log_msg,
                status=RpaHistoricoStatusEnum.Falha,
                tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
            )

        else:
            report_to_user("consultaDJE.xlsx", report, os.path.join(temp_dir))
            feedback_cnpj_str = "\n".join(feedback_cnpj)
            corpo_email = f""" Caros(as), \nEspero que este e-mail o encontre bem ! \n\n Abaixo, segue um resumo de facil acesso as consultas do DJE.
            \n Resumo:
            1 - Para consulta temos acesso a {len(certificate_files)} certificados e o RPA tem a responsabilidade de acessar diariamente o sistema DJE para consulta.\n
            CNPJs buscados e seus respectivos status: \n
            {feedback_cnpj_str}

            Feedback de execução em anexo:
            """
            corpo_email += (
                f"\nEstamos disponivel para quaisquer duvidas ou sugestões.\n"
            )

            try:
                send_email(
                    smtp_server,
                    smtp_port,
                    smtp_user,
                    smtp_pass,
                    corpo_email,
                    "RPA - DJE",
                    emails_to,
                    temp_dir,
                )
                log_msg = f"Processo concluido e e-mail disparado para area de negócio"
                console.print(log_msg)
                return RpaRetornoProcessoDTO(
                    sucesso=True,
                    retorno=log_msg,
                    status=RpaHistoricoStatusEnum.Sucesso,
                )
            except Exception as e:
                log_msg = f"Processo concluido porém não foi possivel enviar o e-mail para area de negóciol: {e}"
                logger.error(log_msg)
                console.print(log_msg, style="bold red")
                return RpaRetornoProcessoDTO(
                    sucesso=False,
                    retorno=log_msg,
                    status=RpaHistoricoStatusEnum.Falha,
                    tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
                )

    except Exception as e:
        log_msg = f"Erro Processo do DJE: {e}"
        logger.error(log_msg)
        console.print(log_msg, style="bold red")
        return RpaRetornoProcessoDTO(
            sucesso=False,
            retorno=log_msg,
            status=RpaHistoricoStatusEnum.Falha,
            tags=[RpaTagDTO(descricao=RpaTagEnum.Tecnico)],
        )

    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

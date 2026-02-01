import dz_lib.bivariate.data as two_d
import dz_lib.univariate.data  as one_d
import openpyxl


def excel_to_array(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        max_rows = sheet.max_row
        max_cols = sheet.max_column
        spreadsheet_data = []
        for row in range(1, max_rows + 1):
            row_data = []
            for col in range(1, max_cols + 1):
                cell_value = sheet.cell(row=row, column=col).value
                cell_value = cell_value if cell_value is not None else None
                row_data.append(cell_value)
            spreadsheet_data.append(row_data)
        return spreadsheet_data
    except Exception as e:
        print(f"Error converting Excel file to array: {e}")
        return None


def read_1d_samples(spreadsheet_array, max_age: int=4500):
    samples = []
    num_cols = len(spreadsheet_array[0])
    
    for i in range(0, num_cols, 2):
        sample_name = str(spreadsheet_array[0][i])
        if sample_name is not None:
            grains = []
            for row_data in spreadsheet_array[1:]:
                if i < len(row_data):
                    age = row_data[i]
                    uncertainty = row_data[i + 1] if i + 1 < len(row_data) else None
                    
                    if (isinstance(age, (float, int)) and 
                        isinstance(uncertainty, (float, int)) and 
                        float(age) < max_age):
                        grains.append(one_d.Grain(float(age), float(uncertainty)))
            
            if grains:
                sample = one_d.Sample(sample_name, grains)
                samples.append(sample)
    return samples


def read_2d_samples(spreadsheet_array, max_age: int=4500):
    samples = []
    for i in range(0, len(spreadsheet_array[0]), 2):
        sample_name = str(spreadsheet_array[0][i])
        if sample_name is not None:
            grains = []
            for row_data in spreadsheet_array[1:]:
                age = row_data[i]
                if not (isinstance(age, float) or isinstance(age, int)):
                    age = None
                hafnium = row_data[i + 1] if i + 1 < len(row_data) else None
                if not (isinstance(hafnium, float) or isinstance(hafnium, int)):
                    hafnium = None
                if age is not None and hafnium is not None and float(age) < max_age:
                    grains.append(two_d.BivariateGrain(float(age), float(hafnium)))
            sample = two_d.BivariateSample(sample_name, grains)
            samples.append(sample)
    return samples

def get_x_max(samples):
    x_max = 0
    for sample in samples:
        for grain in sample.grains:
            if grain.age + grain.uncertainty > x_max:
                x_max = grain.age + grain.uncertainty
    return x_max


def get_x_min(samples):
    x_min = 0
    for sample in samples:
        for grain in sample.grains:
            if grain.age - grain.uncertainty < x_min:
                x_min = grain.age - grain.uncertainty
    return x_min
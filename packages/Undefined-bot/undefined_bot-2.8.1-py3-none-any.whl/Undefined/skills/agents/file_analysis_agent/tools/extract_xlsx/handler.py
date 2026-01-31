from pathlib import Path
from typing import Any, Dict
import logging
import csv

logger = logging.getLogger(__name__)


async def execute(args: Dict[str, Any], context: Dict[str, Any]) -> str:
    file_path: str = args.get("file_path", "")

    path = Path(file_path)

    if not path.exists():
        return f"错误：文件不存在 {file_path}"

    if not path.is_file():
        return f"错误：{file_path} 不是文件"

    try:
        info: list[str] = []
        info.append(f"文件大小：{path.stat().st_size} 字节")

        if path.suffix.lower() == ".csv":
            return await _extract_csv(path, info)
        else:
            return await _extract_excel(path, info)

    except Exception as e:
        logger.exception(f"解析 Excel/CSV 失败: {e}")
        return f"解析 Excel/CSV 失败: {e}"


async def _extract_csv(path: Path, info: list[str]) -> str:
    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            return "CSV 文件为空"

        info.append("工作表：CSV (单个)")
        info.append(f"总行数：{len(rows)}")
        if rows:
            info.append(f"列数：{len(rows[0])}")

        preview = []
        preview.append("\n--- 前 20 行预览 ---\n")
        for i, row in enumerate(rows[:20], 1):
            preview.append(f"行 {i}: {' | '.join(row)}")
        if len(rows) > 20:
            preview.append(f"\n... (共 {len(rows)} 行)")

        return "\n".join(info) + "\n" + "\n".join(preview)

    except Exception as e:
        return f"解析 CSV 失败: {e}"


async def _extract_excel(path: Path, info: list[str]) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)

        sheet_names = wb.sheetnames
        info.append(f"工作表数量：{len(sheet_names)}")
        info.append(f"工作表列表：{', '.join(sheet_names)}")

        all_content = []

        for sheet_name in sheet_names[:5]:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(max_row=100, values_only=True))

            if not rows:
                continue

            max_cols = max(len(row) for row in rows if row)
            info.append(f"\n工作表 '{sheet_name}':")
            info.append(f"  行数（前 100 行）: {len(rows)}")
            info.append(f"  列数: {max_cols}")

            preview = []
            preview.append(f"  --- {sheet_name} 前 15 行预览 ---\n")
            for i, row in enumerate(rows[:15], 1):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                preview.append(f"    行 {i}: {' | '.join(row_values)}")
            if len(rows) > 15:
                preview.append(f"    ... (共 {len(rows)} 行)")
            all_content.append("\n".join(preview))

        if len(sheet_names) > 5:
            info.append(f"\n... (还有 {len(sheet_names) - 5} 个工作表)")

        return "\n".join(info) + "\n" + "\n".join(all_content)

    except Exception as e:
        logger.exception(f"解析 Excel 失败: {e}")
        return f"解析 Excel 失败: {e}"

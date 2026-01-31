"""
/export command - Export to different formats

Works offline by reading local JSON files.
"""

import json
import csv
from datetime import datetime
from pathlib import Path
from typing import Optional

from rich.console import Console

from codrsync.config import require_project_context, ProjectContext
from codrsync.i18n import t


console = Console()


def run(format: str = "excel", output: str = "exports"):
    """Run export command"""
    ctx = require_project_context()
    output_dir = Path(output)
    output_dir.mkdir(exist_ok=True)

    format_lower = format.lower()

    if format_lower == "excel":
        export_excel(ctx, output_dir)
    elif format_lower == "jira":
        export_jira(ctx, output_dir)
    elif format_lower == "trello":
        export_trello(ctx, output_dir)
    elif format_lower == "notion":
        export_notion(ctx, output_dir)
    elif format_lower == "json":
        export_json(ctx, output_dir)
    elif format_lower == "all":
        export_excel(ctx, output_dir)
        export_jira(ctx, output_dir)
        export_trello(ctx, output_dir)
        export_notion(ctx, output_dir)
        export_json(ctx, output_dir)
    else:
        console.print(f"{t('common.error')} {t('export.unknown_format', format=format)}")
        console.print(t("export.available_formats"))


def export_excel(ctx: ProjectContext, output_dir: Path):
    """Export to Excel with multiple tabs"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        console.print(f"{t('common.error')} {t('export.openpyxl_missing')}")
        return

    wb = Workbook()

    # Dashboard tab
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Project"
    ws["B1"] = ctx.project_name
    ws["A2"] = "Phase"
    ws["B2"] = ctx.phase
    ws["A3"] = "Progress"
    ws["B3"] = f"{ctx.overall_progress}%"
    ws["A4"] = "Generated"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Epics tab
    ws_epics = wb.create_sheet("Epics")
    ws_epics.append(["ID", "Name", "Status", "Progress", "Dependencies"])
    for epic in ctx.progress.get("epics", []):
        ws_epics.append([
            epic.get("id", ""),
            epic.get("name", ""),
            epic.get("status", ""),
            epic.get("progress", 0),
            ", ".join(epic.get("dependencies", []))
        ])

    # Stories tab
    ws_stories = wb.create_sheet("Stories")
    ws_stories.append(["ID", "Epic", "Title", "Status", "Points"])
    for story in ctx.progress.get("stories", []):
        ws_stories.append([
            story.get("id", ""),
            story.get("epic", ""),
            story.get("title", ""),
            story.get("status", ""),
            story.get("points", 0)
        ])

    # Milestones tab
    ws_milestones = wb.create_sheet("Milestones")
    ws_milestones.append(["ID", "Name", "Date", "Status"])
    for m in ctx.manifest.get("milestones", []):
        ws_milestones.append([
            m.get("id", ""),
            m.get("name", ""),
            m.get("date", ""),
            m.get("status", "")
        ])

    # PRPs tab
    ws_prps = wb.create_sheet("PRPs")
    ws_prps.append(["ID", "Name", "Status", "Decisions", "Approved"])
    for prp in ctx.manifest.get("prps", []):
        ws_prps.append([
            prp.get("id", ""),
            prp.get("name", ""),
            prp.get("status", ""),
            prp.get("decisions_count", ""),
            prp.get("approved_at", "")
        ])

    # Metrics tab
    ws_metrics = wb.create_sheet("Metrics")
    metrics = ctx.progress.get("metrics", {})
    ws_metrics.append(["Metric", "Value"])
    ws_metrics.append(["Average Velocity", metrics.get("average_velocity", 0)])
    ws_metrics.append(["Total Points Delivered", metrics.get("total_points_delivered", 0)])
    ws_metrics.append(["Sprints Completed", metrics.get("total_sprints_completed", 0)])
    ws_metrics.append(["Decisions Made", metrics.get("decisions_made", 0)])

    # Save
    filename = output_dir / f"{ctx.manifest['project']['slug']}-report.xlsx"
    wb.save(filename)
    console.print(f"{t('common.success')} {t('export.excel_exported', filename=filename)}")


def export_jira(ctx: ProjectContext, output_dir: Path):
    """Export to Jira CSV format"""
    filename = output_dir / "jira-import.csv"

    with open(filename, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Summary", "Description", "Issue Type", "Priority",
            "Epic Link", "Story Points", "Labels"
        ])

        # Export epics
        for epic in ctx.progress.get("epics", []):
            writer.writerow([
                f"[EPIC] {epic.get('name', '')}",
                f"Epic from PRP: {epic.get('prp', '')}",
                "Epic",
                "High",
                "",
                "",
                "codrsync"
            ])

        # Export stories
        for story in ctx.progress.get("stories", []):
            writer.writerow([
                story.get("title", ""),
                f"Story from epic {story.get('epic', '')}",
                "Story",
                "Medium",
                story.get("epic", ""),
                story.get("points", ""),
                "codrsync"
            ])

    console.print(f"{t('common.success')} {t('export.jira_exported', filename=filename)}")
    console.print(f"  {t('export.jira_import_hint')}")


def export_trello(ctx: ProjectContext, output_dir: Path):
    """Export to Trello JSON format"""
    board = {
        "name": ctx.project_name,
        "lists": [
            {"name": "Backlog", "cards": []},
            {"name": "To Do", "cards": []},
            {"name": "In Progress", "cards": []},
            {"name": "Review", "cards": []},
            {"name": "Done", "cards": []}
        ],
        "labels": [
            {"name": "Epic", "color": "purple"},
            {"name": "Story", "color": "blue"},
            {"name": "High Priority", "color": "red"}
        ]
    }

    # Add epics to backlog
    for epic in ctx.progress.get("epics", []):
        card = {
            "name": f"[EPIC] {epic.get('name', '')}",
            "desc": f"Progress: {epic.get('progress', 0)}%",
            "labels": ["Epic"]
        }
        board["lists"][0]["cards"].append(card)

    # Add stories based on status
    status_list_map = {
        "pending": 0, "todo": 1, "in_progress": 2,
        "review": 3, "done": 4
    }

    for story in ctx.progress.get("stories", []):
        card = {
            "name": story.get("title", ""),
            "desc": f"Epic: {story.get('epic', '')}\nPoints: {story.get('points', '')}",
            "labels": ["Story"]
        }
        list_idx = status_list_map.get(story.get("status", "pending"), 0)
        board["lists"][list_idx]["cards"].append(card)

    filename = output_dir / "trello-board.json"
    with open(filename, "w") as f:
        json.dump(board, f, indent=2)

    console.print(f"{t('common.success')} {t('export.trello_exported', filename=filename)}")
    console.print(f"  {t('export.trello_import_hint')}")


def export_notion(ctx: ProjectContext, output_dir: Path):
    """Export to Notion Markdown format"""
    filename = output_dir / "notion-database.md"

    with open(filename, "w") as f:
        f.write(f"# {ctx.project_name} - Backlog\n\n")

        f.write("## Database Properties\n\n")
        f.write("| Property | Type | Options |\n")
        f.write("|----------|------|---------|\n")
        f.write("| Status | Select | Backlog, To Do, In Progress, Review, Done |\n")
        f.write("| Type | Select | Epic, Story, Task |\n")
        f.write("| Priority | Select | Critical, High, Medium, Low |\n")
        f.write("| Points | Number | |\n")
        f.write("| Epic | Relation | |\n")
        f.write("\n---\n\n")

        # Epics
        f.write("## Epics\n\n")
        for epic in ctx.progress.get("epics", []):
            f.write(f"### {epic['id']}: {epic.get('name', '')}\n")
            f.write(f"- **Status**: {epic.get('status', 'pending')}\n")
            f.write(f"- **Progress**: {epic.get('progress', 0)}%\n")
            f.write("\n")

        # Stories
        f.write("## Stories\n\n")
        for story in ctx.progress.get("stories", []):
            f.write(f"### {story['id']}: {story.get('title', '')}\n")
            f.write(f"- **Epic**: {story.get('epic', '')}\n")
            f.write(f"- **Status**: {story.get('status', 'pending')}\n")
            f.write(f"- **Points**: {story.get('points', '')}\n")
            f.write("\n")

    console.print(f"{t('common.success')} {t('export.notion_exported', filename=filename)}")
    console.print(f"  {t('export.notion_import_hint')}")


def export_json(ctx: ProjectContext, output_dir: Path):
    """Export full project data as JSON"""
    data = {
        "meta": {
            "version": "1.0",
            "exported_at": datetime.now().isoformat(),
            "tool": "codrsync"
        },
        "project": ctx.manifest.get("project", {}),
        "developer": ctx.manifest.get("developer", {}),
        "prps": ctx.manifest.get("prps", []),
        "milestones": ctx.manifest.get("milestones", []),
        "current_sprint": ctx.manifest.get("current_sprint"),
        "progress": ctx.progress
    }

    filename = output_dir / "project-data.json"
    with open(filename, "w") as f:
        json.dump(data, f, indent=2)

    console.print(f"{t('common.success')} {t('export.json_exported', filename=filename)}")

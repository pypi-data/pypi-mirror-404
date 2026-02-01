"""Write MIHCSME metadata to Excel format."""

from pathlib import Path
from typing import Dict, List, Any, Union, BinaryIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from mihcsme_py.models import MIHCSMEMetadata


def write_metadata_to_excel(
    metadata: MIHCSMEMetadata, output_path: Union[Path, BinaryIO]
) -> None:
    """
    Write MIHCSME metadata to Excel file.

    :param metadata: MIHCSMEMetadata object to export
    :param output_path: Path to output Excel file, or a file-like object (e.g., BytesIO)
    """
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Write Investigation Information
    if metadata.investigation_information and metadata.investigation_information.groups:
        _write_grouped_sheet(
            wb,
            "InvestigationInformation",
            metadata.investigation_information.groups,
            header_comment="# Investigation Information - Metadata about the overall investigation"
        )

    # Write Study Information
    if metadata.study_information and metadata.study_information.groups:
        _write_grouped_sheet(
            wb,
            "StudyInformation",
            metadata.study_information.groups,
            header_comment="# Study Information - Metadata about the study design"
        )

    # Write Assay Information
    if metadata.assay_information and metadata.assay_information.groups:
        _write_grouped_sheet(
            wb,
            "AssayInformation",
            metadata.assay_information.groups,
            header_comment="# Assay Information - Metadata about the assay protocol"
        )

    # Write Assay Conditions
    if metadata.assay_conditions:
        _write_assay_conditions(wb, metadata.assay_conditions)

    # Write Reference Sheets
    for ref_sheet in metadata.reference_sheets:
        _write_reference_sheet(wb, ref_sheet.name, ref_sheet.data)

    # Save workbook
    wb.save(output_path)


def _write_grouped_sheet(
    wb: Workbook,
    sheet_name: str,
    groups: Dict[str, Dict[str, str]],
    header_comment: str = None
) -> None:
    """
    Write a grouped metadata sheet (Investigation/Study/Assay Information).

    :param wb: Workbook object
    :param sheet_name: Name of the sheet
    :param groups: Dictionary of groups {group_name: {key: value}}
    :param header_comment: Optional comment to add at the top
    """
    ws = wb.create_sheet(sheet_name)

    # Add header comment if provided
    row_num = 1
    if header_comment:
        ws.cell(row=row_num, column=1, value=header_comment)
        ws.cell(row=row_num, column=1).font = Font(italic=True, color="808080")
        row_num += 1

    # Add column headers
    ws.cell(row=row_num, column=1, value="Group")
    ws.cell(row=row_num, column=2, value="Key")
    ws.cell(row=row_num, column=3, value="Value")

    # Style headers
    for col in range(1, 4):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row_num += 1

    # Write data
    for group_name, group_data in groups.items():
        for key, value in group_data.items():
            ws.cell(row=row_num, column=1, value=group_name)
            ws.cell(row=row_num, column=2, value=key)
            ws.cell(row=row_num, column=3, value=value)
            row_num += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50


def _write_assay_conditions(wb: Workbook, assay_conditions: List[Any]) -> None:
    """
    Write AssayConditions sheet.

    :param wb: Workbook object
    :param assay_conditions: List of AssayCondition objects
    """
    ws = wb.create_sheet("AssayConditions")

    # Add header comment
    ws.cell(row=1, column=1, value="# Assay Conditions - Per-well metadata")
    ws.cell(row=1, column=1).font = Font(italic=True, color="808080")

    # Collect all unique condition keys
    all_keys = set()
    for condition in assay_conditions:
        all_keys.update(condition.conditions.keys())

    # Sort keys for consistent ordering
    condition_keys = sorted(all_keys)

    # Write headers
    headers = ["Plate", "Well"] + condition_keys
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write data
    for row_idx, condition in enumerate(assay_conditions, start=3):
        ws.cell(row=row_idx, column=1, value=condition.plate)
        ws.cell(row=row_idx, column=2, value=condition.well)

        for col_idx, key in enumerate(condition_keys, start=3):
            value = condition.conditions.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    for col_idx in range(3, 3 + len(condition_keys)):
        ws.column_dimensions[chr(64 + col_idx)].width = 20


def _write_reference_sheet(wb: Workbook, sheet_name: str, data: Dict[str, Any]) -> None:
    """
    Write a reference sheet (sheets starting with _).

    :param wb: Workbook object
    :param sheet_name: Name of the reference sheet
    :param data: Dictionary of key-value pairs
    """
    # Ensure sheet name starts with underscore
    if not sheet_name.startswith('_'):
        sheet_name = f'_{sheet_name}'

    ws = wb.create_sheet(sheet_name)

    if not data:
        # Empty reference sheet
        ws.cell(row=1, column=1, value="# Empty reference sheet")
        return

    # Write headers
    headers = ["Key", "Value"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write data as key-value pairs
    for row_idx, (key, value) in enumerate(data.items(), start=2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

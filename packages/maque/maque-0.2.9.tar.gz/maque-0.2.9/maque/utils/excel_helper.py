# !/usr/bin/env python3
# -*- coding: utf-8 -*-
"""读取Excel文件，提取图片并将其与表格行对齐"""

import pandas as pd
import os
import hashlib
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import glob


def extract_excel_with_images(
    excel_path,
    image_column_names,  # 改为支持多列：可以是字符串（单列）或列表（多列）
    output_column_name=None,  # 输出合并后的列名，如果为None则使用第一个图像列名
    image_output_dir="extracted_images",
    sheet_name=0,
    use_hash_filename=False,
    save_updated_excel=False,
    image_output_dir_prefix=False,
    use_global_image_folder=False,  # 是否使用全局统一图像文件夹
    dtype_columns=None,  # 指定列的数据类型，格式为 {'列名': 数据类型}，如 {'nid': str}
    use_absolute_path=False,  # 是否使用绝对路径保存图像路径
):
    """
    读取包含图片的Excel文件，提取图片并将其与表格行对齐。

    参数：
    - excel_path: Excel文件路径
    - image_column_names: 图片所在的列名称，可以是字符串（单列）或列表（多列）
    - output_column_name: 输出合并后的列名，如果为None则使用第一个图像列名
    - image_output_dir: 图片保存路径
    - sheet_name: 指定读取的工作表（默认读取第一个）
    - use_hash_filename: 是否使用哈希值作为文件名（默认为False）
    - save_updated_excel: 是否保存更新后的Excel文件（默认为False）
    - image_output_dir_prefix: 图片输出目录前缀, 如果为False，则不使用前缀, 只有文件名，不包含路径
    - use_global_image_folder: 是否使用全局统一图像文件夹（默认为False，每个Excel单独文件夹）
    - dtype_columns: 指定列的数据类型，格式为 {'列名': 数据类型}，如 {'nid': str}（默认为None）
    - use_absolute_path: 是否使用绝对路径保存图像路径（默认为False，使用相对路径）

    返回：
    - 一个带有图片路径的新DataFrame
    """
    # 获取Excel文件名（不含路径和扩展名）
    excel_basename = os.path.splitext(os.path.basename(excel_path))[0]

    # 加载 Excel 文件
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.active

    # 如果sheet_name为None或数字，获取当前活动的工作表名
    actual_sheet_name = sheet_name if isinstance(sheet_name, str) else ws.title

    # 根据use_global_image_folder参数决定图片输出目录结构
    if use_global_image_folder:
        # 使用全局统一图像文件夹
        specific_image_dir = image_output_dir
    else:
        # 创建特定于此Excel的图片输出目录
        # 如果只有一个工作表，不创建额外的工作表子目录
        if len(wb.worksheets) == 1:
            specific_image_dir = os.path.join(image_output_dir, excel_basename)
        else:
            specific_image_dir = os.path.join(
                image_output_dir, excel_basename, actual_sheet_name
            )
    os.makedirs(specific_image_dir, exist_ok=True)

    # 使用openpyxl直接读取数据
    data = []
    headers = []

    # 获取表头
    for cell in ws[1]:
        headers.append(cell.value)

    # 创建列名到索引的映射，用于快速查找需要特殊处理的列
    col_to_index = {col_name: idx for idx, col_name in enumerate(headers)}
    force_string_columns = set()

    # 识别需要强制为字符串的列
    if dtype_columns:
        for col_name, col_dtype in dtype_columns.items():
            if col_name in col_to_index and col_dtype == str:
                force_string_columns.add(col_to_index[col_name])
                print(
                    f"列 '{col_name}' (索引 {col_to_index[col_name]}) 将强制读取为字符串类型"
                )

    # 获取所有数据行
    for row in ws.iter_rows(min_row=2):  # 从第二行开始（跳过表头）
        row_data = []
        for col_idx, cell in enumerate(row):
            if col_idx in force_string_columns:
                # 对于需要强制为字符串的列，直接转换为字符串（保持原始格式）
                row_data.append(str(cell.value) if cell.value is not None else "")
            else:
                # 对于其他列，保持原始值
                row_data.append(cell.value if cell.value is not None else None)
        data.append(row_data)

    # 创建DataFrame
    df = pd.DataFrame(data, columns=headers)

    # 应用其他（非字符串）的数据类型
    if dtype_columns:
        for col_name, col_dtype in dtype_columns.items():
            if (
                col_name in df.columns and col_dtype != str
            ):  # 字符串类型已经在读取时处理了
                try:
                    df[col_name] = df[col_name].astype(col_dtype)

                    print(f"列 '{col_name}' 已设置为 {col_dtype} 类型")
                except Exception as e:
                    print(f"设置列 '{col_name}' 为 {col_dtype} 类型时出错: {e}")
            elif col_name not in df.columns:
                print(f"警告: 列 '{col_name}' 不存在，跳过类型设置")
    # 打印DataFrame信息
    print(f"DataFrame信息:")
    print(f"- 总行数: {len(df)}")
    if dtype_columns:
        print(f"- 已应用的数据类型: {dtype_columns}")
    # print(f"- 列名: {list(df.columns)}")
    # print(f"- 数据类型:\n{df.dtypes}")

    # 处理图像列名参数：支持单列（字符串）和多列（列表）
    if isinstance(image_column_names, str):
        image_columns_list = [image_column_names]
    else:
        image_columns_list = image_column_names

    # 确定输出列名
    if output_column_name is None:
        output_column_name = image_columns_list[0]

    # 确保指定的列名都存在
    cols = list(df.columns)
    image_columns = {}  # column_name: column_index

    for col_name in image_columns_list:
        if col_name not in df.columns:
            raise ValueError(f"列名 '{col_name}' 在Excel文件中不存在")
        image_columns[col_name] = cols.index(col_name) + 1

    # 如果有多列图像，需要准备合并列
    if len(image_columns_list) > 1 and output_column_name not in df.columns:
        # 添加新的输出列
        df[output_column_name] = ""

    # 提取图片和位置
    image_map = {}  # row_num: [image_filenames_list]
    print(f"开始提取图片，总行数: {len(df)}")
    print(f"目标图像列: {list(image_columns.keys())}")

    for img in ws._images:
        anchor = img.anchor._from  # 起始锚点
        row = anchor.row + 1
        col = anchor.col + 1

        # 检查是否在任何目标图像列中
        target_col_name = None
        for col_name, col_index in image_columns.items():
            if col == col_index:
                target_col_name = col_name
                break

        if target_col_name is None:
            continue  # 忽略不在目标列的图片

        if use_hash_filename:
            # 使用图片内容生成哈希值作为文件名
            img_hash = hashlib.md5(img.ref.getvalue()).hexdigest()
            filename = f"{img_hash}.png"
        else:
            # 包含Excel文件名以避免全局文件夹中的命名冲突
            if use_global_image_folder:
                filename = f"{excel_basename}_r{row}_c{col}_{target_col_name}.png"
            else:
                filename = f"image_r{row}_c{col}_{target_col_name}.png"

        save_path = os.path.join(specific_image_dir, filename)
        # 保存图片
        with open(save_path, "wb") as img_file:
            img_file.write(img.ref.getvalue())

        # 根据 use_absolute_path 和 image_output_dir_prefix 参数决定保存的路径格式
        if use_absolute_path:
            # 使用绝对路径
            img_path = os.path.abspath(save_path)
        elif image_output_dir_prefix:
            # 使用相对路径
            img_path = os.path.join(specific_image_dir, filename)
        else:
            # 只使用文件名
            img_path = filename

        # 将图片路径添加到对应行
        if row not in image_map:
            image_map[row] = []
        image_map[row].append(img_path)

    total_images = sum(len(paths) for paths in image_map.values())
    print(f"找到的图片数量: {total_images}，涉及行数: {len(image_map)}")

    # 添加图片路径到输出列，多个图片用\n分割
    for row_num, img_paths in image_map.items():
        df_index = row_num - 2  # 减2：1行表头 + 1行从1开始变成从0
        if 0 <= df_index < len(df):
            # 将多个图片路径用\n连接
            combined_paths = "\n\n".join(img_paths)
            df.loc[df_index, output_column_name] = combined_paths

    # 如果使用了多列图像且创建了新的输出列，清空原始图像列（除了输出列）
    if len(image_columns_list) > 1:
        for col_name in image_columns_list:
            if col_name != output_column_name:
                df[col_name] = ""

    # 生成更新后的Excel文件
    if save_updated_excel:
        output_excel_path = os.path.join(
            os.path.dirname(excel_path),
            f"{excel_basename}_updated{os.path.splitext(excel_path)[1]}",
        )
        df.to_excel(
            output_excel_path,
            sheet_name=actual_sheet_name,
            index=False,
            engine="openpyxl",
        )
        print(f"更新后的Excel文件已保存到: {output_excel_path}")

    return df


def deduplicate_columns(df):
    cols = pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique():
        cols[cols[cols == dup].index.values.tolist()] = [
            dup + "." + str(i) if i != 0 else dup for i in range(sum(cols == dup))
        ]
    df.columns = cols
    return df


def process_directory(
    dir_path,
    output_dir_suffix,
    image_columns=["图片"],
    output_column="图片",
    use_global_folder=False,
    merge_excel_output=False,
    dtype_columns=None,
):
    """处理指定目录下的所有Excel文件

    参数：
    - dir_path: 目录路径
    - output_dir_suffix: 输出目录后缀
    - image_columns: 图像列名列表，默认为["图片"]
    - output_column: 输出列名，默认为"图片"
    - use_global_folder: 是否使用全局统一图像文件夹
    - merge_excel_output: 是否合并所有Excel输出为一个文件
    - dtype_columns: 指定列的数据类型，格式为 {'列名': 数据类型}，如 {'nid': str}（默认为None）
    """
    excel_files = glob.glob(os.path.join(dir_path, "*.xlsx"))
    processed_dataframes = []

    for excel_file in excel_files:
        print(f"正在处理: {excel_file}")

        # 跳过已经更新过的文件
        if "_updated" in excel_file:
            print(f"跳过已更新的文件: {excel_file}")
            continue

        try:
            # 当需要合并Excel输出时，不保存单独的Excel文件
            save_individual = not merge_excel_output

            df = extract_excel_with_images(
                excel_path=excel_file,
                image_column_names=image_columns,
                output_column_name=output_column,
                image_output_dir=f"{output_dir_suffix}-images",
                use_hash_filename=True,
                save_updated_excel=save_individual,
                use_global_image_folder=use_global_folder,
                dtype_columns=dtype_columns,
            )
            df = deduplicate_columns(df)

            if merge_excel_output:
                # 添加源文件名列以便追踪数据来源
                excel_basename = os.path.splitext(os.path.basename(excel_file))[0]
                df["source_file"] = excel_basename
                processed_dataframes.append(df)

            print(f"完成处理: {excel_file}")
        except Exception as e:
            print(f"处理文件 {excel_file} 时出错: {str(e)}")

    # 如果需要合并Excel输出
    if merge_excel_output and processed_dataframes:
        print(f"\n开始合并 {len(processed_dataframes)} 个Excel文件...")

        # 合并所有DataFrame
        merged_df = pd.concat(processed_dataframes, ignore_index=True)

        # 保存合并后的CSV文件
        merged_csv_path = os.path.join(dir_path, f"{output_dir_suffix}_merged.csv")
        merged_df.to_csv(merged_csv_path, index=False, encoding="utf-8")

        print(f"合并后的CSV文件已保存到: {merged_csv_path}")
        print(f"合并数据总行数: {len(merged_df)}")
        print(f"包含源文件: {list(merged_df['source_file'].unique())}")


def insert_images_to_excel(
    excel_path,
    image_column_name,
    images_dir="images",
    sheet_name=0,
    image_row_start=2,
    image_width=100,
    image_height=100,
    create_if_not_exists=True,
    image_file_patterns=None,  # 图像文件匹配模式，如 ["*.jpg", "*.png", "*.jpeg"]
    sort_images=True,  # 是否对图像文件进行排序
):
    """
    将本地图像文件插入到Excel文件的指定列中。
    
    参数：
    - excel_path: Excel文件路径
    - image_column_name: 要插入图像的列名
    - images_dir: 图像文件夹路径，默认为"images"
    - sheet_name: 指定工作表（默认第一个）
    - image_row_start: 开始插入图像的行号，默认从第2行开始（跳过表头）
    - image_width: 图像宽度，像素，默认100
    - image_height: 图像高度，像素，默认100
    - create_if_not_exists: 如果Excel文件不存在是否创建新文件，默认True
    - image_file_patterns: 图像文件匹配模式，默认为常见图像格式
    - sort_images: 是否对图像文件进行排序，默认True
    
    返回：
    - 插入图像的数量
    """
    # 默认图像文件匹配模式
    if image_file_patterns is None:
        image_file_patterns = ["*.jpg", "*.jpeg", "*.png", "*.gif", "*.bmp"]
    
    # 检查图像文件夹是否存在
    if not os.path.exists(images_dir):
        raise ValueError(f"图像文件夹不存在: {images_dir}")
    
    # 获取所有符合条件的图像文件
    image_files = []
    for pattern in image_file_patterns:
        image_files.extend(glob.glob(os.path.join(images_dir, pattern)))
    
    if sort_images:
        image_files.sort()
    
    if not image_files:
        print(f"在 {images_dir} 目录中未找到符合条件的图像文件")
        return 0
    
    print(f"找到 {len(image_files)} 个图像文件: {[os.path.basename(f) for f in image_files]}")
    
    # 处理Excel文件
    if os.path.exists(excel_path):
        # 加载现有Excel文件
        wb = load_workbook(excel_path)
        print(f"加载现有Excel文件: {excel_path}")
    else:
        if create_if_not_exists:
            # 创建新的Excel文件
            wb = Workbook()
            print(f"创建新Excel文件: {excel_path}")
        else:
            raise ValueError(f"Excel文件不存在且不允许创建: {excel_path}")
    
    # 获取工作表
    if isinstance(sheet_name, str):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            print(f"创建新工作表: {sheet_name}")
    else:
        ws = wb.active
    
    # 确保有表头行
    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        # 创建简单的表头
        ws.cell(1, 1, "ID")
        if image_column_name != "ID":
            ws.cell(1, 2, image_column_name)
            target_col = 2
        else:
            target_col = 1
    else:
        # 查找目标列或创建新列
        target_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(1, col).value == image_column_name:
                target_col = col
                break
        
        if target_col is None:
            # 在最后添加新列
            target_col = ws.max_column + 1
            ws.cell(1, target_col, image_column_name)
    
    print(f"将图像插入到第 {target_col} 列 '{image_column_name}'")
    
    # 插入图像
    inserted_count = 0
    current_row = image_row_start
    
    for image_file in image_files:
        try:
            # 创建图像对象
            img = Image(image_file)
            
            # 调整图像尺寸
            img.width = image_width
            img.height = image_height
            
            # 获取单元格位置
            cell_address = ws.cell(current_row, target_col).coordinate
            
            # 插入图像到单元格
            img.anchor = cell_address
            ws.add_image(img)
            
            # 调整行高以适应图像
            row_height_points = image_height * 0.75  # 像素转点数的近似换算
            if ws.row_dimensions[current_row].height is None or ws.row_dimensions[current_row].height < row_height_points:
                ws.row_dimensions[current_row].height = row_height_points
            
            # 调整列宽以适应图像
            column_letter = ws.cell(current_row, target_col).column_letter
            column_width_chars = image_width / 7  # 像素转字符宽度的近似换算
            if ws.column_dimensions[column_letter].width is None or ws.column_dimensions[column_letter].width < column_width_chars:
                ws.column_dimensions[column_letter].width = column_width_chars
            
            # 在同一行添加图像文件名（可选）
            if target_col > 1:  # 如果不是第一列，在第一列添加文件名
                ws.cell(current_row, 1, os.path.splitext(os.path.basename(image_file))[0])
            
            print(f"插入图像 {os.path.basename(image_file)} 到第 {current_row} 行")
            inserted_count += 1
            current_row += 1
            
        except Exception as e:
            print(f"插入图像 {image_file} 时出错: {str(e)}")
    
    # 保存Excel文件
    try:
        wb.save(excel_path)
        print(f"Excel文件已保存: {excel_path}")
        print(f"成功插入 {inserted_count} 个图像")
    except Exception as e:
        print(f"保存Excel文件时出错: {str(e)}")
        return 0
    
    return inserted_count


def create_image_excel_from_folder(
    images_dir="images",
    output_excel_path="images_output.xlsx",
    image_column_name="图片",
    image_width=150,
    image_height=150,
):
    """
    从图像文件夹创建包含所有图像的Excel文件的便利函数。
    
    参数：
    - images_dir: 图像文件夹路径，默认为"images"
    - output_excel_path: 输出Excel文件路径，默认为"images_output.xlsx"
    - image_column_name: 图像列名，默认为"图片"
    - image_width: 图像宽度，像素，默认150
    - image_height: 图像高度，像素，默认150
    
    返回：
    - 插入图像的数量
    
    示例用法：
    >>> count = create_image_excel_from_folder("images", "test_output.xlsx", "图片", 120, 120)
    >>> print(f"成功插入了 {count} 个图像")
    """
    return insert_images_to_excel(
        excel_path=output_excel_path,
        image_column_name=image_column_name,
        images_dir=images_dir,
        image_width=image_width,
        image_height=image_height,
        create_if_not_exists=True,
    )


if __name__ == "__main__":
    # 配置示例
    directories = [
        # (目录名, 输出前缀, 图像列列表, 输出列名, 是否使用全局文件夹, 是否合并Excel输出)
        (
            "origin",
            "target",
            ["黑图索引1", "黑图索引2", "黑图索引3"],
            "image_urls",
            True,
            True,
        ),
        # 更多示例：
        # ("origin", "target", ["图片"], "图片", False, False),  # 单列，每个Excel独立文件夹，不合并
        # ("origin", "target", ["图片1", "图片2"], "合并图片", True, True),  # 多列，全局文件夹，合并Excel
    ]

    for (
        dir_name,
        output_prefix,
        img_cols,
        out_col,
        global_folder,
        merge_excel,
    ) in directories:
        print(f"\n开始处理目录: {dir_name}")
        print(
            f"图像列: {img_cols}，输出列: {out_col}，全局文件夹: {global_folder}，合并Excel: {merge_excel}"
        )
        process_directory(
            dir_name, output_prefix, img_cols, out_col, global_folder, merge_excel
        )
        print(f"完成处理目录: {dir_name}")

    print("\n所有文件处理完成！")

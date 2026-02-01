from io import BytesIO

import pandas as pd
from django.core.files import File
from django.db.models import FileField
from openpyxl.styles import Font, Border, Side
from pandas.api.types import is_datetime64_any_dtype as is_datetime


class XLSXField(FileField):
    max_length = 300

    cell_format = '#,##0.00 ;[Red]-#,##0.00 ;_-* "-"??_-'
    cell_format_without_color = '#,##0.00 ;-#,##0.00 ;_-* "-"??_-'
    boolean_format = '[Green]"TRUE";[Red]"FALSE";[Red]"FALSE";[Red]"FALSE"'


    def get_number_of_rows_to_insert(self, sheet, index_len):
        max_len = 0
        for column_num in range(index_len + 1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=column_num)  # Adjust the row number to the row where you want to start splitting
            if cell.value:  # Check if the cell is not empty
                split_values = cell.value.split(".")
                max_len = max(max_len, len(split_values))

        return max_len

    def insert_rows_before_first_row(self, sheet, num_rows):
        sheet.insert_rows(1, amount=num_rows)

    def split_entries_in_sheet(self, sheet, number_of_inserted_rows, index_len):
        for column_num in range(index_len + 1, sheet.max_column + 1):  # Column F starts at index 6 (1-indexed)
            cell = sheet.cell(row=number_of_inserted_rows+1, column=column_num)  # Adjust the row number to the row where you want to start splitting
            if cell.value:  # Check if the cell is not empty
                split_values = cell.value.split(".")  # Split the entry at every "."
                first_row = 1  # First row to fill the split values
                for idx, split_value in enumerate(split_values):
                    row_num = first_row + idx
                    sheet.cell(row=row_num, column=column_num, value=split_value)
                    # Bold the cell and add outside borders
                    cell_to_format = sheet.cell(row=row_num, column=column_num)
                    cell_to_format.font = Font(bold=True)
                    cell_to_format.border = Border(top=Side(border_style='thin'),
                                                   bottom=Side(border_style='thin'),
                                                   left=Side(border_style='thin'),
                                                   right=Side(border_style='thin'))

    def create_pivotable_row(self, sheet, index_len, number_of_rows_to_be_inserted, range_of_pivot_concatenation=None):
        for column_num in range(index_len + 1, sheet.max_column + 1):
            concatenated_value = ""
            for row_num in range_of_pivot_concatenation:
                cell = sheet.cell(row=row_num, column=column_num)
                if cell.value:
                    concatenated_value += cell.value + " "
            sheet.cell(row=number_of_rows_to_be_inserted+1, column=column_num, value=concatenated_value.strip())

    def create_excel_file_from_dfs(self, path, data_frames, sheet_names=None, merge_cells=False, formats={}, comments={}, index=True, ranges_of_pivot_concatenation={'default': None}):
        """
        :param path: file_path including file_name; if relative, will be saved under self.to+path
        :param data_frames: list of dataframes that will be inserted into an excel tab each
        :param sheet_names: list of sheet names corresponding to the data_frames
        :rtype: None
        """
        if sheet_names is None:
            sheet_names = ['Sheet']
        excel_file = BytesIO()
        writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
        df: pd.DataFrame
        idx_length = 0
        for df, sheet_name in zip(data_frames, sheet_names):
            if df is not None:
                if len(df) == 0:
                    df = pd.concat([df, pd.DataFrame([{}])], ignore_index=True)
                if index:
                    idx_length = df.index.nlevels
                df.to_excel(writer, sheet_name=sheet_name, merge_cells=merge_cells, freeze_panes=(1, idx_length),
                            index=index)

                worksheet = writer.sheets[sheet_name]  # pull worksheet object
                worksheet_comment = comments[sheet_name] if sheet_name in comments else None
                cell_formats = {}
                for format in formats:
                    cell_formats[format] = writer.book.add_format({'num_format': formats[format]})

                if index:
                    index_frame = df.index.to_frame()
                    for idx, col in enumerate(index_frame):  # loop through all columns
                        series = index_frame[col]
                        max_len = max((
                            series.astype(str).map(len).max(),  # len of largest item
                            len(str(series.name))  # len of column name/header
                        )) + 1  # adding a little extra space
                        if is_datetime(series):
                            max_len = 22
                        worksheet.set_column(idx, idx, max_len)  # set column width

                for idx, col in enumerate(df):  # loop through all columns
                    series = df[col]
                    if worksheet_comment is not None:
                        comment = worksheet_comment[col] if col in worksheet_comment else None
                        if comment is not None:
                            worksheet.write_comment(0, idx + idx_length, comment)

                    max_len = max((
                        series.astype(str).map(len).max(),  # len of largest item
                        len(str(series.name))  # len of column name/header
                    )) + 1  # adding a little extra space
                    worksheet.set_column(idx + idx_length, idx + idx_length, max_len)  # set column width
                    # set Cell format
                    if col in formats:
                        worksheet.set_column(idx + idx_length, idx + idx_length, max_len, cell_format=cell_formats[col])
                    elif is_datetime(df[col]):
                        pass
                    else:
                        worksheet.set_column(idx + idx_length, idx + idx_length, max_len,
                                             cell_format=writer.book.add_format({'num_format': XLSXField.cell_format}))
                # Add autofilter:
                if len(df.columns) > 0:
                    worksheet.autofilter(0, 0, len(df), idx_length + len(df.columns)-1)

        writer.close()
        excel_file.seek(0)
        self.save(path, content=File(excel_file), save=False)
        
        return excel_file

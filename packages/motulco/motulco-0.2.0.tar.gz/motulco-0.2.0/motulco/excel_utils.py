import os 
import pandas as pd
import openpyxl
import time
import pythoncom
import win32com.client
from openpyxl import load_workbook


################################################################################################################################
################################################################################################################################
############################################### Carga un archivo de excel como un df ###########################################
################################################################################################################################
################################################################################################################################
def cargaExcel(rutaArchivo,sheetName, AsString=False,engine='openpyxl'):
    df = pd.read_excel(rutaArchivo, sheet_name=sheetName, engine=engine)
    if AsString:
        df=df.astype(str)
    return df

################################################################################################################################
################################################################################################################################
############################################# Ejecuta la funcion Actualizar todo de excel ######################################
################################################################################################################################
################################################################################################################################

def ActualizarTodoExcel(rutaArchivo):
    pythoncom.CoInitialize()  # Inicializa COM correctamente

    try:
        rutaArchivoAbsoluta = os.path.abspath(rutaArchivo)
        print(f"Intentando abrir el archivo: {rutaArchivoAbsoluta}")

        if not os.path.exists(rutaArchivoAbsoluta):
            print(f"⚠️ El archivo no existe en la ruta: {rutaArchivoAbsoluta}")
            return
        # Iniciar Excel y hacerlo visible
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = True  
        excel_app.DisplayAlerts = False
        # Abrir el archivo de Excel
        workbook = excel_app.Workbooks.Open(rutaArchivoAbsoluta)
        print("📂 Archivo abierto correctamente, iniciando actualización...")

        # Activar la primera hoja (por si acaso)
        workbook.Sheets(1).Activate()
        # Iniciar actualización
        workbook.RefreshAll()
        print("🔄 Actualizando conexiones y tablas de Power Query...")
        # Esperar que Excel termine la actualización y cálculos
        while excel_app.CalculationState != 0:  # 0 significa "listo"
            print("⌛ Esperando que Excel termine los cálculos...")
            time.sleep(2)
        print("✅ Actualización completada.")
        # Guardar y cerrar Excel
        print(f'📁 Archivo actualizado y cerrado: {rutaArchivoAbsoluta}')
    except Exception as e:
        print(f"❌ Error al actualizar el archivo {rutaArchivoAbsoluta}: {e}")
    finally:
        # Asegurar que Excel se cierra correctamente
        if 'excel_app' in locals():
            a=10
        pythoncom.CoUninitialize()

################################################################################################################################
################################################################################################################################
#################################### Consolida Archivos de excel cuyo nombre es semejante ######################################
################################################################################################################################
################################################################################################################################
def consolidarArchivosExcelPrefijo(NombreConElQueIniciaLosArchivos,RutaALaCarpeta,TipoArchivos='.xlsx'):
    dataframes = []
    for archivo in os.listdir(RutaALaCarpeta):
        if archivo.endswith(TipoArchivos) and archivo.startswith(NombreConElQueIniciaLosArchivos):
            ruta_archivo = os.path.join(RutaALaCarpeta, archivo)
            df = pd.read_excel(ruta_archivo)
            dataframes.append(df)    
    df_consolidado = pd.concat(dataframes, ignore_index=True)
    return df_consolidado

################################################################################################################################
################################################################################################################################
############################################## Consolida Archivos de excel en una carpeta ######################################
################################################################################################################################
################################################################################################################################
def consolidarArchivosExcelGlobal(NombreConElQueIniciaLosArchivos,RutaALaCarpeta,TipoArchivos='.xlsx'):
    dataframes = []
    for archivo in os.listdir(RutaALaCarpeta):
        if archivo.endswith(TipoArchivos):
            ruta_archivo = os.path.join(RutaALaCarpeta, archivo)
            df = pd.read_excel(ruta_archivo)
            dataframes.append(df)    
    df_consolidado = pd.concat(dataframes, ignore_index=True)
    return df_consolidado


################################################################################################################################
################################################################################################################################
################################## Funcion que retornas el arreglo de hojas visibles (no oculta) en excel    ###################
################################################################################################################################
################################################################################################################################
def obtenerHojasVisiblesExcel(RutaArchivo):
    """Obtiene solo las hojas visibles de un archivo Excel (.xlsx)."""
    try:
        wb = load_workbook(RutaArchivo, read_only=True)
        hojas_visibles = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == "visible"]
        wb.close()
        return hojas_visibles
    except Exception as e:
        print(f"Error al leer hojas de {RutaArchivo}: {e}")
        return []

################################################################################################################################
################################################################################################################################
################################## Funcion que retornas el arreglo de todas las hojas en excel    ##############################
################################################################################################################################
################################################################################################################################
def obtenerTodasLasHojasExcel(RutaArchivo):
    """Obtiene todas las hojas de un archivo Excel (.xlsx), sin importar si están visibles u ocultas."""
    try:
        wb = load_workbook(RutaArchivo, read_only=True)
        hojas = [sheet.title for sheet in wb.worksheets]  # Todas las hojas
        wb.close()
        return hojas
    except Exception as e:
        print(f"Error al leer hojas de {RutaArchivo}: {e}")
        return []
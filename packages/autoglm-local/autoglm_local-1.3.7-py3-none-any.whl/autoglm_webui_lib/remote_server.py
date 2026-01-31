#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
服务端（远程端）- 负责页面展示和配置管理

接收本地端的注册，管理多个本地端
"""

import asyncio
import csv
import io
import json
import os
import time
import uuid
from datetime import datetime
from typing import Optional, List, Dict

import httpx
import uvicorn
from fastapi import (
    FastAPI,
    WebSocket,
    WebSocketDisconnect,
    HTTPException,
    UploadFile,
    File,
    Form,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from loguru import logger

from .config_manager import ConfigManager

# 创建 FastAPI 应用
app = FastAPI(title="AutoGLM 服务端", version="1.0.0")

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 目录配置
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, "static")
DATA_DIR = os.path.join(BASE_DIR, "data")
DOWNLOADS_DIR = os.path.join(STATIC_DIR, "downloads")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(DOWNLOADS_DIR, exist_ok=True)

# 挂载静态文件
if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# 初始化配置管理器
config_manager = ConfigManager(DATA_DIR)

# ============== 本地端管理 ==============

# 存储已注册的本地端 {name: {ip, port, devices, last_heartbeat, ...}}
registered_locals: Dict[str, Dict] = {}
HEARTBEAT_TIMEOUT = 60  # 心跳超时（秒）


class LocalRegisterRequest(BaseModel):
    """本地端注册请求"""
    name: str
    ip: str
    port: int
    devices: list = []
    ios_support: bool = False
    adb_support: bool = True


@app.post("/local/register")
async def register_local(request: LocalRegisterRequest):
    """本地端注册"""
    registered_locals[request.name] = {
        "name": request.name,
        "ip": request.ip,
        "port": request.port,
        "url": f"http://{request.ip}:{request.port}",
        "devices": request.devices,
        "ios_support": request.ios_support,
        "adb_support": request.adb_support,
        "last_heartbeat": time.time(),
        "online": True,
    }
    logger.info(f"✅ 本地端注册: {request.name} ({request.ip}:{request.port})")
    return {"status": "success", "message": f"已注册: {request.name}"}


@app.post("/local/heartbeat")
async def local_heartbeat(request: LocalRegisterRequest):
    """本地端心跳"""
    if request.name in registered_locals:
        registered_locals[request.name].update({
            "ip": request.ip,
            "port": request.port,
            "url": f"http://{request.ip}:{request.port}",
            "devices": request.devices,
            "last_heartbeat": time.time(),
            "online": True,
        })
    else:
        # 自动注册
        await register_local(request)
    return {"status": "success"}


@app.get("/local/list")
async def list_locals():
    """列出所有已注册的本地端"""
    now = time.time()
    result = []
    for name, info in registered_locals.items():
        online = (now - info["last_heartbeat"]) < HEARTBEAT_TIMEOUT
        info["online"] = online
        result.append(info)
    return result


@app.delete("/local/{name}")
async def remove_local(name: str):
    """移除本地端"""
    if name in registered_locals:
        del registered_locals[name]
        return {"status": "success"}
    raise HTTPException(status_code=404, detail="本地端不存在")


def get_active_local_url() -> Optional[str]:
    """获取当前活跃的本地端 URL"""
    now = time.time()
    for name, info in registered_locals.items():
        if (now - info["last_heartbeat"]) < HEARTBEAT_TIMEOUT:
            return info["url"]
    return None


def get_local_url_for_device(device_serial: str) -> Optional[str]:
    """根据设备获取对应的本地端 URL"""
    now = time.time()
    for name, info in registered_locals.items():
        if (now - info["last_heartbeat"]) >= HEARTBEAT_TIMEOUT:
            continue
        for device in info.get("devices", []):
            if device.get("serial") == device_serial:
                return info["url"]
    return None


def get_local_url_by_name(local_name: str) -> Optional[str]:
    """根据本地端名称获取 URL"""
    now = time.time()
    if local_name in registered_locals:
        info = registered_locals[local_name]
        if (now - info["last_heartbeat"]) < HEARTBEAT_TIMEOUT:
            return info["url"]
    return None


# ============== 数据模型 ==============

class APIConfig(BaseModel):
    id: str = ""
    name: str
    base_url: str
    model: str
    api_key: str


class TaskRequest(BaseModel):
    device_id: str
    task: str
    platform: str = "android"
    wda_url: str = ""
    api_config_id: str = ""
    base_url: str = "https://open.bigmodel.cn/api/paas/v4"
    model: str = "autoglm-phone"
    api_key: str = ""
    max_steps: int = 0
    lang: str = "cn"


class BatchTaskRequest(BaseModel):
    device_id: str
    test_cases: list
    platform: str = "android"
    wda_url: str = ""
    api_config_id: str = ""
    base_url: str = "https://open.bigmodel.cn/api/paas/v4"
    model: str = "autoglm-phone"
    api_key: str = ""
    max_steps: int = 0
    lang: str = "cn"
    scenario_name: str = ""  # 场景名称，用于历史记录命名
    local_name: str = ""  # 本地端名称（多用户隔离）


# ============== 页面路由 ==============

@app.get("/")
async def root():
    return RedirectResponse(url="/static/index.html")


# ============== 代理转发到本地端 ==============

async def proxy_to_local(method: str, path: str, device_serial: str = None, local_name: str = None, **kwargs):
    """代理请求到本地端服务
    
    Args:
        method: HTTP 方法
        path: 请求路径
        device_serial: 设备序列号（用于根据设备找本地端）
        local_name: 本地端名称（优先级最高）
        **kwargs: 其他请求参数
    """
    local_url = None
    
    # 优先使用 local_name 找本地端
    if local_name:
        local_url = get_local_url_by_name(local_name)
    
    # 其次根据设备找本地端
    if not local_url and device_serial:
        local_url = get_local_url_for_device(device_serial)
    
    # 最后使用任意活跃的本地端
    if not local_url:
        local_url = get_active_local_url()
    
    if not local_url:
        raise HTTPException(status_code=503, detail="没有可用的本地端服务，请先在 Mac 上启动本地端")
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            url = f"{local_url}{path}"
            response = await client.request(method, url, **kwargs)
            return response.json()
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="本地端服务超时")
    except httpx.ConnectError:
        raise HTTPException(status_code=503, detail="无法连接到本地端服务")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============== 设备管理 ==============

@app.get("/devices")
async def get_devices(local_name: Optional[str] = None):
    """获取所有本地端的设备，可按 local_name 过滤"""
    all_devices = []
    now = time.time()
    
    for name, info in registered_locals.items():
        if (now - info["last_heartbeat"]) >= HEARTBEAT_TIMEOUT:
            continue
        
        # 如果指定了 local_name，只返回该本地端的设备
        if local_name and name != local_name:
            continue
        
        for device in info.get("devices", []):
            device["local_name"] = name
            device["local_url"] = info["url"]
            all_devices.append(device)
    
    return all_devices


@app.get("/devices/android")
async def get_android_devices(local_name: Optional[str] = None):
    devices = await get_devices(local_name=local_name)
    return [d for d in devices if d.get("platform") == "android"]


@app.get("/devices/ios")
async def get_ios_devices(local_name: Optional[str] = None):
    devices = await get_devices(local_name=local_name)
    return [d for d in devices if d.get("platform") == "ios"]


@app.get("/system/info")
async def get_system_info():
    return {
        "version": "1.0.0",
        "locals_count": len([l for l in registered_locals.values() if l.get("online")]),
        "locals": list(registered_locals.values()),
    }


@app.get("/system_check")
async def system_check():
    return {
        "locals": list(registered_locals.values()),
        "devices": await get_devices(),
    }


# ============== 截图（代理） ==============

@app.get("/screenshot/{device_id}")
async def get_screenshot(device_id: str):
    return await proxy_to_local("GET", f"/screenshot/{device_id}", device_serial=device_id)


@app.get("/remote/screen/{device_id}")
async def get_remote_screen(device_id: str):
    return await proxy_to_local("GET", f"/remote/screen/{device_id}", device_serial=device_id)


# ============== 输入法（代理） ==============

@app.get("/input_methods/{device_id}")
async def get_device_input_methods(device_id: str):
    return await proxy_to_local("GET", f"/input_methods/{device_id}", device_serial=device_id)


@app.post("/switch_ime/{device_id}")
async def switch_device_ime(device_id: str, ime: str = Form(...)):
    local_url = get_local_url_for_device(device_id)
    if not local_url:
        local_url = get_active_local_url()
    if not local_url:
        raise HTTPException(status_code=503, detail="没有可用的本地端服务")
    
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.post(f"{local_url}/switch_ime/{device_id}", data={"ime": ime})
            return response.json()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============== 远程控制（代理） ==============

@app.post("/remote/action/{device_id}")
async def remote_action(device_id: str, action: dict):
    return await proxy_to_local("POST", f"/remote/action/{device_id}", device_serial=device_id, json=action)


@app.websocket("/ws/remote/{device_id}")
async def remote_control_ws_proxy(websocket: WebSocket, device_id: str):
    await websocket.accept()
    
    local_url = get_local_url_for_device(device_id) or get_active_local_url()
    if not local_url:
        await websocket.send_json({"type": "error", "message": "没有可用的本地端服务"})
        await websocket.close()
        return
    
    import websockets
    
    try:
        ws_url = local_url.replace("http://", "ws://").replace("https://", "wss://")
        async with websockets.connect(f"{ws_url}/ws/remote/{device_id}") as local_ws:
            
            async def forward_to_local():
                try:
                    while True:
                        data = await websocket.receive_text()
                        await local_ws.send(data)
                except WebSocketDisconnect:
                    pass
            
            async def forward_from_local():
                try:
                    async for message in local_ws:
                        await websocket.send_text(message)
                except:
                    pass
            
            await asyncio.gather(forward_to_local(), forward_from_local())
    
    except Exception as e:
        logger.error(f"WebSocket proxy error: {e}")
        await websocket.close()


# ============== iOS WDA（代理） ==============

@app.post("/ios/wda/config/{device_id}")
async def set_ios_wda_config(device_id: str, config: dict):
    return await proxy_to_local("POST", f"/ios/wda/config/{device_id}", device_serial=device_id, json=config)


@app.get("/ios/wda/config/{device_id}")
async def get_ios_wda_config(device_id: str):
    return await proxy_to_local("GET", f"/ios/wda/config/{device_id}", device_serial=device_id)


@app.get("/ios/wda/status/{device_id}")
async def check_ios_wda_status(device_id: str):
    return await proxy_to_local("GET", f"/ios/wda/status/{device_id}", device_serial=device_id)


@app.get("/ios/remote/screen/{device_id}")
async def get_ios_remote_screen(device_id: str):
    return await proxy_to_local("GET", f"/ios/remote/screen/{device_id}", device_serial=device_id)


@app.post("/ios/remote/action/{device_id}")
async def ios_remote_action(device_id: str, action: dict):
    return await proxy_to_local("POST", f"/ios/remote/action/{device_id}", device_serial=device_id, json=action)


# ============== 任务执行（代理） ==============

def _fill_api_config(request_dict: dict) -> dict:
    """根据 api_config_id 填充 API 配置"""
    api_config_id = request_dict.get("api_config_id")
    if api_config_id:
        api_configs = config_manager.load_api_configs()
        for cfg in api_configs:
            if cfg.get("id") == api_config_id:
                request_dict["base_url"] = cfg.get("base_url", request_dict.get("base_url"))
                request_dict["model"] = cfg.get("model", request_dict.get("model"))
                request_dict["api_key"] = cfg.get("api_key", request_dict.get("api_key"))
                break
    return request_dict


@app.post("/run_task")
async def run_task(request: TaskRequest):
    request_dict = _fill_api_config(request.dict())
    return await proxy_to_local("POST", "/run_task", device_serial=request.device_id, json=request_dict)


@app.post("/run_batch")
async def run_batch(request: BatchTaskRequest):
    request_dict = _fill_api_config(request.dict())
    return await proxy_to_local(
        "POST", "/run_batch", 
        device_serial=request.device_id, 
        local_name=request.local_name,  # 使用 local_name 路由到指定本地端
        json=request_dict
    )


@app.get("/task/{task_id}")
async def get_task_status(task_id: str):
    # 尝试所有本地端
    for name, info in registered_locals.items():
        if not info.get("online"):
            continue
        try:
            async with httpx.AsyncClient(timeout=5.0) as client:
                response = await client.get(f"{info['url']}/task/{task_id}")
                if response.status_code == 200:
                    return response.json()
        except:
            pass
    raise HTTPException(status_code=404, detail="任务不存在")


@app.post("/stop_task/{task_id}")
async def stop_task(task_id: str):
    # 尝试所有本地端
    for name, info in registered_locals.items():
        if not info.get("online"):
            continue
        try:
            async with httpx.AsyncClient(timeout=5.0) as client:
                response = await client.post(f"{info['url']}/stop_task/{task_id}")
                if response.status_code == 200:
                    return response.json()
        except:
            pass
    raise HTTPException(status_code=404, detail="任务不存在")


@app.websocket("/ws/{task_id}")
async def websocket_endpoint_proxy(websocket: WebSocket, task_id: str):
    await websocket.accept()
    
    local_url = get_active_local_url()
    if not local_url:
        await websocket.send_json({"type": "error", "message": "没有可用的本地端服务"})
        await websocket.close()
        return
    
    import websockets
    
    try:
        ws_url = local_url.replace("http://", "ws://").replace("https://", "wss://")
        local_ws = await websockets.connect(f"{ws_url}/ws/{task_id}", close_timeout=5, ping_interval=None)
        
        try:
            # 只做一件事：从 local 转发消息到前端，直到收到完成消息或连接断开
            async for message in local_ws:
                try:
                    await websocket.send_text(message)
                    # 检查是否完成
                    try:
                        msg_data = json.loads(message)
                        if msg_data.get("type") in ["complete", "batch_complete"]:
                            logger.info(f"任务 {task_id} 完成，关闭 WebSocket")
                            break
                    except:
                        pass
                except Exception as e:
                    logger.debug(f"发送到前端失败: {e}")
                    break
        except Exception as e:
            logger.debug(f"本地端 WebSocket 异常: {e}")
        finally:
            # 关闭本地连接
            try:
                await local_ws.close()
            except:
                pass
    
    except Exception as e:
        if "assert" not in str(e).lower():
            logger.debug(f"WebSocket proxy 连接失败: {e}")
        # 连接失败时通知前端
        try:
            await websocket.send_json({"type": "error", "message": f"连接本地端失败: {str(e)}"})
        except:
            pass
    finally:
        # 关闭前端连接
        try:
            await websocket.close()
        except:
            pass


# ============== Scrcpy（代理） ==============

@app.post("/scrcpy/start/{device_id}")
async def start_scrcpy(device_id: str, max_size: int = 720, bit_rate: int = 2000000):
    return await proxy_to_local(
        "POST",
        f"/scrcpy/start/{device_id}",
        device_serial=device_id,
        params={"max_size": max_size, "bit_rate": bit_rate}
    )


@app.post("/scrcpy/stop/{device_id}")
async def stop_scrcpy(device_id: str):
    return await proxy_to_local("POST", f"/scrcpy/stop/{device_id}", device_serial=device_id)


@app.get("/scrcpy/status/{device_id}")
async def scrcpy_status(device_id: str):
    return await proxy_to_local("GET", f"/scrcpy/status/{device_id}", device_serial=device_id)


# ============== API 配置管理（本地存储） ==============

@app.get("/api_configs")
async def get_api_configs():
    return config_manager.load_api_configs()


@app.post("/api_configs")
async def add_api_config(config: APIConfig):
    result = config_manager.add_api_config(config.dict())
    return {"status": "success", "config": result}


@app.put("/api_configs/{config_id}")
async def update_api_config(config_id: str, config: APIConfig):
    if config_manager.update_api_config(config_id, config.dict()):
        return {"status": "success"}
    raise HTTPException(status_code=404, detail="配置不存在")


@app.delete("/api_configs/{config_id}")
async def delete_api_config(config_id: str):
    config_manager.delete_api_config(config_id)
    return {"status": "success"}


# ============== 场景管理（本地存储） ==============

@app.get("/scenarios")
async def get_scenarios():
    return config_manager.load_scenarios()


@app.post("/scenarios")
async def add_scenario(name: str = Form(...), test_cases: str = Form(...)):
    test_cases_list = json.loads(test_cases)
    scenario = config_manager.add_scenario(name, test_cases_list)
    return {"status": "success", "scenario": scenario}


@app.delete("/scenarios/{scenario_id}")
async def delete_scenario(scenario_id: str):
    config_manager.delete_scenario(scenario_id)
    return {"status": "success"}


# ============== 截图文件访问 ==============

@app.get("/data/screenshots/{batch_id}/{filename}")
async def get_screenshot_file(batch_id: str, filename: str):
    """获取截图文件"""
    filepath = config_manager.get_screenshot_path(f"screenshots/{batch_id}/{filename}")
    if os.path.exists(filepath):
        return FileResponse(filepath, media_type="image/png")
    raise HTTPException(status_code=404, detail="截图文件不存在")


# ============== 历史记录（分文件存储） ==============

@app.get("/history")
async def get_history():
    return config_manager.load_history()


@app.get("/history/{record_id}")
async def get_history_detail(record_id: str):
    """获取历史记录详情"""
    detail = config_manager.get_history_detail(record_id)
    if detail:
        return detail
    raise HTTPException(status_code=404, detail="记录不存在")


@app.delete("/history/{record_id}")
async def delete_history(record_id: str):
    config_manager.delete_history(record_id)
    return {"status": "success"}


@app.post("/history")
async def add_history(record: dict):
    """接收本地端推送的历史记录 - 直接保存，不重新生成name"""
    record_id = record.get("batch_id") or record.get("id")
    # 直接保存本地端发送的完整数据，不重新生成name
    config_manager.save_history_direct(record_id, record)
    return {"status": "success", "id": record_id}


# ============== 测试报告 ==============

class ReportRequest(BaseModel):
    record_ids: List[str]
    report_name: str = ""


@app.post("/report/generate")
async def generate_report(request: ReportRequest):
    """生成测试报告"""
    report_id = config_manager.generate_report(request.record_ids, request.report_name)
    if report_id:
        return {"status": "success", "report_id": report_id}
    raise HTTPException(status_code=400, detail="生成报告失败，请检查记录是否存在")


@app.get("/reports")
async def list_reports():
    """列出所有报告"""
    return config_manager.list_reports()


@app.get("/report/{report_id}")
async def get_report(report_id: str):
    """获取报告 HTML（在浏览器中直接显示）"""
    report_path = config_manager.get_report_path(report_id)
    if report_path:
        from fastapi.responses import HTMLResponse
        with open(report_path, "r", encoding="utf-8") as f:
            html_content = f.read()
        return HTMLResponse(content=html_content)
    raise HTTPException(status_code=404, detail="报告不存在")


@app.delete("/report/{report_id}")
async def delete_report(report_id: str):
    """删除报告"""
    config_manager.delete_report(report_id)
    return {"status": "success"}


# ============== CSV/Excel 导入导出 ==============

@app.post("/upload_csv")
async def upload_csv(file: UploadFile = File(...)):
    try:
        content = await file.read()
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        test_cases = []
        
        for i, row in enumerate(reader):
            name = row.get("name") or row.get("用例名称") or row.get("名称") or f"用例{i+1}"
            desc = row.get("description") or row.get("描述") or row.get("测试步骤") or row.get("steps") or ""
            expected = row.get("expected") or row.get("预期结果") or ""
            
            if desc:
                test_cases.append({
                    "id": str(uuid.uuid4())[:8],
                    "name": name,
                    "description": desc,
                    "expected": expected,
                    "selected": False,
                })
        
        return {"status": "success", "test_cases": test_cases, "count": len(test_cases)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)}")


@app.get("/template_csv")
async def get_template_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["name", "description", "expected"])
    writer.writerow(["登录测试", "1. 打开应用\n2. 点击登录\n3. 输入账号密码\n4. 点击确认", "登录成功"])
    writer.writerow(["搜索测试", "1. 点击搜索框\n2. 输入关键词\n3. 点击搜索", "显示搜索结果"])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=test_case_template.csv"}
    )


@app.get("/export_excel/{record_id}")
async def export_excel(record_id: str):
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        import base64
    except ImportError:
        raise HTTPException(status_code=500, detail="请安装 openpyxl 和 Pillow")
    
    # 使用分文件存储的详情获取
    record = config_manager.get_history_detail(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    def format_time(iso_str):
        """格式化时间为 年-月-日 时:分:秒"""
        if not iso_str:
            return "-"
        try:
            from datetime import datetime as dt
            if "T" in iso_str:
                d = dt.fromisoformat(iso_str.replace("Z", "+00:00"))
            else:
                d = dt.fromisoformat(iso_str)
            return d.strftime("%Y-%m-%d %H:%M:%S")
        except:
            return iso_str
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试结果"
    
    # 去掉开始时间列，只保留完成时间
    headers = ["序号", "用例名称", "状态", "结果", "完成时间", "关键截图"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for idx, case in enumerate(record.get("case_results", []), 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=case.get("case_name", ""))
        ws.cell(row=row, column=3, value="通过" if case.get("status") == "success" else "失败")
        ws.cell(row=row, column=4, value=case.get("result", "")[:200])
        ws.cell(row=row, column=5, value=format_time(case.get("end_time", "")))
        
        screenshots = case.get("screenshots", [])
        if screenshots:
            screenshot = screenshots[-1]
            try:
                img_data = base64.b64decode(screenshot.get("image", ""))
                img_io = io.BytesIO(img_data)
                pil_img = PILImage.open(img_io)
                pil_img.thumbnail((150, 300))
                img_io2 = io.BytesIO()
                pil_img.save(img_io2, format="PNG")
                img_io2.seek(0)
                
                xl_img = XLImage(img_io2)
                ws.add_image(xl_img, f"F{row}")
                ws.row_dimensions[row].height = 200
            except:
                ws.cell(row=row, column=6, value="截图加载失败")
    
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"test_result_{record_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/download_adbkeyboard")
async def download_adbkeyboard():
    apk_path = os.path.join(DOWNLOADS_DIR, "ADBKeyboard.apk")
    if os.path.exists(apk_path):
        return FileResponse(apk_path, filename="ADBKeyboard.apk")
    raise HTTPException(status_code=404, detail="文件未找到")


# ============== 启动函数 ==============

def main():
    """启动服务端"""
    # 配置日志过滤器，隐藏 websockets 库的 AssertionError
    import logging
    
    class WebSocketErrorFilter(logging.Filter):
        def filter(self, record):
            msg = str(record.getMessage()).lower()
            if "assertionerror" in msg or "data transfer failed" in msg:
                return False
            if "waiter is none or waiter.cancelled" in msg:
                return False
            return True
    
    for name in ["uvicorn.error", "websockets", "websockets.legacy.protocol"]:
        log = logging.getLogger(name)
        log.addFilter(WebSocketErrorFilter())
    
    print(f"""
╔═══════════════════════════════════════════════════════════════╗
║                                                               ║
║    🌐 AutoGLM 服务端                                          ║
║                                                               ║
║    访问地址: http://0.0.0.0:8792                              ║
║                                                               ║
║    功能:                                                      ║
║    • 页面展示                                                 ║
║    • API 配置管理                                             ║
║    • 场景/历史记录管理                                        ║
║    • 自动发现本地端                                           ║
║                                                               ║
║    等待本地端注册...                                          ║
║    Mac 上运行: autoglm-local --server http://服务器IP:8792    ║
║                                                               ║
║    作者: chenwenkun                                           ║
║                                                               ║
╚═══════════════════════════════════════════════════════════════╝
    """)
    
    uvicorn.run(app, host="0.0.0.0", port=8792)


if __name__ == "__main__":
    main()

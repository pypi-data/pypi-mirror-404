"""
MS Excel Tool migrated to use AbstractDocumentTool framework.
"""
import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Literal, Union
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf.style import Style, TableCellProperties, TextProperties
from pydantic import Field, field_validator, ConfigDict
from .document import AbstractDocumentTool, DocumentGenerationArgs


class ExcelArgs(DocumentGenerationArgs):
    """Arguments schema for Excel/ODS Document generation."""

    content: Union[pd.DataFrame, List[Dict[str, Any]]] = Field(
        ...,
        description="Pandas DataFrame or Dataset to export to Excel/ODS format"
    )
    sheet_name: str = Field(
        "Sheet1",
        description="Name of the worksheet"
    )
    template_file: Optional[str] = Field(
        None,
        description="Path to Excel/ODS template file to use as base"
    )
    output_format: Literal["excel", "ods"] = Field(
        "excel",
        description="Export format - 'excel' for .xlsx or 'ods' for OpenDocument"
    )
    header_styles: Optional[Dict[str, Any]] = Field(
        None,
        description="Dictionary of styles to apply to headers (font, color, etc.)"
    )
    data_styles: Optional[Dict[str, Any]] = Field(
        None,
        description="Dictionary of styles to apply to data cells"
    )
    auto_adjust_columns: bool = Field(
        True,
        description="Whether to auto-adjust column widths based on content"
    )
    freeze_header: bool = Field(
        False,
        description="Whether to freeze the header row"
    )

    model_config = ConfigDict(arbitrary_types_allowed=True)

    @field_validator('content')
    @classmethod
    def validate_content(cls, v):
        if isinstance(v, pd.DataFrame):
            if v.empty:
                raise ValueError("DataFrame content cannot be empty")
        elif isinstance(v, list):
            if not v:
                raise ValueError("Content list cannot be empty")
            if not all(isinstance(item, dict) for item in v):
                raise ValueError("Content list must contain only dictionaries")
        if v is None:
            raise ValueError("content cannot be empty")
        return v

    @field_validator('sheet_name')
    @classmethod
    def validate_sheet_name(cls, v):
        if not v or not v.strip():
            raise ValueError("sheet_name cannot be empty")
        # Excel sheet name limitations
        invalid_chars = r'[:/\\?*\[\]]'
        if any(char in v for char in invalid_chars):
            raise ValueError(f"sheet_name contains invalid characters: {invalid_chars}")
        if len(v) > 31:
            raise ValueError("sheet_name cannot exceed 31 characters")
        return v.strip()


class ExcelTool(AbstractDocumentTool):
    """
    Microsoft Excel/OpenDocument Spreadsheet Generation Tool.

    This tool exports pandas DataFrames to Excel (.xlsx) or OpenDocument (.ods) files
    with support for custom styling, templates, and advanced formatting features.

    Features:
    - Export DataFrames to Excel or ODS formats
    - Custom header and data cell styling
    - Template support for both formats
    - Auto-adjusting column widths
    - Header row freezing
    - Professional spreadsheet formatting
    - Comprehensive error handling and validation
    """

    name = "excel_generator"
    description = (
        "Export pandas DataFrames to Excel (.xlsx) or OpenDocument (.ods) files "
        "with custom styling, templates, and professional formatting. "
        "Supports both Excel and ODS formats with advanced styling options."
    )
    args_schema = ExcelArgs

    # Document type configuration
    document_type = "spreadsheet"
    default_extension = "xlsx"
    supported_extensions = [".xlsx", ".xls", ".ods"]

    def __init__(
        self,
        templates_dir: Optional[Path] = None,
        default_format: Literal["excel", "ods"] = "excel",
        **kwargs
    ):
        """
        Initialize the Excel Tool.

        Args:
            templates_dir: Directory containing Excel/ODS templates
            default_format: Default output format ('excel' or 'ods')
            **kwargs: Additional arguments for AbstractDocumentTool
        """
        super().__init__(templates_dir=templates_dir, **kwargs)
        self.default_format = default_format

    def _detect_output_format(self, output_format: str, filename: Optional[str] = None) -> str:
        """Detect output format from format parameter or filename extension."""
        if output_format:
            return output_format

        if filename:
            ext = Path(filename).suffix.lower()
            if ext in ['.ods']:
                return 'ods'
            elif ext in ['.xlsx', '.xls']:
                return 'excel'

        return self.default_format

    def _parse_content_to_dataframe(self, content: Union[str, List[Dict], pd.DataFrame]) -> pd.DataFrame:
        """
        Parse content into a pandas DataFrame.

        Args:
            content: Can be a DataFrame, list of dictionaries, or JSON string

        Returns:
            pandas DataFrame
        """
        if isinstance(content, pd.DataFrame):
            return content
        elif isinstance(content, dict):
            # Single dictionary, convert to DataFrame
            return pd.DataFrame(content)
        elif isinstance(content, list):
            if not content:
                raise ValueError("Content list cannot be empty")
            if not all(isinstance(item, dict) for item in content):
                raise ValueError("Content list must contain only dictionaries")
            return pd.DataFrame(content)
        elif isinstance(content, str):
            try:
                # Try to parse as JSON
                data = self._json_decoder(content)
                if isinstance(data, list):
                    return pd.DataFrame(data)
                elif isinstance(data, dict):
                    # Single dictionary, wrap in list
                    return pd.DataFrame([data])
                else:
                    raise ValueError("JSON content must be a list of objects or a single object")
            except (TypeError, ValueError):
                raise ValueError("String content must be valid JSON")
        else:
            raise ValueError(
                "Content must be a pandas DataFrame, list of dictionaries, or JSON string"
            )

    def _get_extension_for_format(self, output_format: str) -> str:
        """Get file extension for the specified format."""
        return "ods" if output_format == "ods" else "xlsx"

    def _create_excel_workbook(
        self,
        dataframe: pd.DataFrame,
        sheet_name: str,
        template_file: Optional[str] = None,
        header_styles: Optional[Dict[str, Any]] = None,
        data_styles: Optional[Dict[str, Any]] = None,
        auto_adjust_columns: bool = True,
        freeze_header: bool = False
    ) -> Workbook:
        """Create Excel workbook with DataFrame data and styling."""
        # Load template or create new workbook
        if template_file:
            template_path = self._get_template_path(template_file)
            if template_path and template_path.exists():
                wb = load_workbook(str(template_path))
                self.logger.info(f"Loaded Excel template: {template_path}")

                # Clear existing data in the target sheet if it exists
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet(sheet_name)
            else:
                self.logger.warning(f"Template not found: {template_file}, creating new workbook")
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

        # Convert DataFrame to rows
        rows = dataframe_to_rows(dataframe, index=False, header=True)

        # Add data to worksheet
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Apply header styles
                if r_idx == 1 and header_styles:
                    self._apply_excel_cell_style(cell, header_styles)
                # Apply data styles
                elif r_idx > 1 and data_styles:
                    self._apply_excel_cell_style(cell, data_styles)

        # Apply default header styling if no custom styles provided
        if not header_styles:
            self._apply_default_excel_header_styles(ws, len(dataframe.columns))

        # Auto-adjust column widths
        if auto_adjust_columns:
            self._adjust_excel_column_widths(ws)

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = ws['A2']

        return wb

    def _apply_excel_cell_style(self, cell, styles: Dict[str, Any]) -> None:
        """Apply styles to an Excel cell."""
        # Font styling
        if any(key in styles for key in ['font_name', 'font_size', 'bold', 'italic', 'font_color']):
            font_kwargs = {}
            if 'font_name' in styles:
                font_kwargs['name'] = styles['font_name']
            if 'font_size' in styles:
                font_kwargs['size'] = styles['font_size']
            if 'bold' in styles:
                font_kwargs['bold'] = styles['bold']
            if 'italic' in styles:
                font_kwargs['italic'] = styles['italic']
            if 'font_color' in styles:
                font_kwargs['color'] = styles['font_color']

            cell.font = Font(**font_kwargs)

        # Background color
        if 'background_color' in styles:
            cell.fill = PatternFill(
                start_color=styles['background_color'],
                end_color=styles['background_color'],
                fill_type="solid"
            )

        # Alignment
        if any(key in styles for key in ['horizontal', 'vertical', 'wrap_text']):
            align_kwargs = {}
            if 'horizontal' in styles:
                align_kwargs['horizontal'] = styles['horizontal']
            if 'vertical' in styles:
                align_kwargs['vertical'] = styles['vertical']
            if 'wrap_text' in styles:
                align_kwargs['wrap_text'] = styles['wrap_text']

            cell.alignment = Alignment(**align_kwargs)

        # Borders
        if 'border' in styles:
            border_style = styles['border']
            side = Side(style=border_style, color="000000")
            cell.border = Border(left=side, right=side, top=side, bottom=side)

    def _apply_default_excel_header_styles(self, ws, num_columns: int) -> None:
        """Apply default styling to header row."""
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _adjust_excel_column_widths(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_ods_document(
        self,
        dataframe: pd.DataFrame,
        sheet_name: str,
        header_styles: Optional[Dict[str, Any]] = None,
        data_styles: Optional[Dict[str, Any]] = None
    ) -> OpenDocumentSpreadsheet:
        """Create ODS document with DataFrame data and styling."""
        # Create new ODS document
        doc = OpenDocumentSpreadsheet()

        # Create styles
        header_style = self._create_ods_header_style(doc, header_styles)
        data_style = self._create_ods_data_style(doc, data_styles)

        # Create table (sheet)
        table = Table(name=sheet_name)

        # Add header row
        header_row = TableRow()
        for col_name in dataframe.columns:
            cell = TableCell(stylename=header_style)
            cell.addElement(P(text=str(col_name)))
            header_row.addElement(cell)
        table.addElement(header_row)

        # Add data rows
        for _, row in dataframe.iterrows():
            data_row = TableRow()
            for value in row:
                cell = TableCell(stylename=data_style)
                cell.addElement(P(text=str(value) if pd.notna(value) else ""))
                data_row.addElement(cell)
            table.addElement(data_row)

        doc.spreadsheet.addElement(table)
        return doc

    def _create_ods_header_style(self, doc, custom_styles: Optional[Dict[str, Any]]) -> str:
        """Create ODS style for headers."""
        style = Style(name="HeaderStyle", family="table-cell")

        # Default header properties
        cell_props = TableCellProperties(
            backgroundcolor="#366092",
            border="1pt solid #000000"
        )
        text_props = TextProperties(
            fontweight="bold",
            color="#FFFFFF",
            fontfamily="Calibri",
            fontsize="12pt"
        )

        # Apply custom styles if provided
        if custom_styles:
            if 'background_color' in custom_styles:
                cell_props.backgroundcolor = custom_styles['background_color']
            if 'font_color' in custom_styles:
                text_props.color = custom_styles['font_color']
            if 'font_name' in custom_styles:
                text_props.fontfamily = custom_styles['font_name']
            if 'font_size' in custom_styles:
                text_props.fontsize = f"{custom_styles['font_size']}pt"
            if 'bold' in custom_styles:
                text_props.fontweight = "bold" if custom_styles['bold'] else "normal"

        style.addElement(cell_props)
        style.addElement(text_props)
        doc.styles.addElement(style)

        return "HeaderStyle"

    def _create_ods_data_style(self, doc, custom_styles: Optional[Dict[str, Any]]) -> str:
        """Create ODS style for data cells."""
        style = Style(name="DataStyle", family="table-cell")

        # Default data properties
        cell_props = TableCellProperties(border="1pt solid #CCCCCC")
        text_props = TextProperties(
            fontfamily="Calibri",
            fontsize="11pt"
        )

        # Apply custom styles if provided
        if custom_styles:
            if 'background_color' in custom_styles:
                cell_props.backgroundcolor = custom_styles['background_color']
            if 'font_color' in custom_styles:
                text_props.color = custom_styles['font_color']
            if 'font_name' in custom_styles:
                text_props.fontfamily = custom_styles['font_name']
            if 'font_size' in custom_styles:
                text_props.fontsize = f"{custom_styles['font_size']}pt"
            if 'bold' in custom_styles:
                text_props.fontweight = "bold" if custom_styles['bold'] else "normal"

        style.addElement(cell_props)
        style.addElement(text_props)
        doc.styles.addElement(style)

        return "DataStyle"

    async def _generate_document_content(self, content: str, **kwargs) -> bytes:
        """
        Generate Excel/ODS document content from structured data.

        Args:
            content: Structured data - can be JSON string, list of dicts, or DataFrame
            **kwargs: Additional arguments from ExcelArgs

        Returns:
            Excel/ODS document as bytes
        """
        try:
            # Extract arguments
            sheet_name = kwargs.get('sheet_name', 'Sheet1')
            template_file = kwargs.get('template_file')
            output_format = kwargs.get('output_format', 'excel')
            header_styles = kwargs.get('header_styles')
            data_styles = kwargs.get('data_styles')
            auto_adjust_columns = kwargs.get('auto_adjust_columns', True)
            freeze_header = kwargs.get('freeze_header', False)

            # Parse content to DataFrame
            dataframe = self._parse_content_to_dataframe(content)

            self.logger.info(
                f"Generating {output_format} document with {len(dataframe)} rows and "
                f"{len(dataframe.columns)} columns"
            )

            # Create document based on format
            doc_bytes = io.BytesIO()

            if output_format == "ods":
                # Create ODS document
                doc = self._create_ods_document(
                    dataframe=dataframe,
                    sheet_name=sheet_name,
                    header_styles=header_styles,
                    data_styles=data_styles
                )
                doc.save(doc_bytes)
            else:
                # Create Excel document
                wb = self._create_excel_workbook(
                    dataframe=dataframe,
                    sheet_name=sheet_name,
                    template_file=template_file,
                    header_styles=header_styles,
                    data_styles=data_styles,
                    auto_adjust_columns=auto_adjust_columns,
                    freeze_header=freeze_header
                )
                wb.save(doc_bytes)

            doc_bytes.seek(0)
            return doc_bytes.getvalue()

        except Exception as e:
            self.logger.error(f"Error generating {output_format} document: {e}")
            raise

    async def _execute(
        self,
        content: Union[str, List[Dict], pd.DataFrame],
        sheet_name: str = "Sheet1",
        template_file: Optional[str] = None,
        output_format: Literal["excel", "ods"] = "excel",
        header_styles: Optional[Dict[str, Any]] = None,
        data_styles: Optional[Dict[str, Any]] = None,
        auto_adjust_columns: bool = True,
        freeze_header: bool = False,
        output_filename: Optional[str] = None,
        file_prefix: str = "spreadsheet",
        output_dir: Optional[str] = None,
        overwrite_existing: bool = False,
        **kwargs
    ) -> Dict[str, Any]:
        """
        Execute Excel/ODS document generation (AbstractTool interface).

        Args:
            content: Structured data - DataFrame, list of dicts, or JSON string
            sheet_name: Name of the worksheet
            template_file: Excel/ODS template file path
            output_format: Export format ('excel' or 'ods')
            header_styles: Custom header styling
            data_styles: Custom data cell styling
            auto_adjust_columns: Whether to auto-adjust column widths
            freeze_header: Whether to freeze the header row
            output_filename: Custom filename (without extension)
            file_prefix: Prefix for auto-generated filenames
            output_dir: Custom output directory
            overwrite_existing: Whether to overwrite existing files
            **kwargs: Additional arguments

        Returns:
            Dictionary with document generation results
        """
        try:
            # Parse content to get DataFrame info for logging
            dataframe = self._parse_content_to_dataframe(content)

            # Detect format and set appropriate extension
            detected_format = self._detect_output_format(output_format, output_filename)
            extension = self._get_extension_for_format(detected_format)

            self.logger.info(
                f"Starting {detected_format.upper()} document generation with "
                f"{len(dataframe)} rows and {len(dataframe.columns)} columns"
            )

            # Use the safe document creation workflow
            result = await self._create_document_safely(
                content=dataframe,
                output_filename=output_filename,
                file_prefix=file_prefix,
                output_dir=output_dir,
                overwrite_existing=overwrite_existing or self.overwrite_existing,
                extension=extension,
                sheet_name=sheet_name,
                template_file=template_file,
                output_format=detected_format,
                header_styles=header_styles,
                data_styles=data_styles,
                auto_adjust_columns=auto_adjust_columns,
                freeze_header=freeze_header
            )

            if result['status'] == 'success':
                # Add Excel-specific metadata
                result['metadata'].update({
                    'format': detected_format,
                    'sheet_name': sheet_name,
                    'rows': len(dataframe),
                    'columns': len(dataframe.columns),
                    'column_names': list(dataframe.columns)
                })

                self.logger.info(
                    f"{detected_format.upper()} document created successfully: "
                    f"{result['metadata']['filename']}"
                )

            return result

        except Exception as e:
            self.logger.error(f"Error in Excel document generation: {e}")
            raise

    # Convenience methods for direct DataFrame export
    async def export_dataframe(
        self,
        dataframe: pd.DataFrame,
        output_format: Literal["excel", "ods"] = "excel",
        **kwargs
    ) -> Dict[str, Any]:
        """
        Convenience method to directly export a DataFrame.

        Args:
            dataframe: DataFrame to export
            output_format: Export format
            **kwargs: Additional arguments for _execute

        Returns:
            Dictionary with export results
        """
        return await self._execute(
            content=dataframe,
            output_format=output_format,
            **kwargs
        )

    async def export_data(
        self,
        data: Union[List[Dict], pd.DataFrame, str],
        output_format: Literal["excel", "ods"] = "excel",
        **kwargs
    ) -> Dict[str, Any]:
        """
        Convenience method to export various data formats.

        Args:
            data: Data to export (DataFrame, list of dicts, or JSON string)
            output_format: Export format
            **kwargs: Additional arguments for _execute

        Returns:
            Dictionary with export results
        """
        return await self._execute(
            content=data,
            output_format=output_format,
            **kwargs
        )

    def get_format_info(self) -> Dict[str, Any]:
        """Get information about supported formats."""
        return {
            "supported_formats": ["excel", "ods"],
            "excel_extensions": [".xlsx", ".xls"],
            "ods_extensions": [".ods"],
            "default_format": self.default_format,
            "features": {
                "templates": True,
                "custom_styling": True,
                "auto_column_width": True,
                "freeze_panes": True,
                "multiple_sheets": False  # Could be extended in future
            }
        }


class DataFrameToExcelTool(ExcelTool):
    """
    Simplified Excel tool that focuses purely on DataFrame export.

    This is a convenience wrapper around ExcelTool for users who primarily
    need to export DataFrames without complex document features.
    """

    name = "dataframe_to_excel"
    description = (
        "Simple tool to export pandas DataFrames to Excel or ODS files. "
        "Focused on quick DataFrame export with minimal configuration."
    )

    async def quick_export(
        self,
        data: Union[pd.DataFrame, List[Dict], str],
        filename: Optional[str] = None,
        format: Literal["excel", "ods"] = "excel"
    ) -> str:
        """
        Quick export method that returns just the file path.

        Args:
            data: Data to export (DataFrame, list of dicts, or JSON string)
            filename: Optional filename
            format: Export format

        Returns:
            Path to the created file
        """
        result = await self.export_data(
            data=data,
            output_filename=filename,
            output_format=format
        )

        if result['status'] == 'success':
            return result['metadata']['file_path']
        else:
            raise Exception(
                f"Export failed: {result.get('error', 'Unknown error')}"
            )

#!/usr/bin/env python3
"""
SCF Excel Data Extraction Script
Extracts controls and framework mappings from SCF 2025.4 Excel file

Usage:
    python scf-extract-starter.py /path/to/secure-controls-framework-scf-2025-4.xlsx

Output:
    - scf-controls.json: Full control database
    - framework-to-scf.json: Reverse index
    - extraction-stats.txt: Statistics report
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


def extract_scf_controls(excel_path: Path) -> dict:
    """Extract SCF controls from Excel file."""
    print(f"📂 Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb["SCF 2025.4"]

    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"✅ Found {len(headers)} columns")

    # Define framework columns to extract
    framework_columns = {
        # Original 16 frameworks
        "nist_csf_2.0": "NIST\nCSF\n2.0",
        "nist_800_53_r5": "NIST\n800-53\nR5",
        "iso_27001_2022": "ISO\n27001\n2022",
        "iso_27002_2022": "ISO\n27002\n2022",
        "cis_csc_8.1": "CIS\nCSC\n8.1",
        "pci_dss_4.0.1": "PCI DSS\n4.0.1",
        "cmmc_2.0_level_1": "US\nCMMC 2.0\nLevel 1",
        "cmmc_2.0_level_2": "US\nCMMC 2.0\nLevel 2",
        "soc_2_tsc": "AICPA\nTSC 2017:2022 (used for SOC 2)",
        "dora": "EMEA\nEU\nDORA",
        "nis2": "EMEA\nEU\nNIS2",
        "gdpr": "EMEA\nEU\nGDPR",
        "ncsc_caf_4.0": "EMEA\nUK\nCAF 4.0",
        "uk_cyber_essentials": "EMEA\nUK\nCyber Essentials",
        "fedramp_r5_moderate": "US\nFedRAMP R5\n(moderate)",
        "hipaa_security_rule": "US\nHIPAA\nSecurity Rule / NIST SP 800-66 R2",
        # Tier 1: APAC (3 frameworks)
        "australia_essential_8": "APAC\nAustralia\nEssential 8",
        "australia_ism_2024": "APAC\nAustralia\nISM\nJune 2024",
        "singapore_mas_trm_2021": "APAC\nSingapore MAS\nTRM 2021",
        # Tier 1: Industry/Privacy (2 frameworks)
        "swift_cscf_2023": "SWIFT\nCSF\n2023",
        "nist_privacy_framework_1_0": "NIST Privacy Framework\n1.0",
        # Tier 2: European National (6 frameworks)
        "netherlands": "EMEA\nNetherlands",
        "norway": "EMEA\nNorway",
        "sweden": "EMEA\nSweden",
        "germany": "EMEA\nGermany",
        "germany_bait": "EMEA\nGermany\nBanking Supervisory Requirements for IT (BAIT)",
        "germany_c5_2020": "EMEA\nGermany\nC5\n2020",
        # Tier 3: Cloud (1 framework)
        "csa_ccm_4": "CSA\nCCM\n4",
    }

    # Build column index map
    col_indices = {}
    for fw_key, fw_header in framework_columns.items():
        try:
            col_indices[fw_key] = headers.index(fw_header)
        except ValueError:
            print(f"⚠️  Warning: Framework column not found: {fw_header}")

    # Core column indices
    scf_domain_idx = headers.index("SCF Domain")
    scf_control_idx = headers.index("SCF Control")
    scf_num_idx = headers.index("SCF #")
    scf_desc_idx = headers.index("Secure Controls Framework (SCF)\nControl Description")
    weight_idx = headers.index("Relative Control Weighting")
    pptdf_idx = headers.index("PPTDF\nApplicability")
    cadence_idx = headers.index("Conformity Validation\nCadence")

    # Extract controls
    controls = []
    print("🔄 Extracting controls...")

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[scf_num_idx]:  # Skip empty rows
            continue

        control = {
            "id": row[scf_num_idx],
            "domain": row[scf_domain_idx],
            "name": row[scf_control_idx],
            "description": row[scf_desc_idx],
            "weight": row[weight_idx],
            "pptdf": row[pptdf_idx],
            "validation_cadence": row[cadence_idx],
            "framework_mappings": {},
        }

        # Extract framework mappings
        for fw_key, fw_idx in col_indices.items():
            mapping = row[fw_idx]
            if mapping and mapping != "None":
                # Split newline-separated values and clean
                mapped_controls = [c.strip() for c in str(mapping).split("\n") if c.strip()]
                control["framework_mappings"][fw_key] = mapped_controls
            else:
                control["framework_mappings"][fw_key] = None

        controls.append(control)

        if row_num % 100 == 0:
            print(f"  Processed {row_num - 1} controls...")

    print(f"✅ Extracted {len(controls)} SCF controls")
    return {"controls": controls, "version": "2025.4"}


def build_reverse_index(data: dict) -> dict:
    """Build framework → SCF control reverse index."""
    print("🔄 Building reverse index...")

    reverse_index = defaultdict(lambda: defaultdict(list))

    for control in data["controls"]:
        scf_id = control["id"]

        for fw_key, fw_controls in control["framework_mappings"].items():
            if fw_controls:
                for fw_control in fw_controls:
                    reverse_index[fw_key][fw_control].append(scf_id)

    # Convert to regular dict
    reverse_index = {fw: dict(controls) for fw, controls in reverse_index.items()}

    print(f"✅ Built reverse index for {len(reverse_index)} frameworks")
    return reverse_index


def generate_stats(data: dict, reverse_index: dict) -> str:
    """Generate extraction statistics report."""
    controls = data["controls"]

    stats = []
    stats.append("=" * 80)
    stats.append("SCF Data Extraction Statistics")
    stats.append("=" * 80)
    stats.append(f"SCF Version: {data['version']}")
    stats.append(f"Total Controls: {len(controls)}")
    stats.append("")

    # Framework mapping stats
    stats.append("Framework Mapping Coverage:")
    stats.append("-" * 80)

    # Count controls per framework
    framework_counts = {}
    for control in controls:
        for fw_key, fw_mappings in control["framework_mappings"].items():
            if fw_mappings:
                framework_counts[fw_key] = framework_counts.get(fw_key, 0) + 1

    # Sort by count descending
    for fw_key, count in sorted(framework_counts.items(), key=lambda x: x[1], reverse=True):
        fw_name = fw_key.replace("_", " ").upper()
        stats.append(f"  {fw_name:30s}: {count:4d} SCF controls mapped")

    stats.append("")

    # Domain distribution
    stats.append("Controls by Domain:")
    stats.append("-" * 80)

    domain_counts = defaultdict(int)
    for control in controls:
        domain = control["domain"]
        domain_counts[domain] += 1

    for domain, count in sorted(domain_counts.items()):
        stats.append(f"  {domain:50s}: {count:4d} controls")

    stats.append("")
    stats.append("=" * 80)

    return "\n".join(stats)


def main():
    if len(sys.argv) < 2:
        print("Usage: python scf-extract-starter.py <path-to-scf-excel-file>")
        print("\nExample:")
        print(
            "  python scf-extract-starter.py /tmp/scf-repo/secure-controls-framework-scf-2025-4.xlsx"
        )
        sys.exit(1)

    excel_path = Path(sys.argv[1])

    if not excel_path.exists():
        print(f"❌ Error: File not found: {excel_path}")
        sys.exit(1)

    if excel_path.suffix not in [".xlsx", ".xlsm"]:
        print(f"❌ Error: Not an Excel file: {excel_path}")
        sys.exit(1)

    print("🚀 SCF Data Extraction Started")
    print("=" * 80)

    # Extract controls
    data = extract_scf_controls(excel_path)

    # Build reverse index
    reverse_index = build_reverse_index(data)

    # Save controls
    controls_file = Path("scf-controls.json")
    print(f"💾 Saving controls to: {controls_file}")
    with open(controls_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # Save reverse index
    reverse_file = Path("framework-to-scf.json")
    print(f"💾 Saving reverse index to: {reverse_file}")
    with open(reverse_file, "w", encoding="utf-8") as f:
        json.dump(reverse_index, f, indent=2, ensure_ascii=False)

    # Generate and save stats
    stats = generate_stats(data, reverse_index)
    stats_file = Path("extraction-stats.txt")
    print(f"💾 Saving statistics to: {stats_file}")
    with open(stats_file, "w", encoding="utf-8") as f:
        f.write(stats)

    print("\n" + stats)
    print("\n✅ Extraction complete!")
    print("\nGenerated files:")
    print(f"  - {controls_file} ({controls_file.stat().st_size / 1024 / 1024:.1f} MB)")
    print(f"  - {reverse_file} ({reverse_file.stat().st_size / 1024 / 1024:.1f} MB)")
    print(f"  - {stats_file} ({stats_file.stat().st_size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()

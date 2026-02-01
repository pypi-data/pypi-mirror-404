"""Data dictionary export — Excel (.xlsx) and Markdown (.md) formats."""

from datetime import datetime
from pathlib import Path
from typing import Any

from urimai.storage.metadata_store import MetadataStore
from urimai.storage.models import (
    DatabaseInfo,
    SchemaInfo,
    TableProfile,
    QualityReport,
)


async def gather_export_data(
    db_name: str,
    include_profile: bool,
    include_quality: bool,
    include_sample_data: bool,
) -> dict[str, Any]:
    """Load all metadata for a database into a structured dict."""
    store = MetadataStore()
    await store.initialize()

    db_info = await store.get_database(db_name)
    if not db_info:
        raise ValueError(f"Database '{db_name}' not found")

    schemas = await store.get_schemas(db_info.id)

    profiles: dict[str, TableProfile] = {}
    quality: dict[str, QualityReport] = {}

    if include_profile:
        profile_list = await store.list_table_profiles(db_info.id)
        profiles = {p.table_name: p for p in profile_list}

    if include_quality:
        for schema in schemas:
            report = await store.get_quality_report(db_info.id, schema.table_name)
            if report:
                quality[schema.table_name] = report

    return {
        "database": db_info,
        "schemas": schemas,
        "profiles": profiles,
        "quality": quality,
        "include_profile": include_profile,
        "include_quality": include_quality,
        "include_sample_data": include_sample_data,
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_db_context(db_info: DatabaseInfo) -> dict[str, Any]:
    """Extract database_context fields safely."""
    ctx = db_info.database_context or {}
    return {
        "domain": ctx.get("domain", ""),
        "purpose": ctx.get("purpose", ""),
        "key_entities": ctx.get("key_entities", []),
    }


def _col_description(schema: SchemaInfo, col_name: str) -> str:
    """Get enriched description for a column."""
    enriched = schema.enriched_metadata or {}
    descs = enriched.get("column_descriptions", {})
    if isinstance(descs, dict):
        return descs.get(col_name, "")
    return ""


def _table_purpose(schema: SchemaInfo) -> str:
    enriched = schema.enriched_metadata or {}
    return enriched.get("table_purpose", "")


def _business_context(schema: SchemaInfo) -> str:
    enriched = schema.enriched_metadata or {}
    return enriched.get("business_context", "")


def _relationships(schema: SchemaInfo) -> list[str]:
    enriched = schema.enriched_metadata or {}
    return enriched.get("relationships", [])


def _constraints(schema: SchemaInfo) -> list[dict[str, str]]:
    """Extract constraints from original_metadata into flat rows."""
    meta = schema.original_metadata
    rows: list[dict[str, str]] = []

    # Primary keys
    for pk in meta.get("primary_keys", []):
        if isinstance(pk, str):
            rows.append({"type": "PRIMARY KEY", "column": pk, "references": ""})
        elif isinstance(pk, dict):
            rows.append({"type": "PRIMARY KEY", "column": pk.get("column_name", str(pk)), "references": ""})

    # Foreign keys
    for fk in meta.get("foreign_keys", []):
        if isinstance(fk, dict):
            src = fk.get("column", fk.get("from", ""))
            ref_table = fk.get("referenced_table", fk.get("table", ""))
            ref_col = fk.get("referenced_column", fk.get("to", ""))
            rows.append({
                "type": "FOREIGN KEY",
                "column": src,
                "references": f"{ref_table}.{ref_col}" if ref_table else "",
            })

    # Unique constraints
    for idx in meta.get("indexes", []):
        if isinstance(idx, dict) and idx.get("unique"):
            cols = idx.get("columns", [])
            col_str = ", ".join(cols) if isinstance(cols, list) else str(cols)
            rows.append({"type": "UNIQUE", "column": col_str, "references": ""})

    return rows


def _fmt(val: Any) -> str:
    """Format a value for display, handling None gracefully."""
    if val is None:
        return ""
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(data: dict[str, Any], output_path: Path) -> None:
    """Write data dictionary to .xlsx using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _write_headers(ws, headers: list[str], row: int = 1) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _auto_width(ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    db_info: DatabaseInfo = data["database"]
    schemas: list[SchemaInfo] = data["schemas"]
    profiles: dict[str, TableProfile] = data["profiles"]
    quality: dict[str, QualityReport] = data["quality"]
    ctx = _get_db_context(db_info)

    # ---- Sheet 1: Overview ----
    ws = wb.active
    ws.title = "Overview"
    overview_rows = [
        ("Database Name", db_info.name),
        ("Domain", ctx["domain"]),
        ("Purpose", ctx["purpose"]),
        ("Key Entities", ", ".join(ctx["key_entities"]) if isinstance(ctx["key_entities"], list) else str(ctx["key_entities"])),
        ("Tables", len(schemas)),
        ("Last Synced", db_info.last_synced_at),
        ("Exported At", datetime.now().isoformat()),
    ]
    _write_headers(ws, ["Field", "Value"])
    for r, (field, value) in enumerate(overview_rows, 2):
        ws.cell(row=r, column=1, value=field).border = thin_border
        ws.cell(row=r, column=2, value=str(value)).border = thin_border
    _auto_width(ws)

    # ---- Sheet 2: Tables ----
    ws_tables = wb.create_sheet("Tables")
    table_headers = ["Table", "Purpose", "Columns", "Rows", "Relationships", "Quality Score"]
    _write_headers(ws_tables, table_headers)
    for r, schema in enumerate(schemas, 2):
        row_count = ""
        if schema.table_name in profiles:
            row_count = profiles[schema.table_name].row_count
        rels = "; ".join(_relationships(schema))
        q_score = ""
        if schema.table_name in quality:
            q_score = f"{quality[schema.table_name].overall_score:.1f}"
        num_cols = len(schema.original_metadata.get("columns", []))

        for ci, val in enumerate([schema.table_name, _table_purpose(schema), num_cols, row_count, rels, q_score], 1):
            ws_tables.cell(row=r, column=ci, value=val if val != "" else None).border = thin_border
    _auto_width(ws_tables)

    # ---- Sheet 3: Columns ----
    ws_cols = wb.create_sheet("Columns")
    col_headers = ["Table", "Column", "Type", "Nullable", "Default", "Description"]
    if data["include_profile"]:
        col_headers += ["Null%", "Distinct%", "Role", "Min", "Max", "Mean"]
    _write_headers(ws_cols, col_headers)

    row_num = 2
    for schema in schemas:
        profile = profiles.get(schema.table_name)
        profile_map: dict[str, Any] = {}
        if profile:
            profile_map = {cp.column_name: cp for cp in profile.columns}

        for col in schema.original_metadata.get("columns", []):
            col_name = col.get("column_name", "")
            values = [
                schema.table_name,
                col_name,
                col.get("data_type", ""),
                "Yes" if col.get("nullable") else "No",
                _fmt(col.get("default_value")),
                _col_description(schema, col_name),
            ]
            if data["include_profile"]:
                cp = profile_map.get(col_name)
                if cp:
                    values += [
                        f"{cp.null_percentage:.1f}",
                        f"{cp.distinct_percentage:.1f}",
                        cp.inferred_role or "",
                        _fmt(cp.min_value),
                        _fmt(cp.max_value),
                        _fmt(cp.mean),
                    ]
                else:
                    values += [""] * 6

            for ci, val in enumerate(values, 1):
                ws_cols.cell(row=row_num, column=ci, value=val if val != "" else None).border = thin_border
            row_num += 1
    _auto_width(ws_cols)

    # ---- Sheet 4: Constraints ----
    ws_const = wb.create_sheet("Constraints")
    _write_headers(ws_const, ["Table", "Constraint Type", "Column", "References"])
    row_num = 2
    for schema in schemas:
        for c in _constraints(schema):
            for ci, val in enumerate([schema.table_name, c["type"], c["column"], c["references"]], 1):
                ws_const.cell(row=row_num, column=ci, value=val if val != "" else None).border = thin_border
            row_num += 1
    _auto_width(ws_const)

    # ---- Sheet 5: Quality ----
    if data["include_quality"] and quality:
        ws_q = wb.create_sheet("Quality")
        _write_headers(ws_q, ["Table", "Overall", "Completeness", "Uniqueness", "Validity", "Consistency"])
        row_num = 2
        for tname, report in quality.items():
            for ci, val in enumerate([
                tname,
                f"{report.overall_score:.1f}",
                f"{report.completeness_score:.1f}",
                f"{report.uniqueness_score:.1f}",
                f"{report.validity_score:.1f}",
                f"{report.consistency_score:.1f}",
            ], 1):
                ws_q.cell(row=row_num, column=ci, value=val).border = thin_border
            row_num += 1

        # Blank row then issues
        row_num += 1
        _write_headers(ws_q, ["Table", "Severity", "Category", "Column", "Issue", "Recommendation"], row=row_num)
        row_num += 1
        for tname, report in quality.items():
            for issue in report.issues:
                for ci, val in enumerate([
                    tname,
                    issue.severity,
                    issue.category,
                    issue.column_name or "",
                    issue.issue,
                    issue.recommendation,
                ], 1):
                    ws_q.cell(row=row_num, column=ci, value=val if val != "" else None).border = thin_border
                row_num += 1
        _auto_width(ws_q)

    # ---- Sheet 6+: Sample Data ----
    if data["include_sample_data"]:
        for schema in schemas:
            if not schema.sample_data:
                continue
            sheet_name = f"Sample — {schema.table_name}"[:31]  # Excel 31-char limit
            ws_s = wb.create_sheet(sheet_name)
            headers = list(schema.sample_data[0].keys())
            _write_headers(ws_s, headers)
            for r, row_data in enumerate(schema.sample_data, 2):
                for ci, h in enumerate(headers, 1):
                    ws_s.cell(row=r, column=ci, value=_fmt(row_data.get(h))).border = thin_border
            _auto_width(ws_s)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Markdown export
# ---------------------------------------------------------------------------

def export_to_markdown(data: dict[str, Any], output_path: Path) -> None:
    """Write data dictionary to a single .md file."""
    db_info: DatabaseInfo = data["database"]
    schemas: list[SchemaInfo] = data["schemas"]
    profiles: dict[str, TableProfile] = data["profiles"]
    quality: dict[str, QualityReport] = data["quality"]
    ctx = _get_db_context(db_info)

    lines: list[str] = []

    def _md_table(headers: list[str], rows: list[list[str]]) -> None:
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join("---" for _ in headers) + " |")
        for row in rows:
            lines.append("| " + " | ".join(str(v) for v in row) + " |")
        lines.append("")

    # Header
    lines.append(f"# Data Dictionary: {db_info.name}")
    lines.append("")
    key_ent = ", ".join(ctx["key_entities"]) if isinstance(ctx["key_entities"], list) else str(ctx["key_entities"])
    lines.append(
        f"**Domain:** {ctx['domain']} | "
        f"**Purpose:** {ctx['purpose']} | "
        f"**Tables:** {len(schemas)} | "
        f"**Last Synced:** {db_info.last_synced_at}"
    )
    lines.append("")

    if key_ent:
        lines.append(f"**Key Entities:** {key_ent}")
        lines.append("")

    lines.append(f"**Exported At:** {datetime.now().isoformat()}")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Per-table sections
    for schema in schemas:
        profile = profiles.get(schema.table_name)
        report = quality.get(schema.table_name)

        lines.append(f"## Table: {schema.table_name}")
        lines.append("")

        purpose = _table_purpose(schema)
        if purpose:
            lines.append(f"**Purpose:** {purpose}")
            lines.append("")
        bctx = _business_context(schema)
        if bctx:
            lines.append(f"**Business Context:** {bctx}")
            lines.append("")
        if profile:
            lines.append(f"**Row Count:** {profile.row_count}")
            lines.append("")

        # Columns
        lines.append("### Columns")
        lines.append("")
        col_headers = ["Column", "Type", "Nullable", "Description"]
        if data["include_profile"]:
            col_headers += ["Null%", "Distinct%", "Role"]

        profile_map: dict[str, Any] = {}
        if profile:
            profile_map = {cp.column_name: cp for cp in profile.columns}

        col_rows = []
        for col in schema.original_metadata.get("columns", []):
            col_name = col.get("column_name", "")
            row = [
                col_name,
                col.get("data_type", ""),
                "Yes" if col.get("nullable") else "No",
                _col_description(schema, col_name),
            ]
            if data["include_profile"]:
                cp = profile_map.get(col_name)
                if cp:
                    row += [
                        f"{cp.null_percentage:.1f}",
                        f"{cp.distinct_percentage:.1f}",
                        cp.inferred_role or "",
                    ]
                else:
                    row += ["", "", ""]
            col_rows.append(row)
        _md_table(col_headers, col_rows)

        # Constraints
        constraints = _constraints(schema)
        if constraints:
            lines.append("### Constraints")
            lines.append("")
            _md_table(
                ["Type", "Column", "References"],
                [[c["type"], c["column"], c["references"]] for c in constraints],
            )

        # Relationships
        rels = _relationships(schema)
        if rels:
            lines.append("### Relationships")
            lines.append("")
            for rel in rels:
                lines.append(f"- {rel}")
            lines.append("")

        # Quality
        if data["include_quality"] and report:
            lines.append(f"### Quality (Score: {report.overall_score:.0f}/100)")
            lines.append("")
            _md_table(
                ["Dimension", "Score"],
                [
                    ["Completeness", f"{report.completeness_score:.1f}"],
                    ["Uniqueness", f"{report.uniqueness_score:.1f}"],
                    ["Validity", f"{report.validity_score:.1f}"],
                    ["Consistency", f"{report.consistency_score:.1f}"],
                ],
            )
            if report.issues:
                lines.append("**Issues:**")
                lines.append("")
                _md_table(
                    ["Severity", "Category", "Column", "Issue", "Recommendation"],
                    [
                        [
                            issue.severity,
                            issue.category,
                            issue.column_name or "",
                            issue.issue,
                            issue.recommendation,
                        ]
                        for issue in report.issues
                    ],
                )

        # Sample data
        if data["include_sample_data"] and schema.sample_data:
            lines.append("### Sample Data")
            lines.append("")
            headers = list(schema.sample_data[0].keys())
            sample_rows = [[_fmt(row.get(h)) for h in headers] for row in schema.sample_data]
            _md_table(headers, sample_rows)

        lines.append("---")
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")

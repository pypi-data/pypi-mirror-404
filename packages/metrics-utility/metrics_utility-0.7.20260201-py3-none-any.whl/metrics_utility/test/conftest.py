import datetime as dt_actual
import os

from unittest.mock import MagicMock

import openpyxl
import pandas as pd
import pytest

from metrics_utility import prepare
from metrics_utility.automation_controller_billing.dataframe_engine.db_dataframe_host_metric import DBDataframeHostMetric
from metrics_utility.test.util import generate_renewal_guidance_dataframe


# this updates python paths to include mock_awx/ (and not fail to find awx imports)
# has to happen before any tests that import code that imports awx are imported
prepare()


@pytest.fixture
def fixed_now():
    """
    Provides a fixed, timezone-aware datetime for deterministic tests.
    """
    return dt_actual.datetime(2025, 6, 3, 10, 0, 0, tzinfo=dt_actual.timezone.utc)


@pytest.fixture
def setup_processed_dataframe(fixed_now):
    """
    Sets up a processed Pandas DataFrame ready for use by report methods.
    It mocks the DBDataframeHostMetric's data extraction to provide consistent data.
    """
    mock_full_dataframe_raw = generate_renewal_guidance_dataframe(is_empty=False, current_datetime=fixed_now)

    mock_df_for_batch_processed = mock_full_dataframe_raw.copy()
    for col in ['first_automation', 'last_automation']:
        mock_df_for_batch_processed[col] = pd.to_datetime(mock_df_for_batch_processed[col], utc=True).dt.tz_localize(None)

    mock_df_for_batch_processed['last_deleted'] = mock_df_for_batch_processed['last_deleted'].apply(
        lambda x: (x.isoformat(timespec='seconds') if pd.notna(x) and isinstance(x, dt_actual.datetime) else None)
    )

    mock_batches = [{'host_metric': mock_df_for_batch_processed}]

    mock_extractor = MagicMock()
    mock_extractor.iter_batches.return_value = (batch for batch in mock_batches)

    db_host_metric_instance = DBDataframeHostMetric(
        extractor=mock_extractor,
        month=fixed_now.strftime('%Y-%m'),
        extra_params={},
    )

    processed_df = db_host_metric_instance.build_dataframe()

    # Assertions to confirm the generated DataFrame is as expected
    assert processed_df is not None, 'Fixture: build_dataframe should return a DataFrame'
    assert not processed_df.empty, 'Fixture: DataFrame should not be empty'
    assert len(processed_df) == len(mock_full_dataframe_raw), 'Fixture: DataFrame should contain all mock rows'
    assert pd.api.types.is_datetime64_any_dtype(processed_df['first_automation'])
    assert pd.api.types.is_datetime64_any_dtype(processed_df['last_automation'])
    assert pd.api.types.is_datetime64_any_dtype(processed_df['last_deleted'])

    yield processed_df


def validate_sheet_tab_names(file_path, expected_sheets, excluded_sheets=[]):
    """Test the sheet names in the Excel file."""

    wb = openpyxl.load_workbook(file_path)
    try:
        actual_tab_names = wb.sheetnames
        actual_tab_names = [name for name in actual_tab_names if name not in excluded_sheets]
        assert actual_tab_names == list(expected_sheets.keys()), 'Sheet names do not match.'
    finally:
        wb.close()


def normalize_column(col):
    """Remove whitespace, newlines, and uppercase chars from column name."""
    if not col:
        return ''
    return col.strip().replace('\n', ' ').lower()


def validate_sheet_columns(file_path, expected_sheets, usage_reporting_min_row):
    """Test the column names for each sheet."""

    # Determine the min_row (first row) to identify.
    def get_min_row(sheet_name):
        return usage_reporting_min_row if sheet_name == 'Usage Reporting' else 1

    # Determine the expected column headers.
    def get_column_headers(expected_column_data):
        expected_column_headers = []
        for column_group in expected_column_data:
            expected_column_headers.extend(normalize_column(col) for col in column_group.keys())
        return expected_column_headers

    wb = openpyxl.load_workbook(file_path)
    try:
        for sheet_name, expected_column_data in expected_sheets.items():
            sheet = wb[sheet_name]

            min_row = get_min_row(sheet_name)

            # Determine all actual column headers.
            actual_column_headers = [normalize_column(cell.value) for cell in next(sheet.iter_rows(min_row=min_row, max_row=min_row))]

            # Call the get_column_headers() method and assign return value to expected_column_headers.
            expected_column_headers = get_column_headers(expected_column_data)

            # Assert the actual headers equal expected headers.
            assert actual_column_headers == expected_column_headers, f'Column names do not match for sheet: {sheet_name}'

            # Iterate through each expected column group
            for column_group in expected_column_data:
                for expected_col_name, expected_column_values in column_group.items():
                    # Find the actual column index for this column
                    try:
                        col_index = actual_column_headers.index(normalize_column(expected_col_name)) + 1
                    except ValueError:
                        raise AssertionError(f"Expected column '{expected_col_name}' not found in actual columns for sheet: {sheet_name}")

                    # Extract actual values for this column (skip the header).
                    actual_column_values = [cell.value for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index) for cell in row]

                    # Assert that the actual column values equal the expected column values.
                    assert actual_column_values == expected_column_values, (
                        f"Column values do not match for column '{expected_col_name}' in sheet '{sheet_name}'"
                    )
    finally:
        wb.close()


@pytest.fixture
def cleanup(request):
    file_path = request.param
    """Fixture to clean up the generated file at the start and end of test."""
    # Cleanup at the beginning
    if os.path.exists(file_path):
        os.remove(file_path)
    yield
    # Cleanup at the end
    if os.path.exists(file_path):
        os.remove(file_path)


def transform_sheet(sheet):
    """
    Transforms a sheet dictionary in column-wise format into a row-wise dictionary.

    Parameters:
        sheet (dict): A dictionary where keys are column names and values are dictionaries
                      mapping row indices to cell values.

    Returns:
        dict: A dictionary where each key is a row index and each value is a dictionary mapping
              column names to cell values for that row.
    """
    rows = {}
    # Iterate over each column and its data
    for col, col_data in sheet.items():
        col = col.replace('\n', ' ')
        # For each row in the column
        for row_index, value in col_data.items():
            # Initialize the row if it hasn't been created yet
            if row_index not in rows:
                rows[row_index] = {}
            # Set the value for the column in that row
            rows[row_index][col] = value
    return rows


def validate_cell(wb, sheet_name, address, value):
    sheet = wb[sheet_name]
    assert sheet[address].value == value, f'Sheet: {sheet_name}, cell: {address}, expected value: {value}, actual value: {sheet[address].value}'

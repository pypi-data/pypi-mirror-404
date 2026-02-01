######################################
# Code for building the spreadsheet
######################################
import json

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from metrics_utility.automation_controller_billing.dataframe_engine.base import merge_sets
from metrics_utility.automation_controller_billing.helpers import merge_arrays, merge_json_sets
from metrics_utility.metric_utils import INDIRECT


class Base:
    BLACK_COLOR_HEX = '00000000'
    WHITE_COLOR_HEX = '00FFFFFF'
    BLUE_COLOR_HEX = '000000FF'
    RED_COLOR_HEX = 'FF0000'
    LIGHT_BLUE_COLOR_HEX = 'd4eaf3'
    GREEN_COLOR_HEX = '92d050'
    YELLOW_COLOR_HEX = 'ffcc17'

    FONT = 'Arial'
    PRICE_FORMAT = '$#,##0.00'
    HOST_NAME = 'Host name'
    JOB_RUNS = 'Job runs'
    NUM_OF_TASKS_OR_RUNS = 'Number of task\nruns'
    HOST_RUNS_UNIQUE = 'Unique managed nodes\nautomated'
    HOST_RUNS = 'Non-unique managed\nnodes automated'
    DURATION = 'Duration of task\nruns [seconds]'

    # Deduplication column labels
    DEDUP_COLUMN_LABELS = {
        'host_names_before_dedup': 'Host names\nbefore\ndeduplication',
        'host_names_before_dedup_count': 'Host names\nbefore\ndeduplication\ncount',
    }

    def optional_report_sheets(self):
        return self.extra_params.get('optional_sheets')

    def has_dedup_enabled(self):
        """Check if experimental deduplication is enabled."""
        return self.extra_params.get('deduplicator') == 'ccsp-experimental'

    def calculate_dedup_count(self, series):
        """Calculate count for dedup column more efficiently."""
        return series.map(lambda x: len(x) if isinstance(x, (set, list)) else 1)

    def add_dedup_count_column(self, dataframe, base_col_name, count_col_name):
        """Add deduplication count column if base column exists."""
        if base_col_name in dataframe.columns:
            dataframe[count_col_name] = self.calculate_dedup_count(dataframe[base_col_name])
        return dataframe

    def handle_dedup_columns_for_scope(self, dataframe, columns, convert_cols):
        """Handle deduplication columns for inventory scope if experimental dedup is enabled."""
        if self.has_dedup_enabled() and 'host_names_before_dedup' in dataframe.columns:
            convert_cols.append('host_names_before_dedup')
            self.add_dedup_count_column(dataframe, 'host_names_before_dedup', 'host_names_before_dedup_count')
            columns += ['host_names_before_dedup', 'host_names_before_dedup_count']
        return columns, convert_cols

    def handle_dedup_aggregation(self, agg_dict):
        """Add deduplication aggregation if experimental dedup is enabled."""
        if self.has_dedup_enabled():
            # Use merge_sets since the data already contains sets from initial aggregation
            agg_dict['host_names_before_dedup'] = ('host_names_before_dedup', merge_sets)
        return agg_dict

    def handle_dedup_columns_for_usage(self, dataframe, columns, convert_cols):
        """Handle deduplication columns for usage tables if experimental dedup is enabled."""
        if self.has_dedup_enabled():
            self.add_dedup_count_column(dataframe, 'host_names_before_dedup', 'host_names_before_dedup_count')
            columns += ['host_names_before_dedup', 'host_names_before_dedup_count']
            if 'host_names_before_dedup' in dataframe.columns:
                convert_cols.append('host_names_before_dedup')
        return columns, convert_cols

    def add_dedup_labels_if_needed(self, labels, column_names):
        """Add deduplication labels for specified columns if experimental dedup is enabled."""
        if self.has_dedup_enabled():
            labels.update({k: v for k, v in self.DEDUP_COLUMN_LABELS.items() if k in column_names})
        return labels

    def convert_cell(self, cell):
        # If the cell is a dictionary, convert each set value to a sorted list, then dump as a JSON string.
        if isinstance(cell, dict):
            new_cell = {k: sorted(list(v)) if isinstance(v, set) else v for k, v in cell.items()}
            return json.dumps(new_cell)
        # If the cell itself is a set, convert it to a sorted list and then to a JSON string.
        elif isinstance(cell, set):
            return json.dumps(sorted(list(cell)))
        # If the cell is a list, convert any set elements inside to sorted lists and dump as a JSON string.
        elif isinstance(cell, list):
            new_cell = [sorted(list(item)) if isinstance(item, set) else item for item in cell]
            # Sort the list itself if it contains strings
            if new_cell and all(isinstance(item, str) for item in new_cell):
                new_cell = sorted(new_cell)
            return json.dumps(new_cell)
        # Otherwise, return the cell unchanged.
        return cell

    def add_sheet(self, title, sheet_index, widths=None):
        self.wb.create_sheet(title=title)
        ws = self.wb.worksheets[sheet_index]
        if widths:
            self.set_widths(ws, widths)
        return ws

    def set_widths(self, ws, widths):
        for key, value in widths.items():
            ws.column_dimensions[get_column_letter(key)].width = value

    def _fix_event_host_names(self, mapping_dataframe, destination_dataframe):
        if destination_dataframe is None:
            return None

        def concatenate_columns_mapping(row):
            return f'{row["original_host_name"]}__{str(row["install_uuid"])}__{str(row["job_remote_id"])}'

        def concatenate_columns_destination(row):
            return f'{row["host_name"]}__{str(row["install_uuid"])}__{str(row["job_remote_id"])}'

        # Apply the function to each row of the DataFrame
        mapping_dataframe['host_composite_id'] = mapping_dataframe.apply(concatenate_columns_mapping, axis=1)
        mapping_dataframe = mapping_dataframe.set_index('host_composite_id')
        mapping_dataframe = mapping_dataframe['host_name'].astype(str).to_dict()

        def apply_mapping(row):
            return mapping_dataframe.get(f'{row["host_name"]}__{str(row["install_uuid"])}__{row["job_remote_id"]}', row['host_name'])

        destination_dataframe['host_name'] = destination_dataframe.apply(apply_mapping, axis=1)
        destination_dataframe['host_composite_id'] = destination_dataframe.apply(concatenate_columns_destination, axis=1)

        return destination_dataframe

    def _build_data_section_scope(self, current_row, ws, dataframe, mode=None):
        header_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX, bold=True)
        value_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX)

        ccsp_report_dataframe = dataframe.copy()

        # Convert arrays and dict fields into string, so they can be rendered into xlsx
        convert_cols = ['organizations', 'inventories', 'canonical_facts', 'facts']
        columns = [
            'host_name',
            'last_automation',
            'organizations',
            'inventories',
            'canonical_facts',
            'facts',
        ]

        # Handle deduplication columns if enabled
        columns, convert_cols = self.handle_dedup_columns_for_scope(ccsp_report_dataframe, columns, convert_cols)

        for col in convert_cols:
            if col in ccsp_report_dataframe.columns:
                ccsp_report_dataframe[col] = ccsp_report_dataframe[col].apply(self.convert_cell)

        # We're not showing cluster/install_uuid until we support multi-cluster view officially
        if 'install_uuid' in ccsp_report_dataframe.columns:
            del ccsp_report_dataframe['install_uuid']

        labels = {
            'host_name': self.HOST_NAME,
            'last_automation': 'Last\nAutomation',
            'organizations': 'Organizations',
            'inventories': 'Inventories',
            'canonical_facts': 'Canonical Facts',
            'facts': 'Facts',
        }

        # Add dedup labels if needed
        self.add_dedup_labels_if_needed(labels, ['host_names_before_dedup', 'host_names_before_dedup_count'])

        # Filter columns that exist
        columns = [col for col in columns if col in ccsp_report_dataframe.columns]
        ccsp_report_dataframe = ccsp_report_dataframe[columns]

        labels = {k: v for k, v in labels.items() if k in columns}
        ccsp_report_dataframe = ccsp_report_dataframe.rename(columns=labels)

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_infrastructure_summary(self, current_row, ws, dataframe):
        header_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX, bold=True)
        value_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX)

        # Extract infrastructure facts from indirect nodes
        indirect_nodes = dataframe[dataframe['managed_node_type'] == INDIRECT].copy()

        if indirect_nodes.empty:
            # If no indirect nodes, show empty message
            cell = ws.cell(row=current_row, column=1)
            cell.value = 'No indirect nodes found'
            cell.font = value_font
            return current_row + 1

        # Parse facts to extract infra_type, infra_bucket, and device_type
        def extract_infra_info(facts):
            def extract_value(value):
                if isinstance(value, set):
                    return list(value)[0] if value else 'Unknown'
                elif isinstance(value, list):
                    return value[0] if value else 'Unknown'
                elif isinstance(value, str):
                    return value
                else:
                    return 'Unknown'

            if isinstance(facts, dict):
                return {
                    'infra_type': extract_value(facts.get('infra_type', 'Unknown')),
                    'infra_bucket': extract_value(facts.get('infra_bucket', 'Unknown')),
                    'device_type': extract_value(facts.get('device_type', 'Unknown')),
                }
            elif isinstance(facts, str):
                import json

                try:
                    facts_dict = json.loads(facts)
                    return {
                        'infra_type': extract_value(facts_dict.get('infra_type', 'Unknown')),
                        'infra_bucket': extract_value(facts_dict.get('infra_bucket', 'Unknown')),
                        'device_type': extract_value(facts_dict.get('device_type', 'Unknown')),
                    }
                except Exception:
                    pass
            return {'infra_type': 'Unknown', 'infra_bucket': 'Unknown', 'device_type': 'Unknown'}

        # Extract infrastructure information
        infra_info = indirect_nodes['facts'].apply(extract_infra_info)
        indirect_nodes['infra_type'] = infra_info.apply(lambda x: x['infra_type'])
        indirect_nodes['infra_bucket'] = infra_info.apply(lambda x: x['infra_bucket'])
        indirect_nodes['device_type'] = infra_info.apply(lambda x: x['device_type'])

        # Group by infra_type, infra_bucket, and device_type
        agg_dict = {
            'indirect_hosts_unique': ('host_name', 'nunique'),
            'indirect_hosts_total': ('host_name', 'count'),
        }

        summary_df = indirect_nodes.groupby(['infra_type', 'infra_bucket', 'device_type'], dropna=False).agg(**agg_dict)
        summary_df = summary_df.reset_index()

        # Sort by infrastructure type, then bucket, then device type
        summary_df = summary_df.sort_values(['infra_type', 'infra_bucket', 'device_type'])

        # Add all column headers
        headers = ['Infrastructure', 'Device Category', 'Device Type', 'Unique Nodes', 'Total Nodes']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left')

        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Create hierarchical display
        prev_infra_type = None
        prev_infra_bucket = None

        for _, row in summary_df.iterrows():
            infra_type = row['infra_type']
            infra_bucket = row['infra_bucket']
            device_type = row['device_type']

            # Write infrastructure type header when it changes
            if infra_type != prev_infra_type:
                # Merge cells across three columns for infrastructure type
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                cell = ws.cell(row=current_row, column=1)
                cell.value = infra_type
                cell.font = header_font
                cell.alignment = Alignment(horizontal='left')
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                prev_infra_type = infra_type
                prev_infra_bucket = None  # Reset bucket tracking

            # Write infrastructure bucket header when it changes
            if infra_bucket != prev_infra_bucket:
                cell = ws.cell(row=current_row, column=2)
                cell.value = infra_bucket
                cell.font = header_font
                cell.alignment = Alignment(horizontal='left')
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                prev_infra_bucket = infra_bucket

            # Write device type row
            cell = ws.cell(row=current_row, column=3)
            cell.value = device_type
            cell.font = value_font

            cell = ws.cell(row=current_row, column=4)
            cell.value = row['indirect_hosts_unique']
            cell.font = value_font

            cell = ws.cell(row=current_row, column=5)
            cell.value = row['indirect_hosts_total']
            cell.font = value_font

            current_row += 1

        return current_row

    def _build_data_section_usage_by_node(self, current_row, ws, dataframe, mode=None, managed_node_type=None):
        header_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX, bold=True)
        value_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX)

        agg_dict = {
            'organizations': ('organization_name', 'nunique'),
            'host_runs': ('host_name', 'count'),
            'task_runs': ('task_runs', 'sum'),
            'first_automation': ('first_automation', 'min'),
            'last_automation': ('last_automation', 'max'),
            'managed_node_types_set': ('managed_node_types_set', lambda x: merge_arrays(x)),
            'events': ('events', lambda x: merge_arrays(x)),
            'canonical_facts': ('canonical_facts', lambda x: merge_json_sets(x)),
            'facts': ('facts', lambda x: merge_json_sets(x)),
        }

        # Handle deduplication aggregation if enabled
        self.handle_dedup_aggregation(agg_dict)

        # Now pass this dictionary into .agg()
        ccsp_report_dataframe = dataframe.groupby('host_name', dropna=False).agg(**agg_dict)

        # Convert arrays and dict fields into string, so they can be rendered into xlsx
        convert_cols = ['managed_node_types_set', 'events', 'canonical_facts', 'facts']

        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()
        columns = [
            'host_name',
            'organizations',
            'host_runs',
            'task_runs',
            'first_automation',
            'last_automation',
        ]

        # Add facts and canonical facts for managed nodes (direct) when experimental dedup is enabled
        # or always for indirect nodes
        if managed_node_type == 'indirect' or (managed_node_type == 'direct' and self.has_dedup_enabled()):
            columns += ['canonical_facts', 'facts']

        if managed_node_type == 'indirect':
            columns += ['managed_node_types_set', 'events']

        # Handle deduplication columns if enabled
        columns, convert_cols = self.handle_dedup_columns_for_usage(ccsp_report_dataframe, columns, convert_cols)

        for col in convert_cols:
            if col in ccsp_report_dataframe.columns:
                ccsp_report_dataframe[col] = ccsp_report_dataframe[col].apply(self.convert_cell)

        if mode == 'by_organization':
            # Filter some columns out based on mode
            columns = [col for col in columns if col not in ['organizations']]
        ccsp_report_dataframe = ccsp_report_dataframe.reindex(columns=columns)

        labels = {
            'host_name': self.HOST_NAME,
            'organizations': 'Automated by\norganizations',
            'host_runs': self.JOB_RUNS,  # Job runs is the same as host_runs, Non-unique managed nodes automated
            'task_runs': self.NUM_OF_TASKS_OR_RUNS,
            'first_automation': 'First\nautomation',
            'last_automation': 'Last\nautomation',
            'canonical_facts': 'Canonical\nFacts',
            'facts': 'Facts',
        }
        if managed_node_type == 'indirect':
            labels.update(
                {
                    'managed_node_types_set': 'Manage\nNode\nTypes',
                    'events': 'Events',
                }
            )

        # Add deduplication labels if needed
        self.add_dedup_labels_if_needed(labels, ['host_names_before_dedup', 'host_names_before_dedup_count'])

        labels = {k: v for k, v in labels.items() if k in columns}
        ccsp_report_dataframe = ccsp_report_dataframe.rename(columns=labels)

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_collections(self, current_row, ws, dataframe):
        header_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX, bold=True)
        value_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX)

        # Take the content explorer dataframe and extract specific group by
        agg_dict = {
            'host_runs_unique': ('host_name', 'nunique'),
            'host_runs': ('host_composite_id', 'nunique'),
            'task_runs': ('task_runs', 'sum'),
            'duration': ('duration', 'sum'),
        }

        ccsp_report_dataframe = dataframe.groupby(['collection_name'], dropna=False).agg(**agg_dict)

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        rename_columns = {
            'collection_name': 'Collection name',
            'host_runs_unique': self.HOST_RUNS_UNIQUE,
            'host_runs': self.HOST_RUNS,
            'task_runs': self.NUM_OF_TASKS_OR_RUNS,
            'duration': self.DURATION,
        }

        ccsp_report_dataframe = ccsp_report_dataframe.rename(columns=rename_columns)

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_roles(self, current_row, ws, dataframe):
        header_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX, bold=True)
        value_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX)

        # Take the content explorer dataframe and extract specific group by
        agg_dict = {
            'host_runs_unique': ('host_name', 'nunique'),
            'host_runs': ('host_composite_id', 'nunique'),
            'task_runs': ('task_runs', 'sum'),
            'duration': ('duration', 'sum'),
        }

        ccsp_report_dataframe = dataframe.groupby(['role_name'], dropna=False).agg(**agg_dict)

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        rename_columns = {
            'role_name': 'Role name',
            'host_runs_unique': self.HOST_RUNS_UNIQUE,
            'host_runs': self.HOST_RUNS,
            'task_runs': self.NUM_OF_TASKS_OR_RUNS,
            'duration': self.DURATION,
        }

        ccsp_report_dataframe = ccsp_report_dataframe.rename(columns=rename_columns)

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                # cell.border = dotted_border

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_modules(self, current_row, ws, dataframe):
        header_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX, bold=True)
        value_font = Font(name=self.FONT, size=10, color=self.BLACK_COLOR_HEX)

        # Take the content explorer dataframe and extract specific group by
        agg_dict = {
            'host_runs_unique': ('host_name', 'nunique'),
            'host_runs': ('host_composite_id', 'nunique'),
            'task_runs': ('task_runs', 'sum'),
            'duration': ('duration', 'sum'),
        }

        ccsp_report_dataframe = dataframe.groupby(['module_name'], dropna=False).agg(**agg_dict)

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        rename_columns = {
            'module_name': 'Module name',
            'host_runs_unique': self.HOST_RUNS_UNIQUE,
            'host_runs': self.HOST_RUNS,
            'task_runs': self.NUM_OF_TASKS_OR_RUNS,
            'duration': self.DURATION,
        }

        ccsp_report_dataframe = ccsp_report_dataframe.rename(columns=rename_columns)

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

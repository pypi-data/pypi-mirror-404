"""Helper functions for CCSPv2 deduplication tests."""

import hashlib
import shutil

from pathlib import Path

import openpyxl


def transform_sheet(sheet):
    """
    Transforms a sheet dictionary in column-wise format into a row-wise dictionary.
    Handles mixed data types and malformed data gracefully.
    """
    if not isinstance(sheet, dict):
        print(f'⚠ transform_sheet received non-dict data: {type(sheet)}')
        return {}

    rows = {}
    # Iterate over each column and its data
    for col, col_data in sheet.items():
        col = col.replace('\n', ' ')

        # Handle cases where col_data is not a dictionary
        if not isinstance(col_data, dict):
            print(f"⚠ Column '{col}' has non-dict data: {type(col_data)}={col_data}")
            continue

        # For each row in the column
        for row_index, value in col_data.items():
            # Initialize the row if it hasn't been created yet
            if row_index not in rows:
                rows[row_index] = {}
            # Set the value for the column in that row
            rows[row_index][col] = value
    return rows


def get_xlsx_content_hash(file_path):
    """Calculate hash of XLSX content (excluding metadata like timestamps)."""
    try:
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        hash_sha256 = hashlib.sha256()

        # Hash all sheet names and their content
        for sheet_name in sorted(workbook.sheetnames):
            hash_sha256.update(sheet_name.encode('utf-8'))
            sheet = workbook[sheet_name]

            # Hash all cell values
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        hash_sha256.update(str(cell.value).encode('utf-8'))

        workbook.close()
        return hash_sha256.hexdigest()
    except Exception:
        # Fallback to file hash if XLSX parsing fails
        return get_file_hash(file_path)


def get_file_hash(file_path):
    """Calculate SHA256 hash of file content."""
    hash_sha256 = hashlib.sha256()
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b''):
            hash_sha256.update(chunk)
    return hash_sha256.hexdigest()


def copy_if_content_changed(source_path, dest_path):
    """Copy file only if content has changed based on XLSX content comparison."""
    # If destination doesn't exist, copy the file
    if not dest_path.exists():
        shutil.copy2(source_path, dest_path)
        print(f'Created test report: {dest_path}')
        return

    # Calculate content hashes of both XLSX files (ignoring metadata)
    source_hash = get_xlsx_content_hash(source_path)
    dest_hash = get_xlsx_content_hash(dest_path)

    # Only copy if content is different
    if source_hash != dest_hash:
        shutil.copy2(source_path, dest_path)
        print(f'Updated test report (content changed): {dest_path}')
    else:
        print(f'Test report unchanged: {dest_path}')


def sort_json_fields(obj):
    """Recursively sort JSON fields for consistent testing."""
    if isinstance(obj, dict):
        sorted_dict = {}
        for key in sorted(obj.keys()):
            value = obj[key]
            if isinstance(value, list):
                # Sort list values
                sorted_values = []
                for v in value:
                    if v is not None:
                        sorted_values.append(v)
                sorted_values.sort(key=lambda x: str(x))
                sorted_dict[key] = sorted_values
            else:
                sorted_dict[key] = sort_json_fields(value)
        return sorted_dict
    elif isinstance(obj, list):
        # Sort list elements
        sorted_list = []
        for item in obj:
            if item is not None:
                sorted_list.append(sort_json_fields(item))
        sorted_list.sort(key=lambda x: str(x))
        return sorted_list
    else:
        return obj


def get_test_dir():
    """Get the directory where this test file is located."""
    # Return the parent directory of conftest.py
    return Path(__file__).parent

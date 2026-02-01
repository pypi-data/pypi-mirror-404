import copy
import datetime as dt_actual

from unittest.mock import ANY, MagicMock, call, patch

import pandas as pd
import pytest

from openpyxl import Workbook as ActualOpenpyxlWorkbook

from metrics_utility.automation_controller_billing.report.report_renewal_guidance import ReportRenewalGuidance


# 1. Fixture to set up all mocks for build_spreadsheet's dependencies
@pytest.fixture
def setup_build_spreadsheet_mocks(fixed_now):
    """
    Prepares all mocks needed for the build_spreadsheet method.
    """
    mock_wb = MagicMock(spec=ActualOpenpyxlWorkbook)
    mock_active_ws = MagicMock()
    mock_wb.active = mock_active_ws
    mock_wb.remove.return_value = None

    mock_ws_list = [MagicMock(), MagicMock(), MagicMock(), MagicMock(), MagicMock()]
    mock_wb.create_sheet.side_effect = mock_ws_list

    mock_cell = MagicMock()

    mock_time_module = MagicMock()
    mock_time_module.strftime.return_value = 'Jun 03, 2025'

    mocks_for_instance = {
        '_build_heading_h1': MagicMock(return_value=2),
        '_build_header': MagicMock(return_value=3),
        '_build_data_section': MagicMock(),
        '_build_data_section_host_metrics': MagicMock(),
        '_build_data_section_ephemeral_usage': MagicMock(),
        'compute_ephemeral_intervals': MagicMock(return_value=pd.DataFrame()),
        'optional_report_sheets': MagicMock(return_value=['managed_nodes']),
        'dataframe_to_rows_func': MagicMock(return_value=[['mock_header'], ['mock_data']]),
    }

    yield {
        'openpyxl_wb_mock': mock_wb,
        'time_module_mock': mock_time_module,
        'mocks_for_instance': mocks_for_instance,
        'mock_cell': mock_cell,
        'mock_ws_list': mock_ws_list,
    }


# 4. Fixture to set up a ReportRenewalGuidance instance for tests
@pytest.fixture
def setup_report_renewal_guidance_instance(fixed_now, setup_build_spreadsheet_mocks, setup_processed_dataframe):
    patch_target_datetime = 'metrics_utility.automation_controller_billing.report.report_renewal_guidance.datetime'
    patch_target_workbook = 'metrics_utility.automation_controller_billing.report.report_renewal_guidance.Workbook'
    patch_target_time = 'metrics_utility.automation_controller_billing.report.report_renewal_guidance.time'
    patch_target_dataframe_to_rows = 'metrics_utility.automation_controller_billing.report.report_renewal_guidance.dataframe_to_rows'

    with (
        patch(patch_target_datetime) as mock_datetime_module,
        patch(patch_target_workbook, new=setup_build_spreadsheet_mocks['openpyxl_wb_mock']),
        patch(patch_target_time, new=setup_build_spreadsheet_mocks['time_module_mock']),
        patch(patch_target_dataframe_to_rows, new=setup_build_spreadsheet_mocks['mocks_for_instance']['dataframe_to_rows_func']),
    ):
        mock_datetime_module.datetime.now.return_value = fixed_now
        mock_datetime_module.timedelta = dt_actual.timedelta
        mock_datetime_module.timezone = MagicMock(spec=dt_actual.timezone)
        mock_datetime_module.timezone.utc = dt_actual.timezone.utc

        test_extra_params = {
            'ephemeral_days': 30,
            'price_per_node': 0.1,
            'report_period': '2025-01-01,2025-06-03',
            'report_type': 'RENEWAL_GUIDANCE',
            'since_date': '2025-01-01',
            'until_date': '2025-06-03',
        }

        report_instance = ReportRenewalGuidance(
            dataframes={'host_metric': setup_processed_dataframe},
            extra_params=test_extra_params,
        )

        report_instance.wb = setup_build_spreadsheet_mocks['openpyxl_wb_mock']

        for attr, mock_obj in setup_build_spreadsheet_mocks['mocks_for_instance'].items():
            if attr == 'dataframe_to_rows_func':
                continue
            setattr(report_instance, attr, mock_obj)

        yield {
            'report_instance': report_instance,
            'test_extra_params': test_extra_params,
            'processed_df': setup_processed_dataframe,
            'mocks': setup_build_spreadsheet_mocks,
        }


def test_build_spreadsheet_with_ephemeral_data(
    setup_report_renewal_guidance_instance,
):
    """
    Tests the build_spreadsheet method when 'ephemeral_days' is set,
    expecting all ephemeral-related sheets and calculations.
    """
    fixture_context = setup_report_renewal_guidance_instance
    report_instance = fixture_context['report_instance']
    processed_df = fixture_context['processed_df']
    mocks = fixture_context['mocks']

    mock_ephemeral_usage_df = pd.DataFrame(
        {
            'window_start': [dt_actual.datetime(2025, 1, 1)],
            'window_end': [dt_actual.datetime(2025, 1, 30)],
            'ephemeral_hosts': [5],
        }
    )
    mocks['mocks_for_instance']['compute_ephemeral_intervals'].return_value = mock_ephemeral_usage_df
    mocks['mocks_for_instance']['optional_report_sheets'].return_value = ['managed_nodes']

    result_wb = report_instance.build_spreadsheet()

    mocks['openpyxl_wb_mock'].remove.assert_called_once_with(mocks['openpyxl_wb_mock'].active)

    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Usage Reporting')

    mocks['mocks_for_instance']['_build_heading_h1'].assert_called_once_with(1, ANY)
    mocks['mocks_for_instance']['_build_header'].assert_called_once_with(ANY, ANY)
    mocks['mocks_for_instance']['_build_data_section'].assert_called_once_with(
        ANY,
        ANY,
        processed_df,
        mock_ephemeral_usage_df,
    )
    mocks['time_module_mock'].strftime.assert_called_once_with('%b %d, %Y')

    mocks['mocks_for_instance']['compute_ephemeral_intervals'].assert_called_once()

    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Managed nodes')
    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Managed nodes ephemeral')
    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Managed nodes ephemeral usage')
    mocks['mocks_for_instance']['_build_data_section_host_metrics'].assert_any_call(ANY, ANY, ANY)
    mocks['mocks_for_instance']['_build_data_section_ephemeral_usage'].assert_called_once_with(ANY, ANY, mock_ephemeral_usage_df)

    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Deleted Managed nodes')
    mocks['mocks_for_instance']['_build_data_section_host_metrics'].assert_any_call(ANY, ANY, ANY)

    assert result_wb is mocks['openpyxl_wb_mock']

    mocks['mocks_for_instance']['optional_report_sheets'].assert_called_once()


def test_build_spreadsheet_without_ephemeral_data(
    setup_report_renewal_guidance_instance,
):
    """
    Tests the build_spreadsheet method when 'ephemeral_days' is None,
    expecting no ephemeral-related sheets or calculations.
    """
    fixture_context = setup_report_renewal_guidance_instance
    processed_df = fixture_context['processed_df']
    mocks = fixture_context['mocks']

    original_extra_params = fixture_context['test_extra_params']

    modified_extra_params = copy.deepcopy(original_extra_params)

    modified_extra_params['ephemeral_days'] = None

    report_instance = ReportRenewalGuidance(
        dataframes={'host_metric': processed_df},
        extra_params=modified_extra_params,
    )

    for attr, mock_obj in mocks['mocks_for_instance'].items():
        if attr == 'dataframe_to_rows_func':
            continue
        setattr(report_instance, attr, mock_obj)
    report_instance.wb = mocks['openpyxl_wb_mock']

    mocks['mocks_for_instance']['optional_report_sheets'].return_value = ['managed_nodes']

    result_wb = report_instance.build_spreadsheet()

    mocks['openpyxl_wb_mock'].remove.assert_called_once_with(mocks['openpyxl_wb_mock'].active)

    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Usage Reporting')

    mocks['mocks_for_instance']['_build_heading_h1'].assert_called_once_with(1, ANY)
    mocks['mocks_for_instance']['_build_header'].assert_called_once_with(ANY, ANY)
    mocks['mocks_for_instance']['_build_data_section'].assert_called_once_with(
        ANY,
        ANY,
        processed_df,
        None,  # ephemeral_usage_dataframe should be None here
    )
    mocks['time_module_mock'].strftime.assert_called_once_with('%b %d, %Y')

    mocks['mocks_for_instance']['compute_ephemeral_intervals'].assert_not_called()

    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Managed nodes')
    assert call(title='Managed nodes ephemeral') not in mocks['openpyxl_wb_mock'].create_sheet.call_args_list
    assert call(title='Managed nodes ephemeral usage') not in mocks['openpyxl_wb_mock'].create_sheet.call_args_list

    mocks['mocks_for_instance']['_build_data_section_host_metrics'].assert_any_call(ANY, ANY, ANY)
    mocks['mocks_for_instance']['_build_data_section_ephemeral_usage'].assert_not_called()

    mocks['openpyxl_wb_mock'].create_sheet.assert_any_call(title='Deleted Managed nodes')
    mocks['mocks_for_instance']['_build_data_section_host_metrics'].assert_any_call(ANY, ANY, ANY)

    assert result_wb is mocks['openpyxl_wb_mock']

    mocks['mocks_for_instance']['optional_report_sheets'].assert_called_once()

import copy
import json
import os
import re
import shutil
import subprocess
import sys
import warnings

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

import openpyxl
import openpyxl.utils

from metrics_utility.logger import logger


warnings.filterwarnings('ignore', category=ResourceWarning)


@dataclass
class DataShape:
    """
    A data class representing the structure required to run snapshot definitions.

    Attributes:
        env_vars (Dict[str, str]): A dictionary of environment variables.
        params (List[str]): A list of command-line arguments to be passed to the utility.
        custom_params (Dict[str, str]): A dictionary of custom parameters for controlling the run behavior.
                                        "run_command": if yes, it will runs the command using subprocess.
                                        In future we wants it to be also able to run function inside pytest so we can
                                        mock the datetime in order to test relative times like --since=5months
                                        "generated": date when it was generated
    """

    env_vars: Dict[str, str]
    params: List[str]
    custom_params: Dict[str, str]


def create_directory_if_not_exists(directory_path: str) -> None:
    """
    Creates a directory if it does not already exist.

    Args:
        directory_path (str): The path to the directory to create.

    Returns:
        None
    """
    try:
        os.makedirs(directory_path, exist_ok=True)
    except Exception as e:
        logger.error(f'Error creating directory: {e}')


def parse_json_file(file_path: str) -> dict | None:
    """
    Parses a JSON file and returns the data as a dictionary.

    Args:
        file_path (str): The path to the JSON file to parse.

    Returns:
        dict: The parsed JSON data, or None if an error occurred.
    """
    try:
        with open(file_path, 'r') as file:
            data = json.load(file)
            return data
    except json.JSONDecodeError as e:
        logger.error(f'Error decoding JSON from file {file_path}: {e}')
    except FileNotFoundError:
        logger.error(f'File not found: {file_path}')
    except Exception as e:
        logger.error(f'An error occurred: {e}')
    return None


def save_snapshot_definition(data: DataShape, path: str) -> None:
    """
    Saves the snapshot definition (DataShape) to a specified JSON file.

    Args:
        data (DataShape): The DataShape object to save.
        path (str): The path where the JSON file will be stored.

    Returns:
        None
    """
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        # Open the file in write mode
        with open(path, 'w') as json_file:
            # Serialize the dictionary into JSON and write to the file
            json.dump(data, json_file, indent=4)
            json_file.flush()  # Explicitly flush the buffer
        logger.info(f'Data successfully saved to {path}')
    except Exception as e:
        logger.error(f'An error occurred while saving the data: {e} into {path}')


def find_json_files(directory: str) -> List[Path]:
    """
    Finds all JSON files recursively in the given directory.

    Args:
        directory (str): The directory to search for JSON files.

    Returns:
        List[Path]: A list of Path objects pointing to JSON files.
    """
    return list(Path(directory).rglob('*.json'))


def run_and_generate_snapshot_definitions(directory: str) -> None:
    """
    Finds all JSON definitions in a directory, runs each snapshot definition, and generates report files.

    The generated report file is then moved to an 'report.xlsx' within a subdirectory named
    after the JSON file (minus the .json extension).

    Args:
        directory (str): The directory containing the JSON definition files.

    Returns:
        None
    """
    logger.debug(f'\nRun and generate snapshots from {directory}\n')
    json_files = find_json_files(directory)

    for json_file in json_files:
        logger.debug(f'Found {json_file}')
        output_dir = json_file.as_posix().removesuffix('.json')
        create_directory_if_not_exists(output_dir)

        data = parse_json_file(json_file)
        generated_file = run_snapshot_definition(data)
        output_file = output_dir + '/report.xlsx'

        if os.path.exists(output_file):
            os.remove(output_file)

        shutil.move(generated_file, output_file)
        logger.info(f'Report generated and moved from {generated_file} to {output_file}\n')


def run_snapshot_definition(data: DataShape) -> str:
    """
    Runs the report generation process for a given snapshot definition dictionary.

    It sets the environment variables, constructs command-line parameters, and calls
    a subprocess to run the generation command if 'run_command' is set to 'Yes'.

    Args:
        data (dict): A dictionary with keys 'env_vars', 'params', and 'custom_params'.

    Returns:
        str: The path to the generated .xlsx file, or an empty string if it could not be determined.
    """
    env_vars = copy.deepcopy(data['env_vars'])
    env_vars['AWX_LOGGING_MODE'] = 'stdout'

    params = [sys.executable]
    params.extend(data['params'])

    # Runs command, in future, we want to support also calling the test function directly due to mocking datetime.now
    if data['custom_params']['run_command'] == 'Yes':
        result = subprocess.run(
            params,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=env_vars,
        )

        if result.returncode != 0:
            logger.error('Generating of report failed')
            logger.error(result.stderr)
            return ''

        text = result.stderr + '/n' + result.stdout
    else:
        return ''

    generated_file = get_file_name(params, env_vars)

    if generated_file:
        return generated_file

    # Regular expression to capture the file path if get_file_name was not able to compute it
    pattern = r'Report generated into directory: (.*\.xlsx)'

    match = re.search(pattern, text)
    if match:
        return match.group(1)

    return None


def run_and_test_snapshot_definitions(directory: str) -> None:
    """
    Finds all JSON definitions in a directory, runs each snapshot definition, and compares the generated report
    with a previously saved original (i.e., regression testing).

    The function looks for an existing 'report.xlsx' in a folder with the same name as the JSON file minus the .json
    extension and compares it to the newly generated report.

    Args:
        directory (str): The directory containing the JSON definition files.

    Returns:
        None
    """
    json_files = find_json_files(directory)
    for json_file in json_files:
        data_dir = json_file.as_posix().removesuffix('.json')
        data: DataShape = parse_json_file(json_file)

        original_file = './' + data_dir + '/report.xlsx'
        generated_file = run_snapshot_definition(data)

        logger.info(f'Compare {original_file} to {generated_file}')
        # compare the generated and original_file

        if data['env_vars']['METRICS_UTILITY_REPORT_TYPE'] == 'CCSPv2':
            compare_ccspv2_reports(original_file, generated_file)

        if data['env_vars']['METRICS_UTILITY_REPORT_TYPE'] == 'CCSP':
            compare_ccsp_reports(original_file, generated_file)


def compare_ccspv2_reports(original_report_path: str, generated_report_path: str) -> None:
    """
    Compares two CCSPv2 Excel reports by opening them as workbooks and comparing relevant sheets.

    Args:
        original_report_path (str): The path to the original CCSPv2 report.
        generated_report_path (str): The path to the newly generated CCSPv2 report.

    Returns:
        None
    """
    logger.debug(f'Opening {generated_report_path}')
    g_wb = openpyxl.load_workbook(filename=generated_report_path)

    logger.debug(f'Opening {original_report_path}')
    o_wb = openpyxl.load_workbook(filename=original_report_path)

    try:
        compare_worksheets(g_wb, o_wb, 0, ['H1', 'B5'])
        compare_worksheets(g_wb, o_wb, 1, [])
        compare_worksheets(g_wb, o_wb, 2, [])
    finally:
        g_wb.close()
        o_wb.close()


def compare_ccsp_reports(original_report_path: str, generated_report_path: str) -> None:
    """
    Compares two CCSP Excel reports by opening them as workbooks and comparing relevant sheets.

    Args:
        original_report_path (str): The path to the original CCSP report.
        generated_report_path (str): The path to the newly generated CCSP report.

    Returns:
        None
    """
    logger.debug(f'Opening {generated_report_path}')
    g_wb = openpyxl.load_workbook(filename=generated_report_path)

    logger.debug(f'Opening {original_report_path}')
    o_wb = openpyxl.load_workbook(filename=original_report_path)

    try:
        compare_worksheets(g_wb, o_wb, 0, ['B5'])
        compare_worksheets(g_wb, o_wb, 1, [])
    finally:
        g_wb.close()
        o_wb.close()


def compare_worksheets(workbook_generated: openpyxl.Workbook, workbook_original: openpyxl.Workbook, sheet_number: int, exceptions: List[str]) -> None:
    """
    Compares worksheets in two workbooks cell by cell, optionally ignoring cells in the `exceptions` list.

    It asserts that the number of rows and columns match, and that the values in corresponding cells match.

    Args:
        workbook_generated (openpyxl.Workbook): The workbook of the newly generated report.
        workbook_original (openpyxl.Workbook): The workbook of the original report.
        sheet_number (int): The worksheet index to compare.
        exceptions (List[str]): A list of cell addresses (e.g., ['A1', 'B2']) that should be ignored during comparison.

    Returns:
        None

    Raises:
        AssertionError: If the number of rows or columns do not match, or if any corresponding cells differ
            (excluding those listed in exceptions).
    """
    worksheet_generated = workbook_generated.worksheets[sheet_number]
    worksheet_original = workbook_original.worksheets[sheet_number]

    max_row_1 = worksheet_original.max_row
    max_column_1 = worksheet_original.max_column

    max_row_2 = worksheet_generated.max_row
    max_column_2 = worksheet_generated.max_column

    assert max_column_1 == max_column_2, (
        f'Number of columns do not match for sheet number: {sheet_number}. Actual value = {max_column_2}, expected value = {max_column_1}'
    )

    assert max_row_1 == max_row_2, (
        f'Number of rows do not match for sheet number: {sheet_number}. Actual value = {max_row_2}, expected value = {max_row_1}'
    )

    for column in range(1, max_column_1 + 1):
        for row in range(1, max_row_1 + 1):
            addr = openpyxl.utils.get_column_letter(column) + str(row)

            if addr not in exceptions:
                val_g = worksheet_generated[addr].value
                val_o = worksheet_original[addr].value

                assert val_g == val_o, (
                    f'Column names do not match for sheet number: {sheet_number}. Address {addr}. Actual value = {val_g}, expected value = {val_o}'
                )


def get_file_name(params: List[str], env_vars: Dict[str, str]) -> str | None:
    """
    Attempts to deduce the filename for the generated report based on report type and parameters.

    For CCSPv2 reports, it looks for a 'month' parameter and constructs "CCSPv2-{month}.xlsx".

    Args:
        params (List[str]): The command-line parameters.
        env_vars (Dict[str, str]): The environment variables that may determine report type.

    Returns:
        str: The deduced filename if found, otherwise None.
    """
    if env_vars['METRICS_UTILITY_REPORT_TYPE'] == 'CCSPv2':
        month = get_param_value(params, 'month')
        if month:
            return f'CCSPv2-{month}.xlsx'
        # since and until has weirdly generated names, use the file name from terminal output of command
    return None


def get_param_value(params: List[str], name: str) -> str | None:
    """
    Extracts the value from a command-line parameter of the form "name=value".

    Args:
        params (List[str]): A list of parameter strings.
        name (str): The name of the parameter to match (before the '=').

    Returns:
        str: The value of the parameter if found, otherwise None.
    """
    for param in params:
        if param.startswith(name + '='):
            key, value = param.split('=')
            if key == name:
                return value
    return None


def get_entry_point_directory() -> str:
    """
    Retrieves the absolute directory path of the current Python entry point.

    Returns:
        str: The directory containing the entry-point file.
    """
    # Get the absolute path of the entry-point file
    entry_point_file = os.path.abspath(sys.argv[0])
    # Get the directory containing the entry-point file
    entry_point_dir = os.path.dirname(entry_point_file)
    return entry_point_dir

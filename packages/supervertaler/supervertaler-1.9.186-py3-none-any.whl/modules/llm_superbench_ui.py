"""
Superbench - Qt UI Components
==============================

PyQt6 user interface for LLM translation benchmarking.

Features:
- Test dataset selection
- Model selection (checkboxes)
- Benchmark execution with progress
- Results table with comparison
- Summary statistics panel
- Export functionality

Author: Michael Beijer
License: MIT
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel,
    QPushButton, QComboBox, QCheckBox, QTableWidget, QTableWidgetItem,
    QProgressBar, QTextEdit, QSplitter, QHeaderView, QMessageBox,
    QFileDialog, QRadioButton, QSpinBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QPointF
from PyQt6.QtGui import QColor, QFont, QPainter, QPen
from PyQt6.QtWidgets import QStyleOptionButton
from typing import List, Optional, Dict
import json
from pathlib import Path

try:
    from modules.llm_leaderboard import (
        LLMLeaderboard, TestDataset, ModelConfig, BenchmarkResult,
        create_sample_datasets, create_dataset_from_project, CHRF_AVAILABLE
    )
except ImportError:
    from llm_leaderboard import (
        LLMLeaderboard, TestDataset, ModelConfig, BenchmarkResult,
        create_sample_datasets, create_dataset_from_project, CHRF_AVAILABLE
    )


class CheckmarkCheckBox(QCheckBox):
    """Custom checkbox with green background and white checkmark when checked"""

    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setCheckable(True)
        self.setEnabled(True)
        self.setStyleSheet("""
            QCheckBox {
                font-size: 9pt;
                spacing: 6px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #999;
                border-radius: 3px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                background-color: #4CAF50;
                border-color: #4CAF50;
            }
            QCheckBox::indicator:hover {
                border-color: #666;
            }
            QCheckBox::indicator:checked:hover {
                background-color: #45a049;
                border-color: #45a049;
            }
        """)

    def paintEvent(self, event):
        """Override paint event to draw white checkmark when checked"""
        super().paintEvent(event)

        if self.isChecked():
            opt = QStyleOptionButton()
            self.initStyleOption(opt)
            indicator_rect = self.style().subElementRect(
                self.style().SubElement.SE_CheckBoxIndicator,
                opt,
                self
            )

            if indicator_rect.isValid():
                # Draw white checkmark
                painter = QPainter(self)
                try:
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    pen_width = max(2.0, min(indicator_rect.width(), indicator_rect.height()) * 0.12)
                    painter.setPen(QPen(QColor(255, 255, 255), pen_width, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
                    painter.setBrush(QColor(255, 255, 255))

                    # Draw checkmark (✓ shape)
                    x = indicator_rect.x()
                    y = indicator_rect.y()
                    w = indicator_rect.width()
                    h = indicator_rect.height()

                    # Add padding
                    padding = min(w, h) * 0.15
                    x += padding
                    y += padding
                    w -= padding * 2
                    h -= padding * 2

                    # Checkmark path
                    check_x1 = x + w * 0.10
                    check_y1 = y + h * 0.50
                    check_x2 = x + w * 0.35
                    check_y2 = y + h * 0.70
                    check_x3 = x + w * 0.90
                    check_y3 = y + h * 0.25

                    # Draw checkmark lines
                    painter.drawLine(QPointF(check_x2, check_y2), QPointF(check_x3, check_y3))
                    painter.drawLine(QPointF(check_x1, check_y1), QPointF(check_x2, check_y2))
                finally:
                    painter.end()


class CustomRadioButton(QRadioButton):
    """Custom radio button with square indicator, green when checked, white checkmark"""

    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setCheckable(True)
        self.setEnabled(True)
        self.setStyleSheet("""
            QRadioButton {
                font-size: 9pt;
                spacing: 6px;
            }
            QRadioButton::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #999;
                border-radius: 3px;
                background-color: white;
            }
            QRadioButton::indicator:checked {
                background-color: #4CAF50;
                border-color: #4CAF50;
            }
            QRadioButton::indicator:hover {
                border-color: #666;
            }
            QRadioButton::indicator:checked:hover {
                background-color: #45a049;
                border-color: #45a049;
            }
        """)

    def paintEvent(self, event):
        """Override paint event to draw white checkmark when checked"""
        super().paintEvent(event)

        if self.isChecked():
            opt = QStyleOptionButton()
            self.initStyleOption(opt)
            indicator_rect = self.style().subElementRect(
                self.style().SubElement.SE_RadioButtonIndicator,
                opt,
                self
            )

            if indicator_rect.isValid():
                # Draw white checkmark
                painter = QPainter(self)
                try:
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    pen_width = max(2.0, min(indicator_rect.width(), indicator_rect.height()) * 0.12)
                    painter.setPen(QPen(QColor(255, 255, 255), pen_width, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
                    painter.setBrush(QColor(255, 255, 255))

                    # Draw checkmark (✓ shape)
                    x = indicator_rect.x()
                    y = indicator_rect.y()
                    w = indicator_rect.width()
                    h = indicator_rect.height()

                    # Add padding
                    padding = min(w, h) * 0.15
                    x += padding
                    y += padding
                    w -= padding * 2
                    h -= padding * 2

                    # Checkmark path
                    check_x1 = x + w * 0.10
                    check_y1 = y + h * 0.50
                    check_x2 = x + w * 0.35
                    check_y2 = y + h * 0.70
                    check_x3 = x + w * 0.90
                    check_y3 = y + h * 0.25

                    # Draw checkmark lines
                    painter.drawLine(QPointF(check_x2, check_y2), QPointF(check_x3, check_y3))
                    painter.drawLine(QPointF(check_x1, check_y1), QPointF(check_x2, check_y2))
                finally:
                    painter.end()


class BenchmarkThread(QThread):
    """Background thread for running benchmarks without blocking UI"""

    progress_update = pyqtSignal(int, int, str)  # current, total, message
    finished = pyqtSignal()  # Completion signal (no data - avoid Qt signal crash with large lists)
    error = pyqtSignal(str)  # error message

    def __init__(self, leaderboard: LLMLeaderboard, dataset: TestDataset, models: List[ModelConfig]):
        super().__init__()
        self.leaderboard = leaderboard
        self.dataset = dataset
        self.models = models
        self.results = []  # Store results here, access from main thread

    def run(self):
        """Run benchmark in background thread"""
        try:
            print(f"[BENCHMARK THREAD] Starting benchmark with {len(self.models)} models on {len(self.dataset.segments)} segments")
            self.results = self.leaderboard.run_benchmark(
                self.dataset,
                self.models,
                progress_callback=self._on_progress
            )
            print(f"[BENCHMARK THREAD] Benchmark completed with {len(self.results)} results")
            # Don't pass results through signal - causes Qt crash with large lists
            # Main thread will access self.results or self.leaderboard.results directly
            self.finished.emit()
            print(f"[BENCHMARK THREAD] Finished signal emitted successfully")
        except Exception as e:
            print(f"[BENCHMARK THREAD] ERROR: {str(e)}")
            import traceback
            print(f"[BENCHMARK THREAD] TRACEBACK:\n{traceback.format_exc()}")
            self.error.emit(str(e))

    def _on_progress(self, current: int, total: int, message: str):
        """Forward progress updates to main thread"""
        try:
            self.progress_update.emit(current, total, message)
        except Exception as e:
            print(f"[BENCHMARK THREAD] Progress update failed: {str(e)}")


class LLMLeaderboardUI(QWidget):
    """Main UI widget for Superbench"""

    def __init__(self, parent=None, llm_client_factory=None):
        super().__init__(parent)
        self.parent_app = parent
        self.llm_client_factory = llm_client_factory
        self.leaderboard = None
        self.benchmark_thread = None
        self.current_results = []

        # Load sample datasets
        self.datasets = create_sample_datasets()
        self.current_dataset = self.datasets[0] if self.datasets else None
        self.project_dataset = None
        self.project_metadata = None

        self.init_ui()

    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout()
        layout.setSpacing(5)  # Tighter spacing for consistency
        layout.setContentsMargins(10, 10, 10, 10)

        # Header (matches TMX Editor / AutoFingers / PDF Rescue style)
        header = QLabel("📊 Superbench")
        header.setStyleSheet("font-size: 16pt; font-weight: bold; color: #1976D2;")
        layout.addWidget(header, 0)  # 0 = no stretch, stays compact

        # Description box (matches TMX Editor / AutoFingers / PDF Rescue style)
        description = QLabel(
            "LLM Translation Quality Benchmarking System - A Supervertaler Module.\n"
            "Compare translation quality, speed, and cost across multiple LLM providers."
        )
        description.setWordWrap(True)
        description.setStyleSheet("color: #666; padding: 5px; background-color: #E3F2FD; border-radius: 3px;")
        layout.addWidget(description, 0)

        # Spacing after description
        layout.addSpacing(10)

        # Top section: Dataset and Model selection
        top_widget = self._create_top_section()
        layout.addWidget(top_widget)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # Status label
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: #666; font-size: 9pt;")
        layout.addWidget(self.status_label)

        # Splitter for results and log
        splitter = QSplitter(Qt.Orientation.Vertical)

        # Results table
        self.results_table = self._create_results_table()
        splitter.addWidget(self.results_table)

        # Summary panel
        self.summary_panel = self._create_summary_panel()
        splitter.addWidget(self.summary_panel)

        # Log output
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setMaximumHeight(150)
        self.log_output.setPlaceholderText("Benchmark log will appear here...")
        splitter.addWidget(self.log_output)

        splitter.setStretchFactor(0, 3)  # Results table gets most space
        splitter.setStretchFactor(1, 1)  # Summary panel medium space
        splitter.setStretchFactor(2, 1)  # Log output smallest

        layout.addWidget(splitter)

        self.setLayout(layout)

    def _create_top_section(self) -> QWidget:
        """Create dataset selection and model selection section"""
        widget = QWidget()
        layout = QHBoxLayout()

        # Left: Dataset selection
        dataset_group = QGroupBox("Test Dataset")
        dataset_layout = QVBoxLayout()

        # Radio buttons for dataset source
        self.predefined_radio = CustomRadioButton("Pre-defined Datasets")
        self.predefined_radio.setChecked(True)
        self.predefined_radio.toggled.connect(self._on_dataset_source_changed)
        dataset_layout.addWidget(self.predefined_radio)

        # Pre-defined datasets dropdown
        self.dataset_combo = QComboBox()
        for ds in self.datasets:
            self.dataset_combo.addItem(f"{ds.name} ({len(ds.segments)} segments)", ds)
        self.dataset_combo.currentIndexChanged.connect(self._on_dataset_changed)
        dataset_layout.addWidget(self.dataset_combo)

        dataset_layout.addSpacing(10)

        # Current Project option
        self.project_radio = CustomRadioButton("Current Project")
        self.project_radio.toggled.connect(self._on_dataset_source_changed)
        dataset_layout.addWidget(self.project_radio)

        # Project dataset controls (initially hidden)
        self.project_controls_widget = QWidget()
        project_controls_layout = QVBoxLayout()
        project_controls_layout.setContentsMargins(20, 0, 0, 0)  # Indent

        # Sample size
        sample_size_layout = QHBoxLayout()
        sample_size_layout.addWidget(QLabel("Sample size:"))
        self.sample_size_spin = QSpinBox()
        self.sample_size_spin.setRange(1, 50)
        self.sample_size_spin.setValue(10)
        self.sample_size_spin.setToolTip("Number of segments to sample from project")
        sample_size_layout.addWidget(self.sample_size_spin)
        sample_size_layout.addStretch()
        project_controls_layout.addLayout(sample_size_layout)

        # Sampling method
        method_layout = QHBoxLayout()
        method_layout.addWidget(QLabel("Method:"))
        self.sampling_method_combo = QComboBox()
        self.sampling_method_combo.addItems(["Smart Sampling", "Random", "Evenly Spaced"])
        self.sampling_method_combo.setToolTip(
            "Smart: 30% begin, 40% middle, 30% end\n"
            "Random: Random selection\n"
            "Evenly Spaced: Every Nth segment"
        )
        method_layout.addWidget(self.sampling_method_combo)
        method_layout.addStretch()
        project_controls_layout.addLayout(method_layout)

        # Project status info
        self.project_status_label = QLabel("Project status: No project loaded")
        self.project_status_label.setStyleSheet("color: #666; font-size: 9pt; padding: 5px;")
        self.project_status_label.setWordWrap(True)
        project_controls_layout.addWidget(self.project_status_label)

        # Create dataset button
        self.create_dataset_button = QPushButton("📊 Create Test Dataset from Project")
        self.create_dataset_button.clicked.connect(self._on_create_project_dataset)
        self.create_dataset_button.setEnabled(False)
        project_controls_layout.addWidget(self.create_dataset_button)

        self.project_controls_widget.setLayout(project_controls_layout)
        self.project_controls_widget.setVisible(False)
        dataset_layout.addWidget(self.project_controls_widget)

        dataset_layout.addStretch()
        dataset_group.setLayout(dataset_layout)
        layout.addWidget(dataset_group)

        # Update project status on init
        self._update_project_status()

        # Right: Model selection
        model_group = QGroupBox("Model Selection")
        model_layout = QVBoxLayout()

        model_layout.addWidget(QLabel("Select models to test:"))

        # OpenAI models
        self.openai_checkbox = CheckmarkCheckBox("OpenAI (GPT-4o)")
        self.openai_checkbox.setChecked(True)
        model_layout.addWidget(self.openai_checkbox)

        self.openai_model_combo = QComboBox()
        self.openai_model_combo.addItems([
            "gpt-4o",
            "gpt-4o-mini",
            "gpt-5"
        ])
        self.openai_model_combo.setEnabled(True)
        model_layout.addWidget(self.openai_model_combo)

        # Claude models
        self.claude_checkbox = CheckmarkCheckBox("Claude (Sonnet 4.5)")
        self.claude_checkbox.setChecked(True)
        model_layout.addWidget(self.claude_checkbox)

        self.claude_model_combo = QComboBox()
        self.claude_model_combo.addItems([
            "claude-sonnet-4-5-20250929",
            "claude-haiku-4-5-20251001",
            "claude-opus-4-1-20250805"
        ])
        self.claude_model_combo.setEnabled(True)
        model_layout.addWidget(self.claude_model_combo)

        # Gemini models
        self.gemini_checkbox = CheckmarkCheckBox("Gemini (2.5 Flash)")
        self.gemini_checkbox.setChecked(True)
        model_layout.addWidget(self.gemini_checkbox)

        self.gemini_model_combo = QComboBox()
        self.gemini_model_combo.addItems([
            "gemini-2.5-flash",
            "gemini-2.5-flash-lite",
            "gemini-2.5-pro"
        ])
        self.gemini_model_combo.setEnabled(True)
        model_layout.addWidget(self.gemini_model_combo)

        model_layout.addStretch()

        # Run button
        self.run_button = QPushButton("🚀 Run Benchmark")
        self.run_button.setStyleSheet("font-weight: bold; padding: 8px;")
        self.run_button.clicked.connect(self._on_run_benchmark)
        model_layout.addWidget(self.run_button)

        # Cancel button
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.setEnabled(False)
        self.cancel_button.clicked.connect(self._on_cancel_benchmark)
        model_layout.addWidget(self.cancel_button)

        # Export button
        self.export_button = QPushButton("📊 Export Results...")
        self.export_button.setEnabled(False)
        self.export_button.clicked.connect(self._on_export_results)
        model_layout.addWidget(self.export_button)

        model_group.setLayout(model_layout)
        layout.addWidget(model_group)

        widget.setLayout(layout)
        return widget

    def _create_results_table(self) -> QTableWidget:
        """Create results comparison table"""
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "Segment", "Source Text", "Model", "Translation", "Speed (ms)", "Quality"
        ])

        # Set column widths
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)

        table.setAlternatingRowColors(True)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        return table

    def _create_summary_panel(self) -> QWidget:
        """Create summary statistics panel"""
        widget = QGroupBox("Summary Statistics")
        layout = QVBoxLayout()

        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(5)
        self.summary_table.setHorizontalHeaderLabels([
            "Model", "Avg Speed (ms)", "Avg Quality", "Success", "Errors"
        ])

        header = self.summary_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.summary_table.setMaximumHeight(200)
        self.summary_table.setAlternatingRowColors(True)
        self.summary_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        layout.addWidget(self.summary_table)
        widget.setLayout(layout)

        return widget

    def _on_dataset_changed(self, index: int):
        """Handle dataset selection change"""
        self.current_dataset = self.dataset_combo.itemData(index)
        self.log(f"Selected dataset: {self.current_dataset.name}")

    def _on_dataset_source_changed(self):
        """Handle radio button toggle between predefined and project datasets"""
        print(f"[LLM DEBUG] _on_dataset_source_changed() called")
        is_project = self.project_radio.isChecked()
        print(f"[LLM DEBUG] is_project: {is_project}")
        self.dataset_combo.setEnabled(not is_project)
        self.project_controls_widget.setVisible(is_project)

        if is_project:
            # Update project status when switching to project mode
            print(f"[LLM DEBUG] Calling _update_project_status() from _on_dataset_source_changed")
            self._update_project_status()

            # Switch to project dataset if available
            if self.project_dataset:
                self.current_dataset = self.project_dataset
                self.log(f"Using project dataset: {self.current_dataset.name}")
            else:
                self.current_dataset = None
        else:
            # Switch back to predefined dataset
            self.current_dataset = self.dataset_combo.currentData()
            if self.current_dataset:
                self.log(f"Using predefined dataset: {self.current_dataset.name}")

    def _update_project_status(self):
        """Update the project status label based on loaded project"""
        try:
            # Debug to console (always print)
            print(f"[LLM DEBUG] _update_project_status() called")
            print(f"[LLM DEBUG] parent_app exists: {self.parent_app is not None}")
        except Exception as e:
            print(f"[LLM DEBUG] ERROR in _update_project_status: {e}")
            import traceback
            traceback.print_exc()

        if self.parent_app:
            print(f"[LLM DEBUG] parent_app has 'current_project': {hasattr(self.parent_app, 'current_project')}")
            if hasattr(self.parent_app, 'current_project'):
                print(f"[LLM DEBUG] current_project is not None: {self.parent_app.current_project is not None}")
                if self.parent_app.current_project:
                    print(f"[LLM DEBUG] project.segments exists: {hasattr(self.parent_app.current_project, 'segments')}")
                    if hasattr(self.parent_app.current_project, 'segments'):
                        print(f"[LLM DEBUG] project.segments length: {len(self.parent_app.current_project.segments) if self.parent_app.current_project.segments else 0}")

        if not self.parent_app or not hasattr(self.parent_app, 'current_project') or not self.parent_app.current_project:
            self.project_status_label.setText("⚠️ No project loaded")
            self.project_status_label.setStyleSheet("color: #FF6600; font-size: 9pt; padding: 5px;")
            self.create_dataset_button.setEnabled(False)
            return

        project = self.parent_app.current_project
        total_segs = len(project.segments) if project.segments else 0
        print(f"[LLM DEBUG] Project has {total_segs} segments")

        if total_segs == 0:
            self.project_status_label.setText("⚠️ Project has no segments")
            self.project_status_label.setStyleSheet("color: #FF6600; font-size: 9pt; padding: 5px;")
            self.create_dataset_button.setEnabled(False)
            return

        # Count translated segments
        translated = sum(1 for seg in project.segments if seg.target and seg.target.strip())
        pct = (translated / total_segs * 100) if total_segs > 0 else 0

        status_html = f"""
        <b>Project Status:</b><br>
        • Total segments: {total_segs}<br>
        • Translated: {translated} ({pct:.1f}%)<br>
        """

        if translated == 0:
            status_html += "<br>⚠️ <b>No translations yet</b><br>"
            status_html += "Quality scoring unavailable<br>"
            status_html += "Will compare: Speed, Cost, Outputs"
            self.project_status_label.setStyleSheet("color: #FF6600; font-size: 9pt; padding: 5px; background: #FFF8E1; border-radius: 3px;")
        elif translated < total_segs:
            status_html += f"<br>✓ Quality scoring available for {translated} segments"
            self.project_status_label.setStyleSheet("color: #0066CC; font-size: 9pt; padding: 5px; background: #E3F2FD; border-radius: 3px;")
        else:
            status_html += "<br>✓ Quality scoring available (fully translated)"
            self.project_status_label.setStyleSheet("color: #00AA00; font-size: 9pt; padding: 5px; background: #E8F5E9; border-radius: 3px;")

        self.project_status_label.setText(status_html)
        self.create_dataset_button.setEnabled(True)

    def _on_create_project_dataset(self):
        """Create test dataset from current project"""
        if not self.parent_app or not hasattr(self.parent_app, 'current_project') or not self.parent_app.current_project:
            QMessageBox.warning(self, "Error", "No project loaded")
            return

        project = self.parent_app.current_project
        sample_size = self.sample_size_spin.value()

        # Map combo box text to method name
        method_map = {
            "Smart Sampling": "smart",
            "Random": "random",
            "Evenly Spaced": "evenly_spaced"
        }
        sampling_method = method_map.get(self.sampling_method_combo.currentText(), "smart")

        try:
            # Create dataset
            self.project_dataset, self.project_metadata = create_dataset_from_project(
                project,
                sample_size=sample_size,
                sampling_method=sampling_method,
                require_targets=False
            )

            self.current_dataset = self.project_dataset

            # Log creation
            meta = self.project_metadata
            self.log(f"Created project dataset: {self.project_dataset.name}")
            self.log(f"  • Sampled {meta['sampled_segments']} segments from {meta['total_segments']} total")
            self.log(f"  • Method: {sampling_method}")
            self.log(f"  • References available: {meta['segments_with_references']}/{meta['sampled_segments']}")

            if meta['quality_scoring_available']:
                self.log(f"  • ✓ Quality scoring enabled")
            else:
                self.log(f"  • ⚠️ Quality scoring disabled (no reference translations)")

            # Update button text
            self.create_dataset_button.setText(f"✓ Dataset Created ({len(self.project_dataset.segments)} segments)")
            self.create_dataset_button.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")

            QMessageBox.information(
                self,
                "Dataset Created",
                f"Successfully created test dataset with {meta['sampled_segments']} segments.\n\n"
                f"Quality scoring: {'Enabled' if meta['quality_scoring_available'] else 'Disabled (no references)'}\n"
                f"Ready to benchmark!"
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create dataset:\n{str(e)}")
            self.log(f"ERROR creating project dataset: {str(e)}")

    def _on_run_benchmark(self):
        """Start benchmark execution"""
        if not self.llm_client_factory:
            QMessageBox.warning(self, "Error", "LLM client factory not available")
            return

        if not self.current_dataset:
            QMessageBox.warning(self, "Error", "No dataset selected")
            return

        # Get selected models
        models = self._get_selected_models()
        if not models:
            QMessageBox.warning(self, "Error", "Please select at least one model to test")
            return

        # Confirm if sacrebleu not available
        if not CHRF_AVAILABLE:
            reply = QMessageBox.question(
                self,
                "Quality Scoring Unavailable",
                "sacrebleu library is not installed. Quality scores will not be calculated.\n\n"
                "Continue anyway?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # Clear previous results
        self.results_table.setRowCount(0)
        self.summary_table.setRowCount(0)
        self.log_output.clear()
        self.current_results = []

        # Update UI state
        self.run_button.setEnabled(False)
        self.cancel_button.setEnabled(True)
        self.export_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Running benchmark...")

        # Create leaderboard instance
        self.leaderboard = LLMLeaderboard(self.llm_client_factory, self.log)

        # Start benchmark in background thread
        self.benchmark_thread = BenchmarkThread(self.leaderboard, self.current_dataset, models)
        self.benchmark_thread.progress_update.connect(self._on_progress_update)
        self.benchmark_thread.finished.connect(self._on_benchmark_finished)
        self.benchmark_thread.error.connect(self._on_benchmark_error)
        self.benchmark_thread.start()

    def _on_cancel_benchmark(self):
        """Cancel running benchmark"""
        if self.leaderboard:
            self.leaderboard.cancel_benchmark()
            self.log("⚠️ Cancelling benchmark...")

    def _on_progress_update(self, current: int, total: int, message: str):
        """Update progress bar and status"""
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.status_label.setText(f"{message} ({current}/{total})")

    def _on_benchmark_finished(self):
        """Handle benchmark completion"""
        try:
            # Get results from benchmark thread (stored there, not passed through signal)
            if self.benchmark_thread and hasattr(self.benchmark_thread, 'results'):
                results = self.benchmark_thread.results[:]
            elif self.leaderboard and hasattr(self.leaderboard, 'results'):
                results = self.leaderboard.results[:]
            else:
                results = []
            
            print(f"[UI] _on_benchmark_finished called, retrieved {len(results)} results")
            
            # Validate results
            if not results:
                print(f"[UI] WARNING: No results to display")
                self.run_button.setEnabled(True)
                self.cancel_button.setEnabled(False)
                self.progress_bar.setVisible(False)
                self.status_label.setText("⚠️ Benchmark complete but no results")
                self.log("⚠️ Benchmark complete but produced no results")
                return
            
            # Copy results to avoid threading issues
            self.current_results = results

            # Update UI state
            print(f"[UI] Updating UI state...")
            self.run_button.setEnabled(True)
            self.cancel_button.setEnabled(False)
            self.export_button.setEnabled(True)
            self.progress_bar.setVisible(False)
            self.status_label.setText(f"✅ Benchmark complete: {len(results)} results")
            print(f"[UI] UI state updated")

            # Block table signals during population to prevent crashes
            print(f"[UI] Populating results table...")
            self.results_table.blockSignals(True)
            try:
                self._populate_results_table(results)
                print(f"[UI] Results table populated")
            finally:
                self.results_table.blockSignals(False)

            # Populate summary table
            print(f"[UI] Populating summary table...")
            self.summary_table.blockSignals(True)
            try:
                self._populate_summary_table()
                print(f"[UI] Summary table populated")
            finally:
                self.summary_table.blockSignals(False)

            self.log("✅ Benchmark finished successfully")
            print(f"[UI] _on_benchmark_finished completed successfully")
        except Exception as e:
            print(f"[UI] ERROR in _on_benchmark_finished: {str(e)}")
            import traceback
            print(f"[UI] TRACEBACK:\n{traceback.format_exc()}")
            self.log(f"❌ Error displaying results: {str(e)}")
            # Re-enable buttons even on error
            self.run_button.setEnabled(True)
            self.cancel_button.setEnabled(False)
            self.progress_bar.setVisible(False)
            self.status_label.setText("❌ Display error")
            QMessageBox.critical(self, "Display Error", f"Benchmark completed but failed to display results:\n\n{str(e)}")

    def _on_benchmark_error(self, error_msg: str):
        """Handle benchmark error"""
        self.run_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.status_label.setText("❌ Benchmark failed")

        QMessageBox.critical(self, "Benchmark Error", f"An error occurred:\n\n{error_msg}")
        self.log(f"❌ Error: {error_msg}")

    def _populate_results_table(self, results: List[BenchmarkResult]):
        """Populate results table with benchmark data"""
        try:
            # Validate inputs
            if not results:
                print("[UI] _populate_results_table: No results to populate")
                return
            
            if not self.current_dataset or not hasattr(self.current_dataset, 'segments'):
                print("[UI] _populate_results_table: No current dataset")
                return
            
            # Clear existing rows
            self.results_table.setRowCount(0)
            
            # Group results by segment
            segments_dict = {}
            for result in results:
                if not result or not hasattr(result, 'segment_id'):
                    continue
                if result.segment_id not in segments_dict:
                    segments_dict[result.segment_id] = []
                segments_dict[result.segment_id].append(result)

            # Populate table
            row = 0
            for segment_id in sorted(segments_dict.keys()):
                segment_results = segments_dict[segment_id]

                # Get source text from dataset
                source_text = "(source not found)"
                if self.current_dataset and self.current_dataset.segments:
                    for seg in self.current_dataset.segments:
                        if hasattr(seg, 'id') and seg.id == segment_id:
                            source_text = seg.source if hasattr(seg, 'source') else "(no source)"
                            break

                # Truncate source text for display
                if source_text and len(source_text) > 80:
                    source_text = source_text[:77] + "..."

                for result in segment_results:
                    try:
                        self.results_table.insertRow(row)

                        # Segment ID
                        self.results_table.setItem(row, 0, QTableWidgetItem(str(segment_id)))

                        # Source text
                        self.results_table.setItem(row, 1, QTableWidgetItem(source_text))

                        # Model name
                        model_name = result.model_name if hasattr(result, 'model_name') else "Unknown"
                        self.results_table.setItem(row, 2, QTableWidgetItem(model_name))

                        # Translation output
                        output_text = result.output if (hasattr(result, 'output') and result.output) else f"ERROR: {getattr(result, 'error', 'Unknown error')}"
                        if len(output_text) > 100:
                            output_text = output_text[:97] + "..."
                        item = QTableWidgetItem(output_text)
                        if hasattr(result, 'error') and result.error:
                            item.setForeground(QColor("red"))
                        self.results_table.setItem(row, 3, item)

                        # Speed
                        latency = result.latency_ms if hasattr(result, 'latency_ms') else 0.0
                        speed_item = QTableWidgetItem(f"{latency:.0f}")
                        self.results_table.setItem(row, 4, speed_item)

                        # Quality
                        if hasattr(result, 'quality_score') and result.quality_score is not None:
                            quality_item = QTableWidgetItem(f"{result.quality_score:.1f}")
                            self.results_table.setItem(row, 5, quality_item)
                        else:
                            self.results_table.setItem(row, 5, QTableWidgetItem("—"))

                        row += 1
                    except Exception as row_error:
                        print(f"[UI] Error populating row {row}: {row_error}")
                        continue
                        
        except Exception as e:
            print(f"[UI] ERROR in _populate_results_table: {str(e)}")
            import traceback
            print(f"[UI] TRACEBACK:\n{traceback.format_exc()}")
            raise

    def _populate_summary_table(self):
        """Populate summary statistics table"""
        try:
            if not self.leaderboard:
                print("[UI] _populate_summary_table: No leaderboard instance")
                return

            summary = self.leaderboard.get_summary_stats()
            
            if not summary:
                print("[UI] _populate_summary_table: No summary stats available")
                return

            self.summary_table.setRowCount(len(summary))
            row = 0

            for model_name, stats in summary.items():
                try:
                    # Model name
                    self.summary_table.setItem(row, 0, QTableWidgetItem(str(model_name)))

                    # Avg speed
                    avg_speed = stats.get("avg_latency_ms", 0.0)
                    speed_item = QTableWidgetItem(f"{avg_speed:.0f}")
                    self.summary_table.setItem(row, 1, speed_item)

                    # Avg quality
                    avg_quality = stats.get("avg_quality_score")
                    if avg_quality is not None:
                        quality_item = QTableWidgetItem(f"{avg_quality:.1f}")
                        self.summary_table.setItem(row, 2, quality_item)
                    else:
                        self.summary_table.setItem(row, 2, QTableWidgetItem("—"))

                    # Success count
                    success_count = stats.get("success_count", 0)
                    self.summary_table.setItem(row, 3, QTableWidgetItem(str(success_count)))

                    # Error count
                    error_count = stats.get("error_count", 0)
                    error_item = QTableWidgetItem(str(error_count))
                    if error_count > 0:
                        error_item.setForeground(QColor("red"))
                    self.summary_table.setItem(row, 4, error_item)

                    row += 1
                except Exception as row_error:
                    print(f"[UI] Error populating summary row {row}: {row_error}")
                    continue
                    
        except Exception as e:
            print(f"[UI] ERROR in _populate_summary_table: {str(e)}")
            import traceback
            print(f"[UI] TRACEBACK:\n{traceback.format_exc()}")
            raise

    def _on_export_results(self):
        """Export results to file (JSON or Excel)"""
        if not self.current_results:
            QMessageBox.warning(self, "No Results", "No benchmark results to export")
            return

        # Generate filename with dataset info - sanitize for Windows filesystem
        dataset_name = self.current_dataset.name.replace(" ", "_").replace("→", "-")
        # Remove invalid filename characters: < > : " / \ | ? *
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            dataset_name = dataset_name.replace(char, "_")
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"LLM_Leaderboard_{dataset_name}_{timestamp}.xlsx"

        # Ask user for file path and format
        filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Benchmark Results",
            default_filename,
            "Excel Files (*.xlsx);;JSON Files (*.json);;All Files (*)"
        )

        if not filepath:
            return

        try:
            # Determine export format from selected filter or file extension
            if "Excel" in selected_filter or filepath.endswith('.xlsx'):
                self._export_to_excel(filepath)
            else:
                self._export_to_json(filepath)

            QMessageBox.information(self, "Export Complete", f"Results exported to:\n{filepath}")
            self.log(f"OK Results exported to {filepath}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export results:\n{str(e)}")
            self.log(f"ERROR Export error: {e}")

    def _export_to_json(self, filepath: str):
        """Export results to JSON file"""
        export_data = self.leaderboard.export_to_dict()
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

    def _export_to_excel(self, filepath: str):
        """Export results to Excel file with title sheet, detailed results, and summary"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        wb = Workbook()

        # === TITLE/INFO SHEET ===
        ws_info = wb.active
        ws_info.title = "About"

        # Title with emoji (matches UI header style)
        ws_info['A1'] = "📊 Superbench"
        ws_info['A1'].font = Font(size=24, bold=True, color="1976D2")  # Blue color matching UI
        ws_info.merge_cells('A1:D1')

        # Subtitle (matches UI description style)
        ws_info['A2'] = "Translation Quality Benchmarking System"
        ws_info['A2'].font = Font(size=12, italic=True, color="666666")
        ws_info.merge_cells('A2:D2')

        # Supervertaler module branding (matches standardized naming)
        ws_info['A3'] = "A Supervertaler Module"
        ws_info['A3'].font = Font(size=11, color="0066CC", underline="single")
        ws_info['A3'].hyperlink = "https://supervertaler.com/"
        ws_info.merge_cells('A3:D3')

        # Spacing
        ws_info.row_dimensions[4].height = 20

        # Benchmark Info
        info_header_font = Font(bold=True, size=11)
        info_label_font = Font(size=10, color="666666")
        info_value_font = Font(size=10)

        row = 5
        ws_info[f'A{row}'] = "BENCHMARK INFORMATION"
        ws_info[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        row += 1

        # Dataset info
        ws_info[f'A{row}'] = "Test Dataset:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = self.current_dataset.name
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        ws_info[f'A{row}'] = "Description:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = self.current_dataset.description
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        ws_info[f'A{row}'] = "Segments Tested:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = len(self.current_dataset.segments)
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        ws_info[f'A{row}'] = "Date & Time:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        # Models tested
        ws_info[f'A{row}'] = "Models Tested:"
        ws_info[f'A{row}'].font = info_label_font
        models_tested = set(r.model_name for r in self.current_results)
        ws_info[f'B{row}'] = ", ".join(models_tested)
        ws_info[f'B{row}'].font = info_value_font
        row += 2

        # Explanation section
        ws_info[f'A{row}'] = "UNDERSTANDING THE RESULTS"
        ws_info[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        row += 1

        explanations = [
            ("Quality Score (chrF++):", "Character-level metric measuring translation accuracy. Higher is better (0-100). Scores above 80 indicate excellent quality."),
            ("Speed (ms):", "Translation time in milliseconds. Lower is better. Typical range: 1000-5000ms per segment."),
            ("Success Count:", "Number of segments successfully translated without errors."),
            ("Error Count:", "Number of failed translations. Should be 0 for production use."),
            ("Color Coding:", "Green highlighting indicates the best performer in each category (quality/speed).")
        ]

        for label, explanation in explanations:
            ws_info[f'A{row}'] = label
            ws_info[f'A{row}'].font = Font(bold=True, size=10)
            ws_info[f'B{row}'] = explanation
            ws_info[f'B{row}'].font = Font(size=10)
            ws_info[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        row += 1
        ws_info[f'A{row}'] = "NAVIGATION"
        ws_info[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        row += 1

        navigation_items = [
            ("Summary Tab:", "View aggregated statistics and compare models side-by-side."),
            ("Results Tab:", "View detailed translation output for each segment and model."),
        ]

        for label, description in navigation_items:
            ws_info[f'A{row}'] = label
            ws_info[f'A{row}'].font = Font(bold=True, size=10)
            ws_info[f'B{row}'] = description
            ws_info[f'B{row}'].font = Font(size=10)
            row += 1

        # Column widths
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 80
        ws_info.column_dimensions['C'].width = 15
        ws_info.column_dimensions['D'].width = 15

        # === SUMMARY SHEET ===
        ws_summary = wb.create_sheet("Summary")

        # Header row
        summary_headers = ["Model", "Provider", "Model ID", "Avg Speed (ms)", "Avg Quality (chrF++)",
                          "Success Count", "Error Count", "Total Tests"]
        ws_summary.append(summary_headers)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num, _ in enumerate(summary_headers, 1):
            cell = ws_summary.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Get summary statistics
        summary_stats = self.leaderboard.get_summary_stats()

        # Populate summary data
        row_num = 2
        for model_name, stats in summary_stats.items():
            # Extract provider and model ID from results
            provider = ""
            model_id = ""
            for result in self.current_results:
                if result.model_name == model_name:
                    provider = result.provider
                    model_id = result.model_id
                    break

            total_tests = stats["success_count"] + stats["error_count"]

            ws_summary.cell(row_num, 1, model_name)
            ws_summary.cell(row_num, 2, provider)
            ws_summary.cell(row_num, 3, model_id)
            ws_summary.cell(row_num, 4, f"{stats['avg_latency_ms']:.0f}" if stats['avg_latency_ms'] else "")
            ws_summary.cell(row_num, 5, f"{stats['avg_quality_score']:.2f}" if stats['avg_quality_score'] else "")
            ws_summary.cell(row_num, 6, stats["success_count"])
            ws_summary.cell(row_num, 7, stats["error_count"])
            ws_summary.cell(row_num, 8, total_tests)

            # Highlight best quality score
            if stats['avg_quality_score']:
                quality_cell = ws_summary.cell(row_num, 5)
                best_quality = max(s['avg_quality_score'] for s in summary_stats.values() if s['avg_quality_score'])
                if stats['avg_quality_score'] == best_quality:
                    quality_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    quality_cell.font = Font(bold=True)

            # Highlight best speed
            if stats['avg_latency_ms']:
                speed_cell = ws_summary.cell(row_num, 4)
                best_speed = min(s['avg_latency_ms'] for s in summary_stats.values() if s['avg_latency_ms'])
                if stats['avg_latency_ms'] == best_speed:
                    speed_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    speed_cell.font = Font(bold=True)

            row_num += 1

        # Auto-size summary columns
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 30
        ws_summary.column_dimensions['D'].width = 18
        ws_summary.column_dimensions['E'].width = 20
        ws_summary.column_dimensions['F'].width = 15
        ws_summary.column_dimensions['G'].width = 15
        ws_summary.column_dimensions['H'].width = 15

        # === RESULTS SHEET (New Segment-Grouped Format) ===
        ws_results = wb.create_sheet("Results")

        # Group results by segment
        segments_dict = {}
        for result in self.current_results:
            if result.segment_id not in segments_dict:
                segments_dict[result.segment_id] = []
            segments_dict[result.segment_id].append(result)

        # Get list of all models tested
        all_models = []
        model_seen = set()
        for result in self.current_results:
            if result.model_name not in model_seen:
                all_models.append(result.model_name)
                model_seen.add(result.model_name)

        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        segment_header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        segment_header_font = Font(bold=True, size=11)
        label_font = Font(bold=True, size=10)
        best_quality_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Model-specific background colors (alternating for visual clarity)
        model_colors = {
            "GPT-4o": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # Stronger pink/salmon
            "Claude Sonnet 4.5": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),  # Stronger light green
            "Gemini 2.5 Flash": PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid"),  # Stronger light blue
        }
        # Default colors for other models
        default_model_colors = [
            PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),  # Light orange
            PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # Light purple
            PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),  # Light green-2
        ]

        row_num = 1

        # Process each segment
        for segment_id in sorted(segments_dict.keys()):
            segment_results = segments_dict[segment_id]

            # Get source and reference text from dataset
            source_text = ""
            reference_text = ""
            for seg in self.current_dataset.segments:
                if seg.id == segment_id:
                    source_text = seg.source
                    reference_text = seg.reference
                    break

            # Segment header row (spans columns A-B)
            ws_results.merge_cells(f'A{row_num}:B{row_num}')
            segment_header_cell = ws_results.cell(row_num, 1, f"Segment {segment_id}")
            segment_header_cell.fill = segment_header_fill
            segment_header_cell.font = segment_header_font
            segment_header_cell.alignment = Alignment(horizontal="left", vertical="center")
            row_num += 1

            # Source row
            ws_results.cell(row_num, 1, "Source:").font = label_font
            ws_results.cell(row_num, 2, source_text)
            ws_results.cell(row_num, 2).alignment = Alignment(wrap_text=True, vertical="top")
            row_num += 1

            # Reference row (if available)
            if reference_text:
                ws_results.cell(row_num, 1, "Reference:").font = label_font
                ws_results.cell(row_num, 2, reference_text)
                ws_results.cell(row_num, 2).alignment = Alignment(wrap_text=True, vertical="top")
                row_num += 1

            # Find best quality score for this segment (if available)
            best_quality = None
            if reference_text:  # Only if we have references
                quality_scores = [r.quality_score for r in segment_results if r.quality_score is not None]
                if quality_scores:
                    best_quality = max(quality_scores)

            # Model output rows
            for idx, result in enumerate(segment_results):
                # Model name label
                model_cell = ws_results.cell(row_num, 1, result.model_name)
                model_cell.font = label_font

                # Translation output
                output_cell = ws_results.cell(row_num, 2, result.output if result.output else result.error if result.error else "")
                output_cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Apply model-specific background color first
                if result.model_name in model_colors:
                    model_cell.fill = model_colors[result.model_name]
                    output_cell.fill = model_colors[result.model_name]
                else:
                    # Use alternating default colors for unknown models
                    color_idx = idx % len(default_model_colors)
                    model_cell.fill = default_model_colors[color_idx]
                    output_cell.fill = default_model_colors[color_idx]

                # Override with best quality highlight (green wins over model color)
                if best_quality and result.quality_score == best_quality:
                    model_cell.fill = best_quality_fill
                    output_cell.fill = best_quality_fill

                # Override with error highlight (red wins over everything)
                if result.error:
                    error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    model_cell.fill = error_fill
                    output_cell.fill = error_fill

                row_num += 1

            # Add blank row between segments
            row_num += 1

        # Set column widths
        ws_results.column_dimensions['A'].width = 20  # Model name column
        ws_results.column_dimensions['B'].width = 80  # Text column (wider for readability)

        # Save workbook
        wb.save(filepath)

    def _get_selected_models(self) -> List[ModelConfig]:
        """Get list of selected models from UI"""
        models = []

        # Map model IDs to friendly display names
        model_names = {
            # OpenAI
            "gpt-4o": "GPT-4o",
            "gpt-4o-mini": "GPT-4o Mini",
            "gpt-5": "GPT-5 (Reasoning)",

            # Claude
            "claude-sonnet-4-5-20250929": "Claude Sonnet 4.5",
            "claude-haiku-4-5-20251001": "Claude Haiku 4.5",
            "claude-opus-4-1-20250805": "Claude Opus 4.1",

            # Gemini
            "gemini-2.5-flash": "Gemini 2.5 Flash",
            "gemini-2.5-flash-lite": "Gemini 2.5 Flash Lite",
            "gemini-2.5-pro": "Gemini 2.5 Pro",
            "gemini-2.0-flash-exp": "Gemini 2.0 Flash (Exp)"
        }

        if self.openai_checkbox.isChecked():
            model_id = self.openai_model_combo.currentText()
            models.append(ModelConfig(
                name=model_names.get(model_id, model_id),
                provider="openai",
                model_id=model_id
            ))

        if self.claude_checkbox.isChecked():
            model_id = self.claude_model_combo.currentText()
            models.append(ModelConfig(
                name=model_names.get(model_id, model_id),
                provider="claude",
                model_id=model_id
            ))

        if self.gemini_checkbox.isChecked():
            model_id = self.gemini_model_combo.currentText()
            models.append(ModelConfig(
                name=model_names.get(model_id, model_id),
                provider="gemini",
                model_id=model_id
            ))

        return models

    def log(self, message: str):
        """Append message to log output and auto-scroll to bottom"""
        self.log_output.append(message)
        # Auto-scroll to bottom to show latest messages
        scrollbar = self.log_output.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())


# For standalone testing
if __name__ == "__main__":
    import sys
    from PyQt6.QtWidgets import QApplication

    app = QApplication(sys.argv)

    # Mock LLM client factory for testing UI
    def mock_llm_factory(provider, model):
        print(f"Mock: Creating {provider} client with model {model}")
        return None

    window = LLMLeaderboardUI(llm_client_factory=mock_llm_factory)
    window.setWindowTitle("Superbench")
    window.resize(1200, 800)
    window.show()

    sys.exit(app.exec())
